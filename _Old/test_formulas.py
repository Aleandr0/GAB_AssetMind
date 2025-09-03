#!/usr/bin/env python3
"""
Test per verificare la creazione delle formule Excel
"""

from models import PortfolioManager
import pandas as pd

def test_formulas():
    print("=== TEST FORMULE EXCEL ===")
    
    # Inizializza portfolio manager
    pm = PortfolioManager()
    
    # Carica dati attuali
    df = pm.load_data()
    print(f"Asset caricati: {len(df)}")
    print(f"Colonne: {list(df.columns)}")
    
    if len(df) > 0:
        print("Testando salvataggio con formule...")
        
        # Forza il salvataggio con formule
        result = pm.save_data(df)
        print(f"Salvataggio riuscito: {result}")
        
        # Verifica aprendo di nuovo il file
        print("\nVerificando il file Excel...")
        from openpyxl import load_workbook
        
        try:
            wb = load_workbook(pm.excel_file)
            ws = wb.active
            
            print(f"Righe nel file: {ws.max_row}")
            print(f"Colonne nel file: {ws.max_column}")
            
            # Controlla alcune celle per vedere se contengono formule
            for row in range(2, min(4, ws.max_row + 1)):  # Controlla prime 2 righe dati
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if str(cell.value).startswith('='):
                        print(f"FORMULA trovata in {cell.coordinate}: {cell.value}")
                        
        except Exception as e:
            print(f"Errore nel controllo file Excel: {e}")
    else:
        print("Nessun asset trovato per il test")

if __name__ == "__main__":
    test_formulas()