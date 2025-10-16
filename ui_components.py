#!/usr/bin/env python3
"""
Componenti UI separati per GAB AssetMind
Divide la classe monolitica in componenti specializzati e riutilizzabili
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from typing import Optional, Dict, Any, List, Callable, Set
from datetime import datetime

from config import UIConfig, FieldMapping, AssetConfig, Messages
from utils import DateFormatter, CurrencyFormatter, DataValidator, ErrorHandler, safe_execute
from models import Asset, PortfolioManager, MANUAL_UPDATE_NOTE
from logging_config import get_logger
from ui_performance import UIUpdateManager, LazyColumnResizer, UIRefreshOptimizer

class BaseUIComponent:
    """Classe base per tutti i componenti UI"""
    
    def __init__(self, parent, portfolio_manager: PortfolioManager):
        self.parent = parent
        self.portfolio_manager = portfolio_manager
        self.callbacks: Dict[str, Callable] = {}
        self.logger = get_logger(self.__class__.__name__)
    
    def register_callback(self, event_name: str, callback: Callable):
        """Registra un callback per un evento"""
        self.callbacks[event_name] = callback
    
    def trigger_callback(self, event_name: str, *args, **kwargs):
        """Esegue un callback se registrato"""
        if event_name in self.callbacks:
            safe_execute(
                lambda: self.callbacks[event_name](*args, **kwargs),
                error_handler=lambda e: self.logger.error(f"Errore callback {event_name}: {e}")
            )

class NavigationBar(BaseUIComponent):
    """Barra di navigazione principale dell'applicazione"""
    
    def __init__(self, parent, portfolio_manager: PortfolioManager):
        super().__init__(parent, portfolio_manager)
        self.navbar_frame = None
        self.total_value_label = None
        self.selected_value_label = None
        self.counts_label = None
        self.portfolio_selector = None
        self.current_portfolio_file = "portfolio_data.xlsx"
        self.nav_buttons = {}  # Dizionario per tracciare i bottoni di navigazione
        self.current_page = "RoadMap"  # Pagina attiva corrente
        
    def create_navbar(self) -> ctk.CTkFrame:
        """Crea la barra di navigazione completa"""
        self.navbar_frame = ctk.CTkFrame(self.parent, height=80)
        self.navbar_frame.pack(fill="x", padx=10, pady=5)
        self.navbar_frame.pack_propagate(False)
        
        self._create_values_section()
        self._create_portfolio_section()
        self._create_navigation_buttons()
        
        return self.navbar_frame
    
    def _create_values_section(self):
        """Crea la sezione valori (sinistra)"""
        values_frame = ctk.CTkFrame(self.navbar_frame, fg_color="transparent")
        values_frame.pack(side="left", padx=20, pady=7)
        
        # Valore totale
        self.total_value_label = ctk.CTkLabel(
            values_frame, 
            text="Valore Totale: €0",
            font=ctk.CTkFont(**UIConfig.FONTS['header']),
            text_color=(UIConfig.COLORS['primary'], "#14375e")
        )
        self.total_value_label.pack(pady=(3, 0))
        
        # Valore selezionato
        self.selected_value_label = ctk.CTkLabel(
            values_frame,
            text="Valore selezionato: €0",
            font=ctk.CTkFont(**UIConfig.FONTS['text']),
            text_color=(UIConfig.COLORS['primary'], "#14375e")
        )
        self.selected_value_label.pack(pady=(0, 3))

        # Contatori (Record Totali / Asset Correnti)
        self.counts_label = ctk.CTkLabel(
            values_frame,
            text="Record Totali: 0 — Asset Correnti: 0",
            font=ctk.CTkFont(**UIConfig.FONTS['small']),
            text_color=(UIConfig.COLORS['primary'], "#14375e")
        )
        self.counts_label.pack(pady=(0, 3))

    def _create_portfolio_section(self):
        """Crea la sezione selezione portfolio (centro)"""
        portfolio_section = ctk.CTkFrame(self.navbar_frame, fg_color="transparent")
        portfolio_section.pack(side="left", padx=20, pady=7)
        
        # Label portfolio
        portfolio_label = ctk.CTkLabel(
            portfolio_section, 
            text="Portfolio:",
            font=ctk.CTkFont(**UIConfig.FONTS['subheader']),
            text_color=(UIConfig.COLORS['primary'], "#14375e")
        )
        portfolio_label.pack(side="top", pady=(0, 2))
        
        # Container per dropdown e pulsante nuovo
        selector_frame = ctk.CTkFrame(portfolio_section, fg_color="transparent")
        selector_frame.pack(side="top")
        
        # Dropdown selezione portfolio
        self.portfolio_selector = ctk.CTkComboBox(
            selector_frame,
            command=self._on_portfolio_changed,
            width=140, 
            height=28,
            font=ctk.CTkFont(**UIConfig.FONTS['small'])
        )
        self.portfolio_selector.pack(side="left", padx=(0, 5))
        
        # Pulsante per creare nuovo portfolio
        new_portfolio_btn = ctk.CTkButton(
            selector_frame,
            text="+",
            command=self._create_new_portfolio,
            width=30,
            height=28,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=UIConfig.COLORS['success'],
            hover_color=UIConfig.COLORS['success_hover']
        )
        new_portfolio_btn.pack(side="left")
    
    def _create_navigation_buttons(self):
        """Crea i bottoni di navigazione (destra)"""
        nav_buttons_frame = ctk.CTkFrame(self.navbar_frame, fg_color="transparent")
        nav_buttons_frame.pack(side="right", padx=20, pady=7)

        nav_buttons = [
            ("RoadMap AssetMind", "RoadMap"),
            ("Portfolio", "Portfolio"),
            ("Asset", "Asset"),
            ("Grafici", "Grafici"),
            ("Export", "Export"),
        ]
        for text, page in nav_buttons:
            # Colore iniziale: attivo per la pagina corrente, inattivo per le altre
            if page == self.current_page:
                fg_color = UIConfig.COLORS['primary']
                hover_color = UIConfig.COLORS['primary_hover']
            else:
                fg_color = UIConfig.COLORS['secondary']
                hover_color = UIConfig.COLORS['secondary_hover']
            
            btn = ctk.CTkButton(
                nav_buttons_frame,
                text=text,
                command=lambda p=page: self._on_page_changed(p),
                **UIConfig.BUTTON_SIZES['medium'],
                font=ctk.CTkFont(**UIConfig.FONTS['button']),
                fg_color=fg_color,
                hover_color=hover_color
            )
            btn.pack(side="left", padx=2)
            
            # Salva riferimento al bottone
            self.nav_buttons[page] = btn

    def _on_page_changed(self, page: str):
        """Gestisce il cambio di pagina e aggiorna l'evidenziazione dei bottoni"""
        self.current_page = page
        self.update_active_button(page)
        self.trigger_callback('page_changed', page)
    
    def navigate_to(self, page: str) -> None:
        """Naviga programmaticamente verso una pagina aggiornando lo stato della navbar."""
        self._on_page_changed(page)
    
    def _on_portfolio_changed(self, selected_file: str):
        """Gestisce il cambio di portfolio"""
        self.trigger_callback('portfolio_changed', selected_file)
    
    def _create_new_portfolio(self):
        """Crea un nuovo portfolio"""
        self.trigger_callback('new_portfolio_requested')
    
    def update_active_button(self, active_page: str):
        """Aggiorna l'evidenziazione del bottone attivo"""
        for page, button in self.nav_buttons.items():
            if page == active_page:
                # Bottone attivo - colore primario
                try:
                    button.configure(
                        fg_color=UIConfig.COLORS['primary'],
                        hover_color=UIConfig.COLORS['primary_hover']
                    )
                except Exception as e:
                    self.logger.error(f"Errore aggiornamento bottone attivo {page}: {e}")
            else:
                # Bottoni inattivi - colore secondario  
                try:
                    button.configure(
                        fg_color=UIConfig.COLORS['secondary'],
                        hover_color=UIConfig.COLORS['secondary_hover']
                    )
                except Exception as e:
                    self.logger.error(f"Errore aggiornamento bottone inattivo {page}: {e}")
    
    def update_values(self, total_value: float, selected_value: float = 0, percentage: float = 0):
        """Aggiorna i valori visualizzati"""
        safe_execute(lambda: self.total_value_label.configure(
            text=f"Valore Totale: €{total_value:,.2f}"
        ))
        
        if percentage > 0:
            safe_execute(lambda: self.selected_value_label.configure(
                text=f"Valore selezionato: €{selected_value:,.2f} ({percentage:.1f}%)"
            ))
        else:
            safe_execute(lambda: self.selected_value_label.configure(
                text=f"Valore selezionato: €{selected_value:,.2f}"
            ))
    
    def refresh_portfolio_list(self, portfolio_files: List[str], current_file: str):
        """Aggiorna la lista dei portfolio disponibili"""
        safe_execute(lambda: self.portfolio_selector.configure(values=portfolio_files))
        safe_execute(lambda: self.portfolio_selector.set(current_file))

    def update_counts(self, total_records: int, current_assets: int):
        """Aggiorna i contatori di record e asset correnti"""
        safe_execute(lambda: self.counts_label.configure(
            text=f"Record Totali: {total_records} — Asset Correnti: {current_assets}"
        ))

class PortfolioTable(BaseUIComponent):
    """Componente tabella portfolio con filtri e controlli"""
    
    def __init__(self, parent, portfolio_manager: PortfolioManager):
        super().__init__(parent, portfolio_manager)
        self.table_frame = None
        self.portfolio_tree = None
        self.column_filters = {}
        self.show_all_records = False
        self.tree_style = None
        self.zoom_level = 100
        self.active_filter_popup = None
        self.display_columns: List[str] = []

        # Controlli UI
        self.records_btn = None
        self.assets_btn = None
        self.zoom_label = None
        self.v_scrollbar = None
        self.h_scrollbar = None
        self.market_update_btn = None

        # Performance optimizers
        self.update_manager = None
        self.column_resizer = None
        self.refresh_optimizer = None
    
    def create_table(self) -> ctk.CTkFrame:
        """Crea la tabella portfolio completa"""
        self.table_frame = ctk.CTkFrame(self.parent)
        self.table_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        self._create_controls()
        self._create_tree_view()
        self._setup_tree_style()
        self._initialize_performance_optimizers()
        
        return self.table_frame
    
    def _create_controls(self):
        """Crea i controlli sopra la tabella"""
        controls_frame = ctk.CTkFrame(self.table_frame, height=50)
        controls_frame.pack(fill="x", padx=5, pady=5)
        controls_frame.pack_propagate(False)
        
        # Toggle Record/Asset (sinistra)
        toggle_frame = ctk.CTkFrame(controls_frame, fg_color="transparent")
        toggle_frame.pack(side="left", padx=10, pady=10)
        
        self.records_btn = ctk.CTkButton(
            toggle_frame,
            text="Record 0",
            command=self._toggle_to_records,
            **UIConfig.BUTTON_SIZES['medium'],
            font=ctk.CTkFont(**UIConfig.FONTS['button']),
            fg_color=UIConfig.COLORS['secondary'],
            hover_color=UIConfig.COLORS['secondary_hover']
        )
        self.records_btn.pack(side="left", padx=(0, 5))
        
        self.assets_btn = ctk.CTkButton(
            toggle_frame,
            text="Asset 0",
            command=self._toggle_to_assets,
            **UIConfig.BUTTON_SIZES['medium'],
            font=ctk.CTkFont(**UIConfig.FONTS['button']),
            fg_color=UIConfig.COLORS['primary'],
            hover_color=UIConfig.COLORS['primary_hover']
        )
        self.assets_btn.pack(side="left")
        
        # Controlli zoom (centro)
        zoom_frame = ctk.CTkFrame(controls_frame, fg_color="transparent")
        zoom_frame.pack(side="right", padx=10, pady=10)
        
        ctk.CTkLabel(
            zoom_frame,
            text="Zoom:",
            font=ctk.CTkFont(**UIConfig.FONTS['text'])
        ).pack(side="left", padx=(0, 5))
        
        zoom_out_btn = ctk.CTkButton(
            zoom_frame,
            text="-",
            command=self._zoom_out,
            width=30,
            height=28,
            font=ctk.CTkFont(size=16, weight="bold")
        )
        zoom_out_btn.pack(side="left", padx=2)
        
        self.zoom_label = ctk.CTkLabel(
            zoom_frame,
            text="100%",
            font=ctk.CTkFont(**UIConfig.FONTS['text']),
            width=50
        )
        self.zoom_label.pack(side="left", padx=5)
        
        zoom_in_btn = ctk.CTkButton(
            zoom_frame,
            text="+",
            command=self._zoom_in,
            width=30,
            height=28,
            font=ctk.CTkFont(size=16, weight="bold")
        )
        zoom_in_btn.pack(side="left", padx=2)
        
        # Bottone pulisci filtri (accanto ai controlli Record/Asset)
        clear_filters_btn = ctk.CTkButton(
            toggle_frame,
            text="🗑️ Pulisci Filtri",
            command=self.clear_all_filters,
            **UIConfig.BUTTON_SIZES['medium'],
            font=ctk.CTkFont(**UIConfig.FONTS['button']),
            fg_color=UIConfig.COLORS['warning'],
            hover_color=UIConfig.COLORS['warning_hover']
        )
        clear_filters_btn.pack(side="left", padx=(10, 0))
        
        # Pulsante Riordino
        sort_btn = ctk.CTkButton(
            toggle_frame,
            text="📊 Riordino",
            command=self._sort_records,
            **UIConfig.BUTTON_SIZES['medium'],
            font=ctk.CTkFont(**UIConfig.FONTS['button']),
            fg_color=UIConfig.COLORS['info'],
            hover_color=UIConfig.COLORS['info_hover']
        )
        sort_btn.pack(side="left", padx=(10, 0))

        # Pulsante Reset ID
        reset_id_btn = ctk.CTkButton(
            toggle_frame,
            text="🔄 Reset ID",
            command=self._reset_ids,
            **UIConfig.BUTTON_SIZES['medium'],
            font=ctk.CTkFont(**UIConfig.FONTS['button']),
            fg_color=UIConfig.COLORS['secondary'],
            hover_color=UIConfig.COLORS['secondary_hover']
        )
        reset_id_btn.pack(side="left", padx=(10, 0))

        # Pulsante Aggiorna Prezzi
        self.market_update_btn = ctk.CTkButton(
            toggle_frame,
            text="🔁 Aggiorna Prezzi",
            command=self._on_market_update,
            **UIConfig.BUTTON_SIZES['medium'],
            font=ctk.CTkFont(**UIConfig.FONTS['button']),
            fg_color=UIConfig.COLORS['success'],
            hover_color=UIConfig.COLORS['success_hover']
        )
        self.market_update_btn.pack(side="left", padx=(10, 0))

        # Pulsante Colora Storici
        color_btn = ctk.CTkButton(
            toggle_frame,
            text="🎨 Colora Storici",
            command=self._color_historical_records,
            **UIConfig.BUTTON_SIZES['medium'],
            font=ctk.CTkFont(**UIConfig.FONTS['button']),
            fg_color=UIConfig.COLORS['success'],
            hover_color=UIConfig.COLORS['success_hover']
        )
        color_btn.pack(side="left", padx=(10, 0))
        
        # Testo istruzioni filtri (sulla stessa linea)
        instruction_label = ctk.CTkLabel(
            toggle_frame,
            text="Click column header to filter",
            font=ctk.CTkFont(**UIConfig.FONTS['text']),
            text_color=UIConfig.COLORS['secondary']
        )
        instruction_label.pack(side="left", padx=(20, 0))
    
    def _create_tree_view(self):
        """Crea il TreeView per la tabella"""
        tree_container = ctk.CTkFrame(self.table_frame)
        tree_container.pack(fill="both", expand=True, padx=5, pady=5)
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)
        
        # Definizione colonne in base all'ordine del file Excel
        columns = self._determine_display_columns()
        self.display_columns = list(columns)
        
        # TreeView
        self.portfolio_tree = ttk.Treeview(
            tree_container,
            columns=columns,
            show="headings",
            height=15
        )
        self.portfolio_tree.grid(row=0, column=0, sticky="nsew")
        
        # Configurazione colonne
        self._configure_columns(columns)
        
        # Scrollbars
        self.v_scrollbar = ttk.Scrollbar(
            tree_container,
            orient="vertical",
            command=self.portfolio_tree.yview
        )
        self.portfolio_tree.configure(yscrollcommand=self.v_scrollbar.set)
        
        self.h_scrollbar = ttk.Scrollbar(
            tree_container,
            orient="horizontal",
            command=self.portfolio_tree.xview
        )
        self.portfolio_tree.configure(xscrollcommand=self.h_scrollbar.set)
        
        # Posizionamento scrollbars
        self.v_scrollbar.grid(row=0, column=1, sticky="ns")
        self.h_scrollbar.grid(row=1, column=0, sticky="ew")
        
        # Eventi
        self.portfolio_tree.bind("<Double-1>", self._on_double_click)
        
        # Bind click su header per filtri
        for col in columns:
            self.portfolio_tree.heading(col, command=lambda c=col: self._show_column_filter(c))
        
        # Inizializza gli header senza filtri
        self._update_column_headers()
    
    def _configure_columns(self, columns: tuple):
        """Configura le colonne della tabella con auto-fit iniziale"""
        # Larghezze base per calcolo iniziale
        self.base_column_widths = {
            "ID": 60, "Category": 120, "Position": 120, "Asset Name": 200,
            "ISIN": 80, "Ticker": 80, "Risk Level": 120,
            "Created At": 120, "Created Amount": 160, "Created Unit Price": 140,
            "Created Total Value": 150, "Updated At": 120, "Updated Amount": 160,
            "Updated Unit Price": 140, "Updated Total Value": 150,
            "Accumulation Plan": 180, "Accumulation Amount": 180,
            "Income Per Year": 160, "Rental Income": 140, "Note": 250,
            "Return %": 120
        }

        # Titoli colonne
        column_headers = {
            "ID": "ID ▼",
            "Category": "Category ▼",
            "Position": "Position ▼",
            "Asset Name": "Asset Name ▼",
            "ISIN": "ISIN ▼",
            "Ticker": "Ticker ▼",
            "Risk Level": "Risk Level ▼",
            "Created At": "Created At ▼",
            "Created Amount": "Created Amount ▼",
            "Created Unit Price": "Created Price ▼",
            "Created Total Value": "Created Total ▼",
            "Updated At": "Updated At ▼",
            "Updated Amount": "Updated Amount ▼",
            "Updated Unit Price": "Updated Price ▼",
            "Updated Total Value": "Updated Total ▼",
            "Accumulation Plan": "Accumulation Plan ▼",
            "Accumulation Amount": "Accumulation Amount ▼",
            "Income Per Year": "Income Per Year ▼",
            "Rental Income": "Rental Income ▼",
            "Note": "Note ▼",
            "Return %": "Return % ▼"
        }

        for col in columns:
            width = self.base_column_widths.get(col, 120)
            header_text = column_headers.get(col, f"{col}\n▼")
            self.portfolio_tree.heading(col, text=header_text, anchor="w")
            self.portfolio_tree.column(col, width=width, anchor="w", stretch=False)
    
    def _setup_tree_style(self):
        """Configura lo stile del TreeView"""
        self.tree_style = ttk.Style()
        self.tree_style.configure(
            "Portfolio.Treeview",
            background="white",
            foreground="black",
            fieldbackground="white",
            font=("TkDefaultFont", 9),
            rowheight=25
        )
        
        # Stile per record storici (azzurro)
        self.tree_style.configure(
            "Historical.Treeview",
            background="white",
            foreground="#0066CC",  # Azzurro
            fieldbackground="white",
            font=("TkDefaultFont", 9),
            rowheight=25
        )
        
        # Configurazione header con altezza aumentata
        self.tree_style.configure(
            "Portfolio.Treeview.Heading",
            background="#f0f0f0",
            foreground="black",
            font=("TkDefaultFont", 9, "bold"),
            relief="raised",
            borderwidth=1,
            # Forza altezza minima per due righe
            minsize=50,
            padding=[5, 15, 5, 15]  # left, top, right, bottom
        )
        
        # Configurazione diretta dell'altezza dell'header
        try:
            # Imposta altezza fissa agli header
            self.portfolio_tree.tk.call('style', 'configure', 'Portfolio.Treeview.Heading', '-rowheight', 50)
        except:
            pass
            
        self.portfolio_tree.configure(style="Portfolio.Treeview")

    def _initialize_performance_optimizers(self):
        """Inizializza i sistemi di ottimizzazione performance"""
        try:
            self.update_manager = UIUpdateManager(self.parent)
            self.column_resizer = LazyColumnResizer(self.portfolio_tree, self.parent)
            self.refresh_optimizer = UIRefreshOptimizer(self.parent)

            self.logger.info("Performance optimizers inizializzati")
        except Exception as e:
            self.logger.error(f"Errore inizializzazione performance optimizers: {e}")
    
    def _toggle_to_records(self):
        """Passa alla vista record (tutti)"""
        self.show_all_records = True
        self.records_btn.configure(
            fg_color=UIConfig.COLORS['primary'],
            hover_color=UIConfig.COLORS['primary_hover']
        )
        self.assets_btn.configure(
            fg_color=UIConfig.COLORS['secondary'],
            hover_color=UIConfig.COLORS['secondary_hover']
        )
        # Ricarica i dati per la nuova vista
        self._apply_filters()
        self.trigger_callback('view_changed', 'records')
    
    def _toggle_to_assets(self):
        """Passa alla vista asset (solo attuali)"""
        self.show_all_records = False
        self.assets_btn.configure(
            fg_color=UIConfig.COLORS['primary'],
            hover_color=UIConfig.COLORS['primary_hover']
        )
        self.records_btn.configure(
            fg_color=UIConfig.COLORS['secondary'],
            hover_color=UIConfig.COLORS['secondary_hover']
        )
        # Ricarica i dati per la nuova vista
        self._apply_filters()
        self.trigger_callback('view_changed', 'assets')
    
    def _zoom_in(self):
        """Aumenta lo zoom della tabella"""
        if self.zoom_level < 150:
            self.zoom_level += 10
            self._apply_zoom_optimized()

    def _zoom_out(self):
        """Diminuisce lo zoom della tabella"""
        if self.zoom_level > 70:
            self.zoom_level -= 10
            self._apply_zoom_optimized()
    
    def _apply_zoom_optimized(self):
        """Applica il livello di zoom corrente con ottimizzazioni"""
        if not self.update_manager:
            self._apply_zoom()  # Fallback
            return

        # Usa update manager per debounce dello zoom
        self.update_manager.schedule_update(
            "apply_zoom",
            self._apply_zoom_immediate
        )

    def _apply_zoom_immediate(self):
        """Applica zoom senza debouncing - versione ottimizzata"""
        base_font_size = 9
        new_font_size = int(base_font_size * (self.zoom_level / 100))
        new_height = int(25 * (self.zoom_level / 100))

        self.tree_style.configure(
            "Portfolio.Treeview",
            font=("TkDefaultFont", new_font_size),
            rowheight=new_height
        )

        # Ricalcola altezza header con zoom
        zoom_padding_v = max(15, int(15 * (self.zoom_level / 100)))
        zoom_height = max(50, int(50 * (self.zoom_level / 100)))

        self.tree_style.configure(
            "Portfolio.Treeview.Heading",
            font=("TkDefaultFont", new_font_size, "bold"),
            padding=[5, zoom_padding_v, 5, zoom_padding_v]
        )

        try:
            self.portfolio_tree.tk.call('style', 'configure', 'Portfolio.Treeview.Heading', '-rowheight', zoom_height)
        except:
            pass

        # Aggiorna larghezza colonne proporzionalmente allo zoom
        self._update_column_widths()

        # Notifica column resizer del nuovo zoom
        if self.column_resizer:
            self.column_resizer.update_zoom_factor(self.zoom_level)

        self.zoom_label.configure(text=f"{self.zoom_level}%")

        # Aggiorna scrollbar per mostrare quella orizzontale se necessario
        self._update_scrollbars()

        # Usa refresh ottimizzato
        if self.refresh_optimizer:
            self.refresh_optimizer.smart_refresh()

    def _apply_zoom(self):
        """Metodo fallback per zoom senza ottimizzazioni"""
        base_font_size = 9
        new_font_size = int(base_font_size * (self.zoom_level / 100))
        new_height = int(25 * (self.zoom_level / 100))

        self.tree_style.configure(
            "Portfolio.Treeview",
            font=("TkDefaultFont", new_font_size),
            rowheight=new_height
        )

        self.zoom_label.configure(text=f"{self.zoom_level}%")
        self._update_scrollbars()
    
    def _update_column_widths(self):
        """Aggiorna le larghezze delle colonne proporzionalmente al zoom"""
        try:
            # Calcola il fattore di zoom
            zoom_factor = self.zoom_level / 100

            # Aggiorna le larghezze di tutte le colonne
            for col in self.portfolio_tree['columns']:
                base_width = self.base_column_widths.get(col, 120)
                new_width = int(base_width * zoom_factor)
                self.portfolio_tree.column(col, width=new_width)

            self.logger.debug(f"Larghezze colonne aggiornate con zoom {self.zoom_level}%")

        except Exception as e:
            self.logger.error(f"Errore aggiornamento larghezza colonne: {e}")
    
    def _update_scrollbars(self):
        """Aggiorna la visibilità delle scrollbar (sempre mostrate se necessarie)"""
        try:
            # Forza aggiornamento dopo un breve delay per permettere al TreeView di calcolare le dimensioni
            def update():
                try:
                    # Mostra scrollbar verticale se c'è contenuto scrollabile
                    if self.portfolio_tree.yview() != (0.0, 1.0):
                        self.v_scrollbar.grid()

                    # Mostra scrollbar orizzontale se c'è contenuto scrollabile
                    if self.portfolio_tree.xview() != (0.0, 1.0):
                        self.h_scrollbar.grid()
                except:
                    pass

            self.parent.after(100, update)
        except Exception as e:
            self.logger.error(f"Errore aggiornamento scrollbar: {e}")
    
    def _on_double_click(self, event):
        """Gestisce il doppio click su una riga"""
        item = self.portfolio_tree.selection()[0] if self.portfolio_tree.selection() else None
        if item:
            values = self.portfolio_tree.item(item, "values")
            if values:
                asset_id = int(values[0])
                self.trigger_callback('asset_selected', asset_id)
    
    def _show_column_filter(self, column: str):
        """Mostra il filtro per una colonna specifica"""
        try:
            # Chiudi popup precedente se aperto
            if hasattr(self, 'active_filter_popup') and self.active_filter_popup:
                self.active_filter_popup.destroy()
                self.active_filter_popup = None
            
            # Ottieni i dati appropriati per i valori unici
            if self.show_all_records:
                df = self.portfolio_manager.load_data()
            else:
                df = self.portfolio_manager.get_current_assets_only()
            if df.empty:
                return
            
            # Converti il nome colonna display in nome DataFrame
            from config import FieldMapping
            db_column = FieldMapping.DISPLAY_TO_DB.get(column, column.lower().replace(' ', '_'))
            
            if db_column not in df.columns:
                return
            
            # Ottieni valori unici per la colonna
            unique_values = df[db_column].fillna('N/A').astype(str).unique()
            unique_values = sorted([v for v in unique_values if v != ''])
            
            if not unique_values:
                return
            
            # Crea popup filtro
            self._create_filter_popup(column, db_column, unique_values)
            
        except Exception as e:
            self.logger.error(f"Errore nel filtro colonna {column}: {e}")
    
    def _create_filter_popup(self, display_column: str, db_column: str, values: list):
        """Crea il popup per il filtro colonna (stile legacy semplificato)"""
        import customtkinter as ctk
        
        # Crea finestra popup CTk per consistenza
        popup = ctk.CTkToplevel(self.parent)
        popup.title(f"Filter: {display_column}")
        popup.geometry("280x450")
        popup.transient(self.parent)
        popup.grab_set()
        
        self.active_filter_popup = popup
        
        # Header
        header_label = ctk.CTkLabel(popup, text=f"Filter by {display_column}",
                                   font=ctk.CTkFont(**UIConfig.FONTS['subheader']))
        header_label.pack(pady=(15, 10))
        
        # Search box per filtri di testo (solo per campi testuali)
        if db_column in ['asset_name', 'position', 'note', 'isin', 'ticker', 'accumulation_plan']:
            search_entry = ctk.CTkEntry(popup, placeholder_text="Search...", width=240)
            search_entry.pack(pady=(0, 10), padx=20)
        else:
            search_entry = None
        
        # Frame scrollabile per i checkbox
        checkbox_frame = ctk.CTkScrollableFrame(popup, width=240, height=280)
        checkbox_frame.pack(pady=10, padx=20, fill="both", expand=True)
        
        # Ottieni filtri attivi per questa colonna
        active_filters = self.column_filters.get(db_column, set())
        
        # Crea checkbox per ogni valore unico
        checkboxes = {}
        for value in values:
            # Determina se deve essere selezionato (se non ci sono filtri, seleziona tutti)
            is_selected = len(active_filters) == 0 or value in active_filters
            
            checkbox = ctk.CTkCheckBox(
                checkbox_frame,
                text=str(value),
                font=ctk.CTkFont(**UIConfig.FONTS['text'])
            )
            checkbox.pack(anchor="w", pady=2, padx=5)
            
            if is_selected:
                checkbox.select()
            else:
                checkbox.deselect()
                
            checkboxes[value] = checkbox
        
        # Funzione ricerca in tempo reale
        def on_search(*args):
            if search_entry:
                search_text = search_entry.get().lower()
                for value, checkbox in checkboxes.items():
                    value_text = str(value).lower()
                    if search_text in value_text:
                        checkbox.pack(anchor="w", pady=2, padx=5)
                    else:
                        checkbox.pack_forget()
        
        if search_entry:
            search_entry.bind('<KeyRelease>', on_search)
        
        # Frame bottoni
        button_frame = ctk.CTkFrame(popup, fg_color="transparent")
        button_frame.pack(fill="x", pady=10, padx=20)
        
        # Bottone Select All
        def select_all():
            for checkbox in checkboxes.values():
                checkbox.select()
        
        select_all_btn = ctk.CTkButton(
            button_frame, text="Select All", command=select_all,
            width=80, height=28, font=ctk.CTkFont(**UIConfig.FONTS['text'])
        )
        select_all_btn.pack(side="left", padx=(0, 5))
        
        # Bottone Clear All
        def clear_all():
            for checkbox in checkboxes.values():
                checkbox.deselect()
        
        clear_all_btn = ctk.CTkButton(
            button_frame, text="Clear All", command=clear_all,
            width=80, height=28, font=ctk.CTkFont(**UIConfig.FONTS['text']),
            fg_color=UIConfig.COLORS['secondary']
        )
        clear_all_btn.pack(side="left", padx=5)
        
        # Bottone Apply
        def apply_filter():
            # Raccogli valori selezionati
            selected_values = set()
            for value, checkbox in checkboxes.items():
                if checkbox.get():
                    selected_values.add(str(value))
            
            # Applica o rimuovi filtro
            if len(selected_values) == len(checkboxes) or len(selected_values) == 0:
                # Se tutti selezionati o nessuno, rimuovi filtro
                self.column_filters.pop(db_column, None)
            else:
                self.column_filters[db_column] = selected_values

            # Notifica filtri cambiati (wiring globale)
            try:
                self.trigger_callback('filters_changed', {
                    'column_filters': {k: set(v) for k, v in self.column_filters.items()},
                    'show_all_records': bool(self.show_all_records),
                })
            except Exception:
                pass

            # Aggiorna visualizzazione locale (fallback) e header
            self._apply_filters()
            self._update_column_headers()
            
            popup.destroy()
            self.active_filter_popup = None
        
        apply_btn = ctk.CTkButton(
            button_frame, text="Apply", command=apply_filter,
            width=80, height=28, font=ctk.CTkFont(**UIConfig.FONTS['text']),
            fg_color=UIConfig.COLORS['success']
        )
        apply_btn.pack(side="right")
        
        # Chiusura popup
        def on_close():
            self.active_filter_popup = None
            popup.destroy()
        
        popup.protocol("WM_DELETE_WINDOW", on_close)
    
    def _apply_filters(self):
        """Applica i filtri attivi ai dati"""
        try:
            # Carica dati base
            if self.show_all_records:
                df = self.portfolio_manager.load_data()
            else:
                df = self.portfolio_manager.get_current_assets_only()
            
            # Applica filtri colonna
            for column, filter_values in self.column_filters.items():
                if column in df.columns and filter_values:
                    # Converte i valori della colonna in stringa per il confronto
                    df_values = df[column].fillna('N/A').astype(str)
                    df = df[df_values.isin(filter_values)]
            
            # Aggiorna la tabella con i dati filtrati
            self.update_data(df)
            
            # Notifica il cambiamento per aggiornare i valori
            self.trigger_callback('data_filtered', df)
            
        except Exception as e:
            self.logger.error(f"Errore nell'applicazione filtri: {e}")
    
    def _update_column_headers(self):
        """Aggiorna le intestazioni delle colonne per mostrare i filtri attivi con asterisco più grande"""
        try:
            # Mappa per i nomi delle colonne display -> dataframe
            from config import FieldMapping
            
            # Titoli su una riga con mapping per filtri
            column_headers_base = {
                "ID": "ID",
                "Category": "Category",
                "Position": "Position", 
                "Asset Name": "Asset Name",
                "ISIN": "ISIN",
                "Ticker": "Ticker",
                "Risk Level": "Risk Level",
                "Created At": "Created At",
                "Created Amount": "Created Amount",
                "Created Unit Price": "Created Price",
                "Created Total Value": "Created Total",
                "Updated At": "Updated At",
                "Updated Amount": "Updated Amount",
                "Updated Unit Price": "Updated Price",
                "Updated Total Value": "Updated Total",
                "Accumulation Plan": "Accumulation Plan",
                "Accumulation Amount": "Accumulation Amount",
                "Income Per Year": "Income Per Year",
                "Rental Income": "Rental Income",
                "Note": "Note",
                "Return %": "Return %"
            }
            
            column_order = self.display_columns if self.display_columns else list(FieldMapping.DISPLAY_TO_DB.keys())

            for display_name in column_order:
                db_name = FieldMapping.DISPLAY_TO_DB.get(display_name, display_name)
                base_header = column_headers_base.get(display_name, display_name)
                
                if db_name in self.column_filters:
                    # Mostra asterisco doppio più visibile per indicare filtro attivo
                    header_text = f"{base_header} ▼ **"
                else:
                    # Intestazione normale con triangolo
                    header_text = f"{base_header} ▼"
                
                # Aggiorna header (se la colonna esiste nella TreeView)
                try:
                    self.portfolio_tree.heading(display_name, text=header_text)
                except:
                    pass  # Colonna non esistente nel TreeView, ignora
                    
        except Exception as e:
            self.logger.error(f"Errore aggiornamento headers: {e}")
    
    def clear_all_filters(self):
        """Pulisce tutti i filtri attivi"""
        self.column_filters.clear()

        # Notifica filtri rimossi (wiring globale) - IMPORTANTE per aggiornare grafici!
        try:
            self.trigger_callback('filters_changed', {
                'column_filters': {},
                'show_all_records': bool(self.show_all_records),
            })
        except Exception as e:
            self.logger.error(f"Errore notifica clear_all_filters: {e}")

        self._apply_filters()
        self._update_column_headers()

        if hasattr(self, 'active_filter_popup') and self.active_filter_popup:
            self.active_filter_popup.destroy()
            self.active_filter_popup = None

    def cleanup_performance_optimizers(self):
        """Pulisce i sistemi di ottimizzazione performance alla chiusura"""
        try:
            if self.update_manager:
                self.update_manager.cleanup()

            if self.column_resizer:
                self.column_resizer.cached_widths.clear()

            self.logger.debug("Performance optimizers cleanup completato")
        except Exception as e:
            self.logger.error(f"Errore cleanup performance optimizers: {e}")
    
    def update_data(self, df: pd.DataFrame):
        """Aggiorna i dati della tabella"""
        self.logger.debug(f"update_data() chiamato con DataFrame: {len(df)} righe, vuoto: {df.empty}")
        self.logger.debug(f"portfolio_tree exists: {self.portfolio_tree is not None}")
        
        # Pulisce la tabella esistente
        try:
            children_count = len(self.portfolio_tree.get_children())
            self.logger.debug(f"Cancellando {children_count} righe esistenti dalla tabella")
            for item in self.portfolio_tree.get_children():
                self.portfolio_tree.delete(item)
            self.logger.debug("Tabella pulita con successo")
        except Exception as e:
            self.logger.error(f"Errore durante la pulizia della tabella: {e}")
        
        if df.empty:
            self.logger.debug("DataFrame vuoto, esco senza aggiungere righe")
            return

        # Aggiorna l'ordine delle colonne se necessario
        try:
            new_display_columns = [FieldMapping.DB_TO_DISPLAY.get(col, col) for col in df.columns]
            if self.display_columns != new_display_columns:
                self.display_columns = new_display_columns
                self.portfolio_tree["columns"] = new_display_columns
                self._configure_columns(tuple(new_display_columns))
                for col in new_display_columns:
                    self.portfolio_tree.heading(col, command=lambda c=col: self._show_column_filter(c))
                self._update_column_headers()
        except Exception as col_exc:
            self.logger.debug(f"Impossibile aggiornare l'ordine colonne dinamicamente: {col_exc}")

        # Carica i colori dal file Excel
        from openpyxl import load_workbook
        from openpyxl.styles.colors import Color

        excel_colors = {}  # {row_id: {'fg': color, 'bg': color}}
        try:
            wb = load_workbook(self.portfolio_manager.excel_file)
            ws = wb.active

            # Leggi i colori per ogni riga (skip header)
            for row_idx in range(2, ws.max_row + 1):
                id_cell = ws.cell(row=row_idx, column=1)
                try:
                    row_id = int(id_cell.value)
                except (TypeError, ValueError):
                    continue

                # Leggi font color (per record storici = azzurro)
                fg_color = None
                if id_cell.font and id_cell.font.color:
                    if hasattr(id_cell.font.color, 'rgb') and id_cell.font.color.rgb:
                        fg_color = id_cell.font.color.rgb

                # Leggi background color (per alert = rosso)
                bg_color = None
                if id_cell.fill and id_cell.fill.patternType:
                    if id_cell.fill.fgColor and hasattr(id_cell.fill.fgColor, 'rgb'):
                        if id_cell.fill.fgColor.rgb and id_cell.fill.fgColor.rgb != '00000000':
                            bg_color = id_cell.fill.fgColor.rgb

                if fg_color or bg_color:
                    excel_colors[row_id] = {'fg': fg_color, 'bg': bg_color}

            wb.close()
            self.logger.info(f"Caricati colori Excel per {len(excel_colors)} righe")
        except Exception as e:
            self.logger.error(f"Errore caricamento colori da Excel: {e}")

        # Inserisce i dati con colorazione da Excel
        self.logger.debug(f"Iniziando inserimento {len(df)} righe nella tabella")
        rows_inserted = 0

        for _, row in df.iterrows():
            try:
                values = self._format_row_values(row)
                if rows_inserted < 3:
                    self.logger.debug(f"Inserendo riga {rows_inserted + 1}: ID={row['id']}, Asset={row.get('asset_name', 'N/A')}")

                item_id = self.portfolio_tree.insert("", "end", values=values)

                # Applica colori da Excel
                try:
                    row_id = int(row['id'])
                    if row_id in excel_colors:
                        tag_name = f"row_{row_id}"
                        colors = excel_colors[row_id]

                        # Converti colori RGB da Excel a formato #RRGGBB
                        fg_hex = None
                        bg_hex = None

                        if colors['fg'] and len(colors['fg']) >= 6:
                            fg_hex = f"#{colors['fg'][-6:]}"  # Ultimi 6 char = RRGGBB
                        if colors['bg'] and len(colors['bg']) >= 6:
                            bg_hex = f"#{colors['bg'][-6:]}"

                        # Configura tag dinamico per questa riga
                        tag_config = {}
                        if fg_hex:
                            tag_config['foreground'] = fg_hex
                        if bg_hex:
                            tag_config['background'] = bg_hex

                        if tag_config:
                            self.portfolio_tree.tag_configure(tag_name, **tag_config)
                            self.portfolio_tree.item(item_id, tags=(tag_name,))

                except (TypeError, ValueError, KeyError):
                    pass

                rows_inserted += 1
            except Exception as e:
                self.logger.error(f"Errore inserimento riga {rows_inserted}: {e}")

        self.logger.debug(f"Inserimento completato: {rows_inserted} righe inserite su {len(df)} totali")
        
        # Aggiorna contatori
        self._update_button_counts(df)
        
        # Auto-ridimensiona le colonne con sistema ottimizzato
        if self.column_resizer:
            self.column_resizer.invalidate_cache()
            self.column_resizer.schedule_resize()
        # Nota: column_resizer è sempre inizializzato, quindi il fallback non è più necessario

        # Aggiorna scrollbar con debouncing
        if self.update_manager:
            self.update_manager.schedule_update("update_scrollbars", self._update_scrollbars)
        else:
            self._update_scrollbars()
        
        # Refresh ottimizzato dell'interfaccia
        try:
            self.portfolio_tree.update_idletasks()
            self.parent.update_idletasks()
            self.logger.debug("Refresh interfaccia completato")
        except Exception as e:
            self.logger.error(f"Errore durante refresh: {e}")

        self.logger.debug("update_data() COMPLETATO")
    
    def _format_row_values(self, row: pd.Series) -> tuple:
        """Formatta i valori di una riga per la visualizzazione"""
        formatted_values: List[Any] = []
        columns = self.display_columns or list(self.portfolio_tree["columns"])

        for display_col in columns:
            db_field = FieldMapping.DISPLAY_TO_DB.get(display_col, display_col)
            value = row.get(db_field, row.get(display_col))

            if db_field in FieldMapping.DATE_FIELDS:
                formatted_values.append(DateFormatter.format_for_display(value))
            elif db_field in FieldMapping.MONETARY_FIELDS:
                formatted_values.append(CurrencyFormatter.format_for_display(value))
            elif db_field in {'created_amount', 'updated_amount'}:
                try:
                    formatted_values.append(f"{float(value):,.2f}" if pd.notna(value) else "0")
                except (TypeError, ValueError):
                    formatted_values.append("0")
            elif db_field == 'return_percentage':
                formatted_values.append(self._format_return_percentage(value))
            elif db_field == 'note':
                formatted_values.append(str(value) if pd.notna(value) and str(value).strip() else "-")
            elif db_field in {'risk_level'}:
                formatted_values.append(str(value) if pd.notna(value) else "0")
            elif db_field == 'id':
                formatted_values.append(value if pd.notna(value) else "-")
            else:
                formatted_values.append(str(value) if pd.notna(value) and str(value).strip() else "-")

        return tuple(formatted_values)
    
    def _format_return_percentage(self, value: Any) -> str:
        """Formatta il rendimento annualizzato, mostrando '-' se non disponibile."""
        try:
            if value is None or pd.isna(value):
                return "-"
            return f"{float(value):.2f}%"
        except (TypeError, ValueError):
            return "-"

    def _determine_display_columns(self, df: Optional[pd.DataFrame] = None) -> tuple[str, ...]:
        """
        Determina le colonne da visualizzare seguendo l'ordine del file Excel.

        Args:
            df: DataFrame opzionale da cui estrarre le colonne.
                Se None, carica i dati (compatibilità backward)

        Returns:
            Tuple con i nomi delle colonne display
        """
        ordered_columns: List[str] = []
        try:
            # Usa DataFrame passato o carica da disco (fallback)
            df_snapshot = df if df is not None else self.portfolio_manager.load_data()
            if not df_snapshot.empty:
                for db_field in df_snapshot.columns:
                    display_name = FieldMapping.DB_TO_DISPLAY.get(db_field, db_field)
                    ordered_columns.append(display_name)
        except Exception as exc:
            self.logger.debug(f"Impossibile determinare colonne dal DataFrame: {exc}")

        if not ordered_columns:
            ordered_columns = [
                "ID", "Category", "Position", "Asset Name", "ISIN", "Ticker", "Risk Level",
                "Created At", "Created Amount", "Created Unit Price", "Created Total Value",
                "Updated At", "Updated Amount", "Updated Unit Price", "Updated Total Value",
                "Accumulation Plan", "Accumulation Amount", "Income Per Year", "Rental Income",
                "Note", "Return %"
            ]

        return tuple(ordered_columns)
    
    def _update_button_counts(self, df: pd.DataFrame):
        """Aggiorna i contatori sui bottoni Record/Asset"""
        try:
            total_records = len(self.portfolio_manager.load_data())
            current_assets = len(self.portfolio_manager.get_current_assets_only())

            # Salva i valori per uso esterno
            self._last_total_records = total_records
            self._last_current_assets = current_assets

            safe_execute(lambda: self.records_btn.configure(text=f"Record {total_records}"))
            safe_execute(lambda: self.assets_btn.configure(text=f"Asset {current_assets}"))
        except Exception as e:
            self.logger.error(f"Errore aggiornamento contatori: {e}")

    def get_counts(self) -> tuple[int, int]:
        """Restituisce i contatori correnti (total_records, current_assets)"""
        return (
            getattr(self, '_last_total_records', 0),
            getattr(self, '_last_current_assets', 0)
        )
    
    def get_visible_value(self) -> tuple[float, int]:
        """
        Calcola il valore delle righe visibili usando la stessa logica di deduplica di get_portfolio_summary()

        IMPORTANTE: Ora usa direttamente i dati pandas invece di parsare stringhe dalla TreeView,
        garantendo totale coerenza con get_portfolio_summary() ed eliminando errori di arrotondamento.
        """
        import pandas as pd

        # Ottieni gli ID visibili dalla TreeView
        visible_ids = []
        for child in self.portfolio_tree.get_children():
            item_values = self.portfolio_tree.item(child)['values']
            if len(item_values) > 0:
                try:
                    asset_id = int(item_values[0])  # Prima colonna = ID
                    visible_ids.append(asset_id)
                except (ValueError, TypeError, IndexError):
                    continue

        if not visible_ids:
            return 0.0, 0

        # Carica i dati completi dal portfolio manager
        df = self.portfolio_manager.load_data()

        if df.empty:
            return 0.0, 0

        # Filtra solo i record visibili nella TreeView
        df_visible = df[df['id'].isin(visible_ids)].copy()

        if df_visible.empty:
            return 0.0, 0

        # APPLICA LA STESSA LOGICA DI get_portfolio_summary() in models.py
        # Normalizza i campi chiave per deduplica
        def _norm(s):
            if pd.isna(s):
                return ''
            val = str(s).strip()
            if val.lower() in {'na','n/a','none','null','nan',''}:
                return ''
            return val

        for key_col in ['category','asset_name','position','isin']:
            if key_col in df_visible.columns:
                df_visible[key_col] = df_visible[key_col].apply(_norm)
            else:
                df_visible[key_col] = ''

        # Crea chiave asset unica (stessa logica di models.py)
        df_visible['asset_key'] = (df_visible['category'] + '|' +
                                   df_visible['asset_name'] + '|' +
                                   df_visible['position'] + '|' +
                                   df_visible['isin'])

        # Converti date per ordinamento (stessa logica di models.py)
        df_visible['effective_date'] = pd.to_datetime(
            df_visible['updated_at'].replace(['', 'NA', 'N/A', 'na'], pd.NA).fillna(df_visible['created_at']),
            format='%Y-%m-%d', errors='coerce'
        )

        # Per ogni asset unico, prendi solo il record con data più recente (stessa logica di models.py)
        latest_records = df_visible.sort_values('effective_date', ascending=False).groupby('asset_key').first().reset_index()

        # Calcola il totale usando ESATTAMENTE la stessa formula di get_portfolio_summary()
        visible_value = latest_records['updated_total_value'].fillna(latest_records['created_total_value']).sum()
        visible_count = len(latest_records)

        # Log per debug
        if len(df_visible) != visible_count:
            self.logger.debug(f"get_visible_value: {len(df_visible)} record visibili -> "
                            f"{visible_count} asset deduplicati -> €{visible_value:,.2f}")

        return visible_value, visible_count
    
    def _sort_records(self):
        """Riordina i record del file Excel per categoria, posizione, nome asset, ISIN e data update"""
        try:
            # Conferma utente
            result = messagebox.askyesno(
                "Riordino Record", 
                "Vuoi riordinare tutti i record del file Excel?\n\n"
                "I record verranno ordinati per:\n"
                "1. Categoria\n"
                "2. Posizione\n"
                "3. Nome Asset\n"
                "4. ISIN\n"
                "5. Data Update\n\n"
                "Questa operazione modificherà permanentemente il file Excel."
            )
            
            if not result:
                return
            
            # Carica tutti i dati dal file Excel
            import pandas as pd
            df = pd.read_excel(self.portfolio_manager.excel_file)
            
            if df.empty:
                messagebox.showinfo("Info", "Nessun record da riordinare.")
                return
            
            self.logger.info(f"Riordinamento di {len(df)} record...")
            
            # Converte le date in formato datetime per ordinamento corretto
            for date_col in ['updated_at', 'created_at']:
                if date_col in df.columns:
                    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
            
            # Riordina i record secondo l'ordine specificato
            # Usa fillna per gestire valori mancanti nell'ordinamento
            df_sorted = df.sort_values(
                by=['category', 'position', 'asset_name', 'isin', 'updated_at'],
                na_position='last',
                ascending=True
            ).copy()
            
            self.logger.info("Record riordinati con successo")
            if not df_sorted.empty:
                first_row = df_sorted.iloc[0]
                self.logger.debug(f"Primo record: {first_row.get('category', 'N/A')} | {first_row.get('asset_name', 'N/A')}")
            
            # Salva il file Excel riordinato
            df_sorted.to_excel(self.portfolio_manager.excel_file, index=False)
            
            # Ricarica i dati nell'applicazione
            self.trigger_callback('data_changed')
            
            messagebox.showinfo(
                "Riordino Completato", 
                f"Record riordinati con successo!\n\n"
                f"Totale record: {len(df_sorted)}\n"
                f"File aggiornato: {self.portfolio_manager.excel_file}"
            )
            
            self.logger.info("File Excel aggiornato e ricaricato")

        except Exception as e:
            error_msg = ErrorHandler.handle_file_error(e, "riordino record")
            messagebox.showerror("Errore Riordino", f"Errore durante il riordino:\n\n{error_msg}")
            self.logger.error(f"Errore riordino record: {e}")
            import traceback
            self.logger.debug(f"Stack trace: {traceback.format_exc()}")

    def _reset_ids(self):
        """Rinumera progressivamente gli ID da 1 e rimuove tutte le evidenziazioni rosse"""
        try:
            # Conferma utente
            result = messagebox.askyesno(
                "Reset ID",
                "Vuoi resettare gli ID di tutti i record?\n\n"
                "Questa operazione:\n"
                "• Rinumererà gli ID progressivamente da 1\n"
                "• Rimuoverà tutte le evidenziazioni di colore\n"
                "• Modificherà permanentemente il file Excel\n\n"
                "Continuare?"
            )

            if not result:
                return

            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill

            # Carica il file Excel direttamente con openpyxl (NON con pandas)
            wb = load_workbook(self.portfolio_manager.excel_file)
            ws = wb.active

            if ws.max_row <= 1:  # Solo header
                messagebox.showinfo("Info", "Nessun record presente.")
                wb.close()
                return

            total_records = ws.max_row - 1  # Escludi header
            self.logger.info(f"Reset ID per {total_records} record...")

            # Rinumera gli ID nella prima colonna (colonna A)
            for row_idx, new_id in enumerate(range(1, total_records + 1), start=2):
                ws.cell(row=row_idx, column=1).value = new_id

            # Rimuovi solo il riempimento di colore di sfondo (lascia intatti i font)
            default_fill = PatternFill(fill_type=None)

            for row_idx in range(2, ws.max_row + 1):  # Salta l'header (riga 1)
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # Rimuovi solo il fill, NON toccare il font
                    cell.fill = default_fill

            # Rimuovi tutte le regole di formattazione condizionale
            if hasattr(ws, 'conditional_formatting'):
                # Svuota completamente il conditional_formatting
                ws.conditional_formatting._cf_rules.clear()

            wb.save(self.portfolio_manager.excel_file)
            wb.close()

            self.logger.info("ID resettati e evidenziazioni rimosse")

            # Ricarica i dati nell'applicazione
            self.trigger_callback('data_changed')

            messagebox.showinfo(
                "Reset Completato",
                f"Reset completato con successo!\n\n"
                f"• {total_records} ID rinumerati da 1 a {total_records}\n"
                f"• Evidenziazioni di sfondo rimosse\n\n"
                f"File aggiornato: {self.portfolio_manager.excel_file}"
            )

        except Exception as e:
            error_msg = ErrorHandler.handle_file_error(e, "reset ID")
            messagebox.showerror("Errore Reset", f"Errore durante il reset:\n\n{error_msg}")
            self.logger.error(f"Errore reset ID: {e}")
            import traceback
            self.logger.debug(f"Stack trace: {traceback.format_exc()}")

    def _on_market_update(self):
        """Richiede l'aggiornamento dei prezzi di mercato."""
        self.trigger_callback('market_update_requested')

    def set_market_update_state(self, is_running: bool):
        """Aggiorna lo stato del pulsante di aggiornamento prezzi."""
        if not self.market_update_btn:
            return

        new_state = "disabled" if is_running else "normal"
        new_text = "⏳ Aggiornamento..." if is_running else "🔁 Aggiorna Prezzi"
        safe_execute(lambda: self.market_update_btn.configure(state=new_state, text=new_text))

    # Funzione rimossa: mark_alert_rows() - non più necessaria
    # Le evidenziazioni ora vengono lette direttamente dal file Excel

    def _color_historical_records(self):
        """Colora i record storici di azzurro nel file Excel"""
        try:
            # Conferma utente
            result = messagebox.askyesno(
                "Colora Record Storici", 
                "Vuoi colorare i record storici di azzurro nel file Excel?\n\n"
                "I record storici sono quelli che non rappresentano più\n"
                "lo stato attuale di un asset (sostituiti da versioni più recenti).\n\n"
                "Questa operazione modificherà permanentemente il file Excel."
            )
            
            if not result:
                return
            
            # Esegui la colorazione
            self.portfolio_manager.color_historical_records()
            
            messagebox.showinfo(
                "Colorazione Completata", 
                "I record storici sono stati colorati di azzurro!\n\n"
                "Riapri il file Excel per vedere le modifiche."
            )
            
        except Exception as e:
            from utils import ErrorHandler
            error_msg = ErrorHandler.handle_file_error(e, "colorazione record storici")
            messagebox.showerror("Errore Colorazione", f"Errore durante la colorazione:\n\n{error_msg}")
            self.logger.error(f"Errore colorazione record storici: {e}")
            import traceback
            self.logger.debug(f"Stack trace: {traceback.format_exc()}")







