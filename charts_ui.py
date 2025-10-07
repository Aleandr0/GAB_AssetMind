#!/usr/bin/env python3
"""
Componente interfaccia grafici per GAB AssetMind
Gestisce la visualizzazione di grafici e analytics del portfolio
"""

import customtkinter as ctk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import pandas as pd
from typing import Optional, Dict, Any
import numpy as np

# Nessuna dipendenza da scipy - uso interpolazione semplice con numpy

from config import UIConfig
from utils import ErrorHandler, safe_execute
from models import PortfolioManager, apply_global_filters
from ui_components import BaseUIComponent
from logging_config import get_logger
from date_utils import get_date_manager
from typing import Optional, Dict, Any

class ChartsUI(BaseUIComponent):
    """Componente per la visualizzazione di grafici e analytics"""
    
    
    def __init__(self, parent, portfolio_manager: PortfolioManager):
        super().__init__(parent, portfolio_manager)
        self.charts_frame = None
        self.chart_frame = None
        self.chart_type = None
        self.current_chart = None
        self._external_filtered_df = None
        self.start_year = None
        self.end_year = None
        self.available_years = []
        self.logger = get_logger('ChartsUI')

        # Configurazione matplotlib per CustomTkinter
        plt.style.use('default')
        plt.rcParams.update({
            'figure.facecolor': 'white',
            'axes.facecolor': 'white',
            'axes.edgecolor': 'black',
            'axes.labelcolor': 'black',
            'text.color': 'black',
            'xtick.color': 'black',
            'ytick.color': 'black'
        })
    
    def create_charts_interface(self) -> ctk.CTkFrame:
        """Crea l'interfaccia completa per i grafici"""
        self.charts_frame = ctk.CTkFrame(self.parent)
        self.charts_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        self._create_controls()
        self._create_chart_area()
        
        # Carica il grafico di default
        self.chart_type.set("Distribuzione Valore per Categoria")
        self._update_chart()

        return self.charts_frame
    
    def _create_controls(self):
        """Crea i controlli per la selezione del tipo di grafico"""
        self.controls_frame = ctk.CTkFrame(self.charts_frame, height=60)
        self.controls_frame.pack(fill="x", padx=10, pady=5)
        self.controls_frame.pack_propagate(False)
        
        # Label
        ctk.CTkLabel(
            self.controls_frame,
            text="Tipo Grafico:",
            font=ctk.CTkFont(**UIConfig.FONTS['subheader'])
        ).pack(side="left", padx=20, pady=15)

        # Dropdown selezione grafico
        chart_types = [
            "Distribuzione Valore per Categoria",
            "Distribuzione Rischio",
            "Performance per Categoria",
            "Evoluzione Temporale",
            "Suddivisione Asset per Posizione"
        ]

        self.chart_type = ctk.StringVar(value=chart_types[0])

        chart_selector = ctk.CTkComboBox(
            self.controls_frame,
            values=chart_types,
            variable=self.chart_type,
            command=self._on_chart_type_changed,
            width=250,
            font=ctk.CTkFont(**UIConfig.FONTS['text'])
        )
        chart_selector.pack(side="left", padx=20, pady=15)

        # Label filtri attivi (multilinea con wrapping)
        # Riduce font del 30%: se font base è 12, diventa 8.4 ≈ 8
        base_size = UIConfig.FONTS['text'].get('size', 12)
        reduced_size = int(base_size * 0.7)

        self.filter_label = ctk.CTkLabel(
            self.controls_frame,
            text="",
            font=ctk.CTkFont(family=UIConfig.FONTS['text'].get('family', 'Arial'), size=reduced_size),
            text_color=UIConfig.COLORS['secondary'],
            justify="left",
            wraplength=600  # Larghezza massima prima di andare a capo
        )
        self.filter_label.pack(side="left", padx=20, pady=15, fill="x", expand=True)

        # Controlli temporali (inizialmente nascosti)
        # Label Anno Iniziale
        self.start_year_label = ctk.CTkLabel(
            self.controls_frame,
            text="Dal:",
            font=ctk.CTkFont(**UIConfig.FONTS['text'])
        )

        # Dropdown anno iniziale
        self.start_year = ctk.StringVar()
        self.start_year_selector = ctk.CTkComboBox(
            self.controls_frame,
            variable=self.start_year,
            command=self._on_temporal_range_changed,
            width=80,
            font=ctk.CTkFont(**UIConfig.FONTS['text'])
        )
        
        # Label Anno Finale
        self.end_year_label = ctk.CTkLabel(
            self.controls_frame,
            text="Al:",
            font=ctk.CTkFont(**UIConfig.FONTS['text'])
        )

        # Dropdown anno finale
        self.end_year = ctk.StringVar()
        self.end_year_selector = ctk.CTkComboBox(
            self.controls_frame,
            variable=self.end_year,
            command=self._on_temporal_range_changed,
            width=80,
            font=ctk.CTkFont(**UIConfig.FONTS['text'])
        )
        
        # Inizialmente nascosti
        self.temporal_controls = [
            self.start_year_label, self.start_year_selector,
            self.end_year_label, self.end_year_selector
        ]
    
    def _create_chart_area(self):
        """Crea l'area per la visualizzazione dei grafici"""
        self.chart_frame = ctk.CTkFrame(self.charts_frame)
        self.chart_frame.pack(fill="both", expand=True, padx=10, pady=5)
    
    def _on_chart_type_changed(self, selected_type: str):
        """Gestisce il cambio di tipo di grafico"""
        self.logger.debug(f"Cambio grafico a: {selected_type}")
        
        # Mostra/nascondi controlli temporali
        if selected_type == "Evoluzione Temporale":
            self.logger.debug("Mostrando controlli temporali")
            self._setup_temporal_controls()
            # Mostra i controlli sulla stessa riga
            self.start_year_label.pack(side="left", padx=(20, 5), pady=15)
            self.start_year_selector.pack(side="left", padx=(0, 10), pady=15)
            self.end_year_label.pack(side="left", padx=(10, 5), pady=15)
            self.end_year_selector.pack(side="left", padx=(0, 10), pady=15)
            self.logger.debug(" Controlli temporali mostrati")
        else:
            self.logger.debug("Nascondendo controlli temporali")
            # Nascondi i controlli temporali
            for control in self.temporal_controls:
                control.pack_forget()
        
        self._update_chart()
    
    def _on_temporal_range_changed(self, value):
        """Gestisce il cambio del range temporale"""
        self._update_chart()
    
    def _setup_temporal_controls(self):
        """Configura i controlli per la selezione del range temporale"""
        try:
            # Carica tutti i dati per determinare il range disponibile
            all_data = self.portfolio_manager.load_data()
            if all_data.empty:
                return
            
            # Estrai tutti gli anni disponibili
            created_dates = []
            updated_dates = []
            
            for date_str in all_data['created_at'].dropna():
                parsed = self._parse_single_date(date_str)
                if pd.notna(parsed):
                    created_dates.append(parsed)
            
            for date_str in all_data['updated_at'].dropna():
                parsed = self._parse_single_date(date_str)
                if pd.notna(parsed):
                    updated_dates.append(parsed)
            
            all_dates = created_dates + updated_dates
            if not all_dates:
                return
            
            # Estrai anni unici e ordinali
            years = sorted(set([d.year for d in all_dates]))
            self.available_years = [str(year) for year in years]
            
            # Aggiorna i dropdown
            if hasattr(self, 'start_year_selector'):
                self.start_year_selector.configure(values=self.available_years)
            if hasattr(self, 'end_year_selector'):
                self.end_year_selector.configure(values=self.available_years)
            
            # Imposta valori default se non già impostati
            if not self.start_year.get() or self.start_year.get() not in self.available_years:
                self.start_year.set(self.available_years[0])
            if not self.end_year.get() or self.end_year.get() not in self.available_years:
                self.end_year.set(self.available_years[-1])
                
        except Exception as e:
            self.logger.error(f"Errore nella configurazione controlli temporali: {e}")
    
    def _update_chart(self):
        """Aggiorna il grafico corrente"""
        # Pulisce il grafico precedente
        for widget in self.chart_frame.winfo_children():
            safe_execute(lambda: widget.destroy())
        
        try:
            # Aggiorna label filtri nei controlli
            self._update_filter_label()
            # Carica i dati (solo asset più recenti per coerenza)
            if isinstance(self._external_filtered_df, pd.DataFrame) and not self._external_filtered_df.empty:
                df = self._external_filtered_df.copy()
            else:
                df = self.portfolio_manager.get_current_assets_only()
            
            if df.empty:
                self._show_no_data_message()
                return
            
            # Crea il grafico in base al tipo selezionato
            chart_type = self.chart_type.get()
            
            if chart_type == "Distribuzione Valore per Categoria":
                self._create_value_distribution_chart(df)
            elif chart_type == "Distribuzione Rischio":
                self._create_risk_distribution_chart(df)
            elif chart_type == "Performance per Categoria":
                self._create_performance_chart(df)
            elif chart_type == "Evoluzione Temporale":
                self._create_temporal_chart(df)
            elif chart_type == "Suddivisione Asset per Posizione":
                self._create_position_distribution_chart(df)
            
        except Exception as e:
            self._show_error_message(f"Errore nella creazione del grafico: {e}")
    
    def _create_value_distribution_chart(self, df: pd.DataFrame, ax=None):
        """Crea grafico a torta della distribuzione valore per categoria

        Args:
            df: DataFrame con i dati del portfolio
            ax: Asse matplotlib opzionale. Se None, crea una nuova figura.
        """
        try:
            # Calcola i valori per categoria usando la stessa logica del Portfolio Summary
            df['current_value'] = df['updated_total_value'].fillna(df['created_total_value']).fillna(0)
            category_values = df.groupby('category')['current_value'].sum()
            
            # Filtra categorie con valore > 0
            category_values = category_values[category_values > 0]
            
            # Ordina le categorie secondo l'ordine specificato
            category_order = ['Immobiliare', 'Titoli di stato', 'Fondi di investimento', 'Liquidità', 'Criptovalute', 'ETF', 'PAC']
            
            # Riordina mantenendo solo le categorie presenti nei dati
            ordered_categories = [cat for cat in category_order if cat in category_values.index]
            # Aggiungi eventuali categorie non nell'ordine specificato
            for cat in category_values.index:
                if cat not in ordered_categories:
                    ordered_categories.append(cat)
            
            category_values = category_values.reindex(ordered_categories)
            
            if category_values.empty:
                if ax is None:
                    self._show_no_data_message("Nessun valore da visualizzare")
                return

            # Crea il grafico
            if ax is None:
                fig, ax = plt.subplots(figsize=(10, 8))
                external_ax = False
            else:
                fig = ax.figure
                external_ax = True
            
            # Colori personalizzati
            colors = plt.cm.Set3(np.linspace(0, 1, len(category_values)))
            
            # Calcola le percentuali per la legenda
            total_value = category_values.sum()
            percentages = (category_values / total_value * 100).round(2)
            
            wedges, texts, autotexts = ax.pie(
                category_values.values,
                labels=category_values.index,
                autopct='%1.2f%%',
                colors=colors,
                startangle=90,
                textprops={'fontsize': 10},
                pctdistance=0.7  # Posiziona le percentuali più vicino al centro
            )
            
            # Migliora l'aspetto del testo e sfasa le percentuali lungo il raggio
            for i, (autotext, wedge) in enumerate(zip(autotexts, wedges)):
                autotext.set_color('black')
                autotext.set_weight('bold')
                
                # Calcola l'angolo del wedge per sfasare la percentuale
                angle = (wedge.theta2 + wedge.theta1) / 2
                
                # Sfasa la distanza radiale in base all'indice per evitare sovrapposizioni
                base_distance = 0.6  # Più vicino al centro
                offset = (i % 3) * 0.08  # Offset più piccolo
                pct_distance = base_distance + offset
                
                # Calcola nuova posizione
                x = pct_distance * np.cos(np.radians(angle))
                y = pct_distance * np.sin(np.radians(angle))
                autotext.set_position((x, y))
            
            ax.set_title("Distribuzione Valore per Categoria", 
                        fontsize=14, fontweight='bold', pad=20)
            
            # Aggiungi legenda con percentuali, nomi e valori
            legend_labels = [f"{percentages[cat]:.2f}% - {cat}: €{val:,.0f}" 
                           for cat, val in category_values.items()]
            ax.legend(wedges, legend_labels, title="Categorie",
                     loc="center left", bbox_to_anchor=(1.15, 0, 0.5, 1))

            # Mostra il grafico solo se non è un asse esterno
            if not external_ax:
                self._display_chart(fig)
            else:
                fig.canvas.draw_idle()
            
        except Exception as e:
            self._show_error_message(f"Errore nel grafico distribuzione valore: {e}")
    
    def _create_risk_distribution_chart(self, df: pd.DataFrame, ax=None):
        """Crea grafico a barre della distribuzione del rischio

        Args:
            df: DataFrame con i dati del portfolio
            ax: Asse matplotlib opzionale. Se None, crea una nuova figura.
        """
        try:
            # Conta gli asset per livello di rischio
            risk_counts = df['risk_level'].value_counts().sort_index()
            
            if risk_counts.empty:
                if ax is None:
                    self._show_no_data_message("Nessun dato di rischio da visualizzare")
                return

            # Crea il grafico
            if ax is None:
                fig, ax = plt.subplots(figsize=(10, 6))
                external_ax = False
            else:
                fig = ax.figure
                external_ax = True
            
            # Colori dal verde (basso rischio) al rosso (alto rischio)
            colors = ['#16a34a', '#84cc16', '#eab308', '#f97316', '#dc2626']
            bar_colors = [colors[min(int(level)-1, 4)] for level in risk_counts.index]
            
            bars = ax.bar(risk_counts.index, risk_counts.values, color=bar_colors, alpha=0.8)
            
            # Personalizzazione
            ax.set_xlabel("Livello di Rischio", fontsize=12, fontweight='bold')
            ax.set_ylabel("Numero di Asset", fontsize=12, fontweight='bold')
            ax.set_title("Distribuzione del Rischio", fontsize=14, fontweight='bold', pad=20)
            ax.grid(True, alpha=0.3, axis='y')
            
            # Aggiungi etichette sulle barre
            for bar, count in zip(bars, risk_counts.values):
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                       f'{count}', ha='center', va='bottom', fontweight='bold')
            
            # Imposta i tick dell'asse x
            ax.set_xticks(risk_counts.index)
            ax.set_xticklabels([f"Livello {int(x)}" for x in risk_counts.index])

            plt.tight_layout()

            # Mostra il grafico solo se non è un asse esterno
            if not external_ax:
                self._display_chart(fig)
            else:
                fig.canvas.draw_idle()

        except Exception as e:
            self._show_error_message(f"Errore nel grafico distribuzione rischio: {e}")
    
    def _create_performance_chart(self, df: pd.DataFrame, ax=None):
        """Crea grafico a barre delle performance per categoria

        Args:
            df: DataFrame con i dati del portfolio
            ax: Asse matplotlib opzionale. Se None, crea una nuova figura.
        """
        try:
            # Calcola le performance (differenza tra updated e created)
            df['created_value'] = df['created_total_value'].fillna(0)
            df['updated_value'] = df['updated_total_value'].fillna(df['created_total_value']).fillna(0)
            df['performance'] = df['updated_value'] - df['created_value']
            
            # Raggruppa per categoria
            category_performance = df.groupby('category')['performance'].sum()
            
            # Filtra categorie con performance significativa
            category_performance = category_performance[category_performance.abs() > 1]
            
            if category_performance.empty:
                if ax is None:
                    self._show_no_data_message("Nessuna performance da visualizzare")
                return

            # Crea il grafico
            if ax is None:
                fig, ax = plt.subplots(figsize=(12, 6))
                external_ax = False
            else:
                fig = ax.figure
                external_ax = True
            
            # Colori: verde per guadagni, rosso per perdite
            colors = ['#16a34a' if x >= 0 else '#dc2626' for x in category_performance.values]
            
            bars = ax.bar(range(len(category_performance)), category_performance.values, 
                         color=colors, alpha=0.8)
            
            # Personalizzazione
            ax.set_xlabel("Categoria", fontsize=12, fontweight='bold')
            ax.set_ylabel("Performance (€)", fontsize=12, fontweight='bold')
            ax.set_title("Performance per Categoria", fontsize=14, fontweight='bold', pad=20)
            ax.grid(True, alpha=0.3, axis='y')
            ax.axhline(y=0, color='black', linestyle='-', alpha=0.5)
            
            # Etichette asse x
            ax.set_xticks(range(len(category_performance)))
            ax.set_xticklabels(category_performance.index, rotation=45, ha='right')
            
            # Etichette sulle barre
            for bar, value in zip(bars, category_performance.values):
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., 
                       height + (50 if height >= 0 else -100),
                       f'€{value:,.0f}', ha='center', 
                       va='bottom' if height >= 0 else 'top',
                       fontweight='bold', fontsize=9)

            plt.tight_layout()

            # Mostra il grafico solo se non è un asse esterno
            if not external_ax:
                self._display_chart(fig)
            else:
                fig.canvas.draw_idle()

        except Exception as e:
            self._show_error_message(f"Errore nel grafico performance: {e}")
    
    def _create_temporal_chart(self, df: pd.DataFrame, ax=None):
        """Crea grafico dell'evoluzione del patrimonio nel tempo per categoria

        Args:
            df: DataFrame con i dati del portfolio
            ax: Asse matplotlib opzionale. Se None, crea una nuova figura.
        """
        try:
            # Carica TUTTI i dati SENZA deduplicazione per mantenere la storia
            all_data = self.portfolio_manager.load_data()

            # Applica direttamente i filtri di colonna alla storia completa, se disponibili
            try:
                if isinstance(self._filter_info, dict):
                    col_filters = self._filter_info.get('column_filters')
                    if col_filters:
                        before = len(all_data)
                        all_data = apply_global_filters(all_data, col_filters)
                        self.logger.debug(f"Filtro colonne timeline: {before} -> {len(all_data)} righe")
            except Exception as _e:
                pass

            # Limita la storia agli asset selezionati: usa il DF filtrato passato
            try:
                if isinstance(df, pd.DataFrame) and not df.empty:
                    # Normalizzazione per confronto robusto
                    def norm(series):
                        return series.fillna('').astype(str).str.strip().str.lower()

                    sel = df.copy()
                    sel['asset_key'] = (
                        norm(sel['category']) + '|' +
                        norm(sel['asset_name']) + '|' +
                        norm(sel['position']) + '|' +
                        norm(sel['isin'])
                    )

                    # Prepara la stessa chiave su all_data e filtra
                    all_data = all_data.copy()
                    all_data['asset_key'] = (
                        norm(all_data['category']) + '|' +
                        norm(all_data['asset_name']) + '|' +
                        norm(all_data['position']) + '|' +
                        norm(all_data['isin'])
                    )
                    selected_keys = set(sel['asset_key'].unique())
                    before = len(all_data)
                    all_data = all_data[all_data['asset_key'].isin(selected_keys)].copy()
                    self.logger.debug(f"Filtro timeline per selezione: {len(selected_keys)} chiavi, {before} -> {len(all_data)} righe")
            except Exception as _e:
                # In caso di problemi con il filtro, prosegui con tutti i dati
                pass
            
            # VERIFICA: Confronta con il file Excel direttamente
            try:
                import pandas as pd
                excel_data = pd.read_excel(self.portfolio_manager.excel_file)
                excel_records = len(excel_data)
                loaded_records = len(all_data)
                
                self.logger.debug("CONFRONTO CARICAMENTO:")
                self.logger.debug(f"  - Record nel file Excel: {excel_records}")
                self.logger.debug(f"  - Record caricati in memoria: {loaded_records}")
                self.logger.debug(f"  - Tutti i record caricati: {'✓ SÌ' if excel_records == loaded_records else '✗ NO'}")
                
                if excel_records != loaded_records:
                    self.logger.debug(f"  - ATTENZIONE: Mancano {excel_records - loaded_records} record!")
                    self.logger.debug(f"  - ID nel Excel: {sorted(excel_data['id'].tolist())}")
                    self.logger.debug(f"  - ID caricati: {sorted(all_data['id'].tolist())}")
                    
                    # Trova record mancanti
                    excel_ids = set(excel_data['id'].tolist())
                    loaded_ids = set(all_data['id'].tolist())
                    missing_ids = excel_ids - loaded_ids
                    if missing_ids:
                        self.logger.debug(f"  - Record mancanti (ID): {sorted(missing_ids)}")
                        
                        # Mostra info sui record mancanti
                        for missing_id in sorted(missing_ids):
                            missing_record = excel_data[excel_data['id'] == missing_id]
                            if not missing_record.empty:
                                created_at = missing_record['created_at'].iloc[0]
                                category = missing_record['category'].iloc[0] 
                                asset_name = missing_record['asset_name'].iloc[0]
                                self.logger.debug(f"    ID {missing_id}: {created_at} | {category} | {asset_name}")
                
            except Exception as e:
                self.logger.debug(f" Errore nel confronto con Excel: {e}")
            
            self.logger.debug(f"Record caricati: {len(all_data)}")
            
            if all_data.empty:
                self._show_no_data_message("Nessun dato disponibile")
                return
            
            # 1. RACCOLTA DI TUTTE LE DATE UNICHE da created_at e updated_at
            self.logger.debug(f" Analisi formati date:")
            self.logger.debug(f"  - Prime 5 created_at grezze: {all_data['created_at'].head().tolist()}")
            self.logger.debug(f"  - Ultime 5 created_at grezze: {all_data['created_at'].tail().tolist()}")
            
            # PARSING CENTRALIZZATO delle date usando il nuovo sistema
            date_manager = get_date_manager()

            def parse_dates_centralized(date_series, column_name):
                """Parsing ottimizzato usando il sistema centralizzato"""
                valid_dates = []
                failed_count = 0

                for date_value in date_series.dropna():
                    parsed = date_manager.parse_date(date_value, strict=False)
                    if parsed:
                        valid_dates.append(parsed)
                    else:
                        failed_count += 1

                self.logger.debug(f"  - {column_name}: {len(valid_dates)} valide, {failed_count} fallite")
                return pd.Series(valid_dates)

            created_dates = parse_dates_centralized(all_data['created_at'], 'created_at')
            updated_dates = parse_dates_centralized(all_data['updated_at'], 'updated_at')
            
            self.logger.debug(f" Date created_at: {len(created_dates)} valide")
            self.logger.debug(f" Date updated_at: {len(updated_dates)} valide")
            if len(created_dates) > 0:
                self.logger.debug(f" Range created_at: {created_dates.min()} - {created_dates.max()}")
            if len(updated_dates) > 0:
                self.logger.debug(f" Range updated_at: {updated_dates.min()} - {updated_dates.max()}")
            
            # Lista univoca di date senza duplicati
            all_dates = sorted(set(created_dates.tolist() + updated_dates.tolist()))
            all_dates = [d for d in all_dates if not pd.isna(d)]
            
            # Applica filtro temporale se specificato
            if self.start_year.get() and self.end_year.get():
                start_year_int = int(self.start_year.get())
                end_year_int = int(self.end_year.get())
                
                # Filtra le date nel range selezionato
                filtered_dates = []
                for date in all_dates:
                    if start_year_int <= date.year <= end_year_int:
                        filtered_dates.append(date)
                
                all_dates = filtered_dates
                self.logger.debug(f" Filtro temporale applicato: {start_year_int}-{end_year_int}")
            
            self.logger.debug(f" Date univoche totali: {len(all_dates)}")
            self.logger.debug(f" Prima data: {all_dates[0] if all_dates else 'N/A'}")
            self.logger.debug(f" Ultima data: {all_dates[-1] if all_dates else 'N/A'}")
            
            if not all_dates:
                self._show_no_data_message("Nessuna data valida trovata nel range selezionato")
                return
            
            # 2. CALCOLO PATRIMONIO PER OGNI DATA E CATEGORIA
            # Crea chiave univoca per identificare asset duplicati (se non già presente)
            if 'asset_key' not in all_data.columns:
                all_data['asset_key'] = (
                    all_data['category'].fillna('') + '|' +
                    all_data['asset_name'].fillna('') + '|' +
                    all_data['position'].fillna('') + '|' +
                    all_data['isin'].fillna('')
                )
            
            categories = all_data['category'].dropna().unique()
            timeline_data = {}
            
            # Per ogni data nella lista univoca
            for date in all_dates:
                date_values = {}
                
                # Per ogni categoria
                for category in categories:
                    category_value = 0
                    category_assets = all_data[all_data['category'] == category]
                    
                    # Per ogni asset unico di questa categoria
                    for asset_key in category_assets['asset_key'].unique():
                        asset_records = category_assets[category_assets['asset_key'] == asset_key]
                        
                        # Trova il record più appropriato per questa data
                        valid_records = []
                        for _, asset in asset_records.iterrows():
                            # Usa parsing flessibile anche qui
                            created_date = self._parse_single_date(asset['created_at'])
                            updated_date = self._parse_single_date(asset['updated_at'])
                            
                            # L'asset esiste solo se è già stato creato a questa data
                            if pd.notna(created_date) and date >= created_date:
                                # Determinare quale valore usare in base alla data
                                if pd.notna(updated_date) and date >= updated_date:
                                    # Usa valore aggiornato
                                    amount = asset['updated_amount'] if pd.notna(asset['updated_amount']) else 0
                                    price = asset['updated_unit_price'] if pd.notna(asset['updated_unit_price']) else 0
                                    record_date = updated_date
                                else:
                                    # Usa valore iniziale
                                    amount = asset['created_amount'] if pd.notna(asset['created_amount']) else 0
                                    price = asset['created_unit_price'] if pd.notna(asset['created_unit_price']) else 0
                                    record_date = created_date
                                
                                value = amount * price
                                valid_records.append((record_date, value))
                        
                        # Prendi l'ultimo valore valido per questo asset unico a questa data
                        if valid_records:
                            # Ordina per data e prendi il più recente
                            valid_records.sort(key=lambda x: x[0])
                            latest_value = valid_records[-1][1]
                            category_value += latest_value
                    
                    date_values[category] = category_value
                
                timeline_data[date] = date_values
            
            # DEBUG: Salva tabella debug nel file Excel
            self._save_debug_table_to_excel(timeline_data, categories)
            
            # VERIFICA: Confronta l'ultimo valore totale con quello della navbar
            self._verify_total_value(timeline_data)
            
            # 3. PREPARAZIONE DATI PER IL GRAFICO
            dates = sorted(timeline_data.keys())
            
            # Crea DataFrame per il grafico
            chart_data = pd.DataFrame(timeline_data).T
            chart_data = chart_data.fillna(0)

            # 4. CREAZIONE DEL GRAFICO
            # Se ax è fornito dall'esterno, usalo; altrimenti crea nuova figura
            if ax is None:
                fig, ax = plt.subplots(figsize=(16, 10))
                external_ax = False
            else:
                fig = ax.figure
                external_ax = True

            # Pulisci l'asse per evitare sovrapposizioni
            ax.clear()
            
            # Colori per le categorie
            colors = plt.cm.Set3(np.linspace(0, 1, len(categories)))
            
            # Una linea curva per ogni categoria
            lines_added = []
            for i, category in enumerate(categories):
                if category in chart_data.columns:
                    # Filtra solo le date dove la categoria ha valore > 0
                    category_data = chart_data[category]
                    category_data = category_data[category_data > 0]
                    
                    if len(category_data) > 0:
                        # Linee con smoothing semplice usando interpolazione lineare con più punti
                        if len(category_data) > 2:
                            # Aggiungi punti intermedi per smoothing visivo semplice
                            x_dates = category_data.index
                            y_values = category_data.values
                            
                            # Linea semplice con interpolazione numpy - primo piano
                            ax.plot(x_dates, y_values, 
                                   color=colors[i], linewidth=2.5, alpha=0.9, label=category,
                                   antialiased=True, marker='o', markersize=6, zorder=3)
                            
                            # Markers sui punti originali - primo piano
                            ax.scatter(x_dates, y_values, color=colors[i], s=40, zorder=5,
                                     edgecolors='white', linewidths=1.5, alpha=1.0)
                        else:
                            # Linea normale per pochi punti - primo piano
                            ax.plot(category_data.index, category_data.values, 
                                   marker='o', label=category, color=colors[i], 
                                   linewidth=2.5, markersize=5, alpha=0.9,
                                   markerfacecolor=colors[i], markeredgecolor='white', markeredgewidth=1,
                                   antialiased=True, zorder=3)
                        
                        lines_added.append(f"Categoria: {category}")
            
            # UNA SOLA linea curva del totale
            total_values = chart_data.sum(axis=1)
            # Mostra totale solo dove c'è almeno una categoria con valore
            total_values = total_values[total_values > 0]
            
            self.logger.debug("GRAFICO:")
            self.logger.debug(f"  - Linee categorie: {len(lines_added)}")
            self.logger.debug(f"  - Punti totale: {len(total_values)}")
            self.logger.debug(f"  - Valore finale totale: €{total_values.iloc[-1]:,.2f}" if len(total_values) > 0 else "  - Nessun valore totale")
            
            if len(total_values) > 0:
                # Linea del totale semplice
                if len(total_values) > 2:
                    x_dates_total = total_values.index
                    y_values_total = total_values.values
                    
                    # Disegna linea totale più spessa sullo sfondo
                    ax.plot(x_dates_total, y_values_total, 
                           color='black', linewidth=5.0, alpha=0.7, label='TOTALE',
                           antialiased=True, marker='s', markersize=8, zorder=1)
                    
                    # Markers sui punti originali del totale - sopra le linee categorie
                    ax.scatter(x_dates_total, y_values_total, color='black', s=60, marker='s', zorder=4,
                             edgecolors='white', linewidths=2, alpha=1.0)
                else:
                    # Linea normale per pochi punti - più spessa e sullo sfondo
                    ax.plot(total_values.index, total_values.values, 
                           marker='s', label='TOTALE', color='black', 
                           linewidth=5.0, markersize=7, alpha=0.7, zorder=1,
                           markerfacecolor='black', markeredgecolor='white', markeredgewidth=1.5,
                           antialiased=True)
                
                lines_added.append("TOTALE")
            
            self.logger.debug(f"  - Linee totali nel grafico: {lines_added}")
            
            # 5. FORMATTAZIONE DEL GRAFICO
            ax.set_title('Evoluzione Patrimonio per Categoria', fontsize=16, fontweight='bold', pad=20)
            ax.set_xlabel('Anno', fontsize=12)
            ax.set_ylabel('Valore (€)', fontsize=12)
            
            # Seleziona solo primo, ultimo e anni pari per evitare sovrapposizioni
            years = sorted(set([d.year for d in dates]))
            if len(years) > 0:
                # Determina quali anni mostrare: primo, ultimo e anni pari
                first_year = min(years)
                last_year = max(years)
                
                display_years = [first_year]  # Sempre il primo anno
                
                # Genera tutti gli anni pari nell'intervallo completo (non solo quelli presenti nei dati)
                for year in range(first_year, last_year + 1):
                    if year % 2 == 0 and year != first_year:
                        display_years.append(year)
                
                # Sempre l'ultimo anno (se non già presente)
                if last_year not in display_years:
                    display_years.append(last_year)
                
                # Ordina gli anni da mostrare
                display_years = sorted(set(display_years))
                
                # Crea i tick e le etichette
                year_dates = [pd.Timestamp(f'{year}-01-01') for year in display_years]
                ax.set_xticks(year_dates)
                ax.set_xticklabels(display_years)
                
                # Assicurati che tutti i tick siano visibili
                ax.tick_params(axis='x', rotation=0)
            
            # Formattazione asse Y con valuta
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'€{x:,.0f}'))
            
            # Griglia e legenda
            ax.grid(True, alpha=0.3)
            ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')

            # Mostra il grafico solo se non è un asse esterno
            if not external_ax:
                self._display_chart(fig)
            else:
                # Se è un asse esterno, aggiorna solo la figura
                fig.canvas.draw_idle()

        except Exception as e:
            self._show_error_message(f"Errore nella creazione del grafico temporale: {e}")

    def _create_position_distribution_chart(self, df: pd.DataFrame, ax=None):
        """Crea grafico a torta della suddivisione valore per posizione

        Args:
            df: DataFrame con i dati del portfolio
            ax: Asse matplotlib opzionale. Se None, crea una nuova figura.
        """
        try:
            if 'position' not in df.columns:
                self._show_no_data_message("Campo 'position' non presente nei dati")
                return

            # Prepara dati con posizione normalizzata
            df_work = df.copy()
            df_work['position'] = df_work['position'].fillna('Non specificata')

            # Calcola il valore corrente per ogni asset (come nel Portfolio Summary)
            df_work['current_value'] = df_work['updated_total_value'].fillna(df_work['created_total_value']).fillna(0)

            # Calcola VALORE TOTALE per posizione (per il grafico)
            position_values = df_work.groupby('position')['current_value'].sum()

            # Calcola NUMERO di asset per posizione (per la legenda)
            position_counts = df_work.groupby('position').size()

            # Filtra posizioni con valore > 0
            position_values = position_values[position_values > 0]

            if position_values.empty:
                if ax is None:
                    self._show_no_data_message("Nessun valore da visualizzare")
                return

            # Crea il grafico
            if ax is None:
                fig, ax = plt.subplots(figsize=(10, 8))
                external_ax = False
            else:
                fig = ax.figure
                external_ax = True

            # Colori personalizzati
            colors = plt.cm.Set3(np.linspace(0, 1, len(position_values)))

            # Calcola le percentuali sul VALORE (per il grafico)
            total_value = position_values.sum()
            value_percentages = (position_values / total_value * 100).round(2)

            wedges, texts, autotexts = ax.pie(
                position_values.values,
                labels=position_values.index,
                autopct='%1.2f%%',
                colors=colors,
                startangle=90,
                textprops={'fontsize': 10},
                pctdistance=0.7
            )

            # Migliora leggibilità delle etichette
            for text in texts:
                text.set_fontsize(11)
                text.set_weight('bold')

            for autotext in autotexts:
                autotext.set_color('black')
                autotext.set_fontsize(9)
                autotext.set_weight('bold')

            # Titolo
            ax.set_title(
                'Suddivisione Valore per Posizione',
                fontsize=14,
                weight='bold',
                pad=20
            )

            # Legenda con numero di asset, valore totale e percentuale
            legend_labels = [
                f'{pos}: {position_counts[pos]} asset, €{position_values[pos]:,.2f} ({value_percentages[pos]}%)'
                for pos in position_values.index
            ]
            ax.legend(
                legend_labels,
                loc='center left',
                bbox_to_anchor=(1, 0, 0.5, 1),
                fontsize=9
            )

            plt.tight_layout()

            # Mostra il grafico solo se non è un asse esterno
            if not external_ax:
                self._display_chart(fig)
            else:
                fig.canvas.draw_idle()

        except Exception as e:
            self._show_error_message(f"Errore nella creazione del grafico posizione: {e}")

    def _save_debug_table_to_excel(self, timeline_data, categories):
        """Salva tabella debug con date e categorie nel file Excel"""
        try:
            import openpyxl
            from openpyxl import Workbook, load_workbook
            import os
            
            excel_file = self.portfolio_manager.excel_file
            
            # Carica il workbook esistente o ne crea uno nuovo
            if os.path.exists(excel_file):
                wb = load_workbook(excel_file)
            else:
                wb = Workbook()
            
            # Rimuovi foglio debug se esiste
            if 'Debug_Timeline' in wb.sheetnames:
                del wb['Debug_Timeline']
            
            # Crea nuovo foglio debug
            ws_debug = wb.create_sheet('Debug_Timeline')
            
            # Prepara i dati per il foglio
            dates = sorted(timeline_data.keys())
            
            # Header: Data | Categoria1 | Categoria2 | ... | TOTALE
            headers = ['Data'] + list(categories) + ['TOTALE']
            ws_debug.append(headers)
            
            # Righe di dati
            for date in dates:
                row = [date.strftime('%Y-%m-%d')]
                date_total = 0
                
                for category in categories:
                    value = timeline_data[date].get(category, 0)
                    row.append(value)
                    date_total += value
                
                row.append(date_total)
                ws_debug.append(row)
            
            # Formattazione
            # Intestazioni in grassetto
            for cell in ws_debug[1]:
                cell.font = openpyxl.styles.Font(bold=True)
            
            # Formato valuta per le colonne dei valori
            from openpyxl.styles import NamedStyle
            currency_style = NamedStyle(name="currency", number_format="€#,##0.00")
            
            for row in ws_debug.iter_rows(min_row=2, min_col=2):
                for cell in row:
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.style = currency_style
            
            # Auto-width per le colonne
            for column in ws_debug.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws_debug.column_dimensions[column_letter].width = min(max_length + 2, 20)
            
            # Salva il file
            wb.save(excel_file)
            self.logger.debug(f" Salvata tabella debug nel foglio 'Debug_Timeline' di {excel_file}")
            
        except Exception as e:
            self.logger.error(f"Errore nel salvare tabella debug: {e}")
    
    def _verify_total_value(self, timeline_data):
        """Verifica che l'ultimo valore totale corrisponda a quello della navbar"""
        try:
            if not timeline_data:
                return
                
            # Calcola il totale dell'ultima data
            last_date = max(timeline_data.keys())
            last_values = timeline_data[last_date]
            chart_total = sum(last_values.values())
            
            # Ottieni il valore dalla navbar (get_portfolio_summary)
            navbar_summary = self.portfolio_manager.get_portfolio_summary()
            navbar_total = navbar_summary['total_value']
            
            self.logger.debug("VERIFICA:")
            self.logger.debug(f"  - Data più recente: {last_date.strftime('%Y-%m-%d')}")
            self.logger.debug(f"  - Totale grafico: €{chart_total:,.2f}")
            self.logger.debug(f"  - Totale navbar: €{navbar_total:,.2f}")
            self.logger.debug(f"  - Differenza: €{abs(chart_total - navbar_total):,.2f}")
            self.logger.debug(f"  - Corrispondenza: {'✓ SÌ' if abs(chart_total - navbar_total) < 0.01 else '✗ NO'}")
            
        except Exception as e:
            self.logger.error(f"Errore nella verifica del valore totale: {e}")
    
    def _parse_single_date(self, date_str):
        """Parsing flessibile di una singola data"""
        if pd.isna(date_str) or date_str == '':
            return pd.NaT
            
        original_str = str(date_str)
        
        # Prova diversi formati
        formats_to_try = [
            '%Y-%m-%d',     # 2004-05-21
            '%d-%m-%Y',     # 21-05-2004  
            '%d/%m/%Y',     # 21/05/2004
            '%Y/%m/%d',     # 2004/05/21
            '%m/%d/%Y',     # 05/21/2004
            '%d.%m.%Y',     # 21.05.2004
            '%Y.%m.%d'      # 2004.05.21
        ]
        
        # Usa il sistema centralizzato per parsing
        date_manager = get_date_manager()
        parsed = date_manager.parse_date(date_str, strict=False)
        if parsed:
            return pd.Timestamp(parsed)
        return pd.NaT
    
    def _display_chart(self, fig):
        """Visualizza il grafico nel frame"""
        try:
            plt.tight_layout()
            
            # Crea il canvas per CustomTkinter
            canvas = FigureCanvasTkAgg(fig, self.chart_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)
            
            # Salva riferimento per cleanup
            self.current_chart = canvas
            
        except Exception as e:
            self._show_error_message(f"Errore nella visualizzazione: {e}")
        finally:
            # Chiudi la figure per liberare memoria
            plt.close(fig)
    
    def _show_no_data_message(self, message: str = "Nessun dato disponibile per i grafici"):
        """Mostra un messaggio quando non ci sono dati"""
        label = ctk.CTkLabel(
            self.chart_frame,
            text=message,
            font=ctk.CTkFont(**UIConfig.FONTS['header']),
            text_color=UIConfig.COLORS['secondary']
        )
        label.pack(expand=True)
    
    def _show_error_message(self, message: str):
        """Mostra un messaggio di errore"""
        label = ctk.CTkLabel(
            self.chart_frame,
            text=f"❌ {message}",
            font=ctk.CTkFont(**UIConfig.FONTS['text']),
            text_color=UIConfig.COLORS['danger']
        )
        label.pack(expand=True)
    
    def refresh_charts(self):
        """Aggiorna i grafici con i dati più recenti"""
        safe_execute(self._update_chart)

    def refresh_with_filtered_data(self, df: pd.DataFrame, filter_info: Optional[Dict[str, Any]] = None):
        """Aggiorna i grafici usando un DataFrame filtrato e opzionalmente info filtri."""
        try:
            self._external_filtered_df = df.copy() if isinstance(df, pd.DataFrame) else None
            if isinstance(filter_info, dict):
                self._filter_info = filter_info
        except Exception:
            self._external_filtered_df = None
        finally:
            self.refresh_charts()

    def cleanup(self):
        """Pulisce le risorse utilizzate dai grafici"""
        if self.current_chart:
            safe_execute(lambda: self.current_chart.get_tk_widget().destroy())
        plt.close('all')

    def _format_filter_summary(self) -> Optional[str]:
        """Crea una stringa leggibile con i filtri attivi.
        Formato: Campo: valore1, valore2, valore3
        Ogni filtro su una riga separata."""
        try:
            info = self._filter_info or {}
            col_filters = info.get('column_filters') or {}

            # Nessun filtro attivo - non mostrare nulla
            if not col_filters:
                return None

            from config import FieldMapping
            lines = []

            for col, values in col_filters.items():
                disp = FieldMapping.DB_TO_DISPLAY.get(col, col)
                vals = list(sorted({str(v) for v in values}))
                shown = ', '.join(vals)
                lines.append(f"{disp}: {shown}")

            return '\n'.join(lines)
        except Exception:
            return None

    def _update_filter_label(self):
        """Aggiorna la label dei filtri nei controlli."""
        if not hasattr(self, 'filter_label') or not self.filter_label:
            return

        summary = self._format_filter_summary()
        if summary:
            self.filter_label.configure(text=summary)
        else:
            self.filter_label.configure(text="")



