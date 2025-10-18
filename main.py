#!/usr/bin/env python3
"""
GAB AssetMind - Portfolio Management Application (Refactored Version)
Applicazione per la gestione professionale di portafogli diversificati con interfaccia moderna

Architettura modulare con separazione delle responsabilità:
- UI Components: Componenti interfaccia separati e riutilizzabili  
- Configuration: Configurazione centralizzata per colori, dimensioni, mappature
- Utils: Utilità per validazione, formattazione, gestione errori
- Models: Logica di business e persistenza dati

Autore: GAB AssetMind Team
Versione: 2.0 (Refactored)
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox, filedialog
import sys
import os
import glob
import threading
from typing import Optional, Dict, Any, List
import time
import pandas as pd

# Import dei moduli refactored
from config import UIConfig, DatabaseConfig, Messages, get_application_directory
from utils import ErrorHandler, DataCache, safe_execute
from models import PortfolioManager, apply_global_filters
from market_data import MarketDataService, MarketDataError
from ui_components import NavigationBar, PortfolioTable
from asset_form import AssetForm
from charts_ui import ChartsUI
from export_ui import ExportUI
from home_ui import RoadMapDashboard
from logging_config import get_logger, set_debug_mode
from security_validation import PathSecurityValidator, SecurityError

# Configurazione tema dell'applicazione
ctk.set_appearance_mode("light")  # Modalità chiara
ctk.set_default_color_theme("blue")  # Tema blu

class GABAssetMind:
    """
    Applicazione principale per la gestione portfolio (Versione Refactored)
    
    Architettura modulare con componenti specializzati:
    - NavigationBar: Barra navigazione e controlli globali
    - PortfolioTable: Tabella portfolio con filtri avanzati  
    - AssetForm: Form gestione asset con validazione
    - ChartsUI: Interfaccia grafici e analytics
    - ExportUI: Esportazione dati in vari formati
    """
    
    def __init__(self):
        """Inizializza l'applicazione con architettura modulare"""
        # Finestra principale
        self.root = ctk.CTk()
        self.root.title("GAB AssetMind - Portfolio Manager (Refactored)")
        self.root.geometry(UIConfig.WINDOW_SIZES['main'])
        
        # Sistema di gestione dati
        self.portfolio_manager: Optional[PortfolioManager] = None
        self.data_cache = DataCache()
        self.current_portfolio_file = DatabaseConfig.DEFAULT_PORTFOLIO_FILE
        self.path_validator = PathSecurityValidator()
        
        # Stato applicazione
        self.current_page = "RoadMap"
        self.page_frames: Dict[str, ctk.CTkFrame] = {}
        self.active_filter_popup: Optional[tk.Toplevel] = None
        
        # Componenti UI specializzati
        self.roadmap_dashboard: Optional[RoadMapDashboard] = None
        self.navbar: Optional[NavigationBar] = None
        self.portfolio_table: Optional[PortfolioTable] = None
        self.asset_form: Optional[AssetForm] = None
        self.charts_ui: Optional[ChartsUI] = None
        self.export_ui: Optional[ExportUI] = None

        # Stato filtri globale (semplice, in memoria)
        self.filter_state: Dict[str, Any] = {
            'show_all_records': False,
            'column_filters': {}
        }
        self._last_filtered_df: Optional[pd.DataFrame] = None

        # Inizializzazione
        self._initialize_portfolio_system()
        self._setup_ui()
        self._setup_callbacks()
        self._load_initial_data()
        
        # Centra la finestra dopo aver configurato tutto
        self._center_window()
    
    def _initialize_portfolio_system(self):
        """Inizializza il sistema di gestione portfolio"""
        try:
            app_dir = get_application_directory()
            default_path = os.path.join(app_dir, self.current_portfolio_file)
            self.portfolio_manager = PortfolioManager(default_path)
            self.logger = get_logger('main')
            self.logger.info(f"Portfolio system initialized: {default_path}")
            
        except Exception as e:
            error_msg = ErrorHandler.handle_file_error(e, "inizializzazione sistema portfolio")
            messagebox.showerror("Errore Critico", error_msg)
            sys.exit(1)
    
    def _setup_ui(self):
        """Configura l'interfaccia utente modulare"""
        try:
            # Navbar globale
            self.navbar = NavigationBar(self.root, self.portfolio_manager)
            navbar_frame = self.navbar.create_navbar()
            
            # Container per le pagine
            self.main_container = ctk.CTkFrame(self.root)
            self.main_container.pack(fill="both", expand=True, padx=10, pady=(0, 10))
            
            # Crea le pagine
            self._create_page_frames()
            self._setup_specialized_components()
            
            # Aggiorna lista portfolio
            self._refresh_portfolio_list()
            
            # Mostra pagina di default
            self.show_page("RoadMap")
            
        except Exception as e:
            error_msg = ErrorHandler.handle_ui_error(e, "setup interfaccia")
            messagebox.showerror("Errore UI", error_msg)
    
    def _create_page_frames(self):
        """Crea i frame per le diverse pagine"""
        page_names = ["RoadMap", "Portfolio", "Asset", "Grafici", "Export"]
        
        for page_name in page_names:
            frame = ctk.CTkFrame(self.main_container)
            self.page_frames[page_name] = frame
    
    def _setup_specialized_components(self):
        """Configura i componenti specializzati per ogni pagina"""
        # Charts UI Component (creato PRIMA del dashboard perché serve per il rendering)
        self.charts_ui = ChartsUI(self.page_frames["Grafici"], self.portfolio_manager)
        self.charts_ui.create_charts_interface()

        # Dashboard Component (passa charts_ui per riutilizzare la stessa logica di rendering)
        self.roadmap_dashboard = RoadMapDashboard(
            self.page_frames["RoadMap"],
            self.portfolio_manager,
            charts_ui_instance=self.charts_ui,
            on_navigate=self._navigate_from_dashboard,
        )
        self.roadmap_dashboard.refresh()

        # Portfolio Table Component
        self.portfolio_table = PortfolioTable(self.page_frames["Portfolio"], self.portfolio_manager)
        self.portfolio_table.create_table()

        # Asset Form Component
        self.asset_form = AssetForm(self.page_frames["Asset"], self.portfolio_manager)
        self.asset_form.create_form()
        
        # Export UI Component
        self.export_ui = ExportUI(self.page_frames["Export"], self.portfolio_manager)
        self.export_ui.create_export_interface()
    
    def _setup_callbacks(self):
        """Configura i callback tra componenti"""
        # Navbar callbacks
        self.navbar.register_callback('page_changed', self.show_page)
        self.navbar.register_callback('portfolio_changed', self._switch_portfolio)
        self.navbar.register_callback('new_portfolio_requested', self._create_new_portfolio)
        self.navbar.register_callback('market_update_requested', self._on_market_update_requested)
        
        # Portfolio Table callbacks
        self.portfolio_table.register_callback('view_changed', self._on_view_changed)
        self.portfolio_table.register_callback('asset_selected', self._on_asset_selected)
        self.portfolio_table.register_callback('filter_requested', self._show_column_filter)
        self.portfolio_table.register_callback('data_filtered', self._on_data_filtered)
        self.portfolio_table.register_callback('data_changed', self._on_data_changed)
        self.portfolio_table.register_callback('market_update_requested', self._on_market_update_requested)
        # Filtri globali
        self.portfolio_table.register_callback('filters_changed', self._on_filters_changed)
        
        # Asset Form callbacks
        self.asset_form.register_callback('asset_saved', self._on_asset_saved)
        self.asset_form.register_callback('asset_deleted', self._on_asset_deleted)
        self.asset_form.register_callback('form_cleared', self._on_form_cleared)
        
        # Export UI callbacks
        self.export_ui.register_callback('export_completed', self._on_export_completed)
    
    def _navigate_from_dashboard(self, page: str, chart_name: Optional[str] = None) -> None:
        """Gestisce la navigazione richiesta dalla dashboard di apertura."""
        if self.navbar:
            self.navbar.navigate_to(page)
        else:
            self.show_page(page)

        if page == "Grafici" and chart_name and self.charts_ui:
            self.charts_ui.select_chart(chart_name)

    def _load_initial_data(self):
        """Carica i dati iniziali dell'applicazione"""
        safe_execute(
            self._load_portfolio_data,
            error_handler=lambda e: print(f"Errore caricamento iniziale: {e}")
        )
        # Aggiorna i valori della navbar dopo il caricamento iniziale
        safe_execute(
            self._update_navbar_values,
            error_handler=lambda e: print(f"Errore aggiornamento navbar iniziale: {e}")
        )
        safe_execute(
            lambda: self._refresh_dashboard(self._last_filtered_df),
            error_handler=lambda e: print(f"Errore aggiornamento dashboard iniziale: {e}")
        )
    
    def show_page(self, page_name: str):
        """Mostra una pagina specifica nascondendo le altre"""
        try:
            # Nascondi tutte le pagine
            for frame in self.page_frames.values():
                frame.pack_forget()
            
            # Mostra la pagina richiesta
            if page_name in self.page_frames:
                self.page_frames[page_name].pack(fill="both", expand=True)
                self.current_page = page_name
                
                # Aggiorna l'evidenziazione del bottone attivo nella navbar
                if self.navbar:
                    self.navbar.update_active_button(page_name)
                
                # Refresh dei dati per pagine specifiche
                if page_name == "RoadMap":
                    self._refresh_dashboard(self._last_filtered_df)
                    # Aggiorna sempre i contatori navbar quando torni alla RoadMap
                    self._update_navbar_values()
                elif page_name == "Portfolio":
                    self._load_portfolio_data()
                elif page_name == "Grafici":
                    self.charts_ui.refresh_charts()
                elif page_name == "Export":
                    self.export_ui.refresh_stats()
                    
        except Exception as e:
            error_msg = ErrorHandler.handle_ui_error(e, f"navigazione pagina {page_name}")
            messagebox.showerror("Errore Navigazione", error_msg)
    
    def _load_portfolio_data(self):
        """Carica i dati del portfolio e aggiorna l'interfaccia"""
        try:
            self.logger.debug(f"Inizio caricamento dati portfolio")
            self.logger.debug(f"Portfolio table exists: {self.portfolio_table is not None}")
            # Carica base DF secondo stato globale e applica filtri globali
            base_show_all = bool(self.filter_state.get('show_all_records'))
            if base_show_all:
                self.logger.debug("Base: tutti i record")
                df_base = self.portfolio_manager.load_data()
            else:
                self.logger.debug("Base: asset correnti")
                df_base = self.portfolio_manager.get_current_assets_only()
            df = apply_global_filters(df_base, self.filter_state.get('column_filters'))
            self._last_filtered_df = df.copy() if isinstance(df, pd.DataFrame) else None
            # Log semplice sui dati caricati
            self.logger.debug(f"Dati caricati: {len(df)} righe")
            if isinstance(df, pd.DataFrame) and not df.empty:
                self.logger.debug(f"Colonne disponibili: {list(df.columns)}")
            
            # Aggiorna componenti
            if self.portfolio_table:
                self.logger.debug(f"Aggiornando portfolio_table con {len(df)} righe")
                self.portfolio_table.update_data(df)
                self.logger.debug("update_data() completato")
                # Refresh UI ottimizzato - solo update_idletasks
                try:
                    if hasattr(self, 'after'):  # Solo se la finestra è inizializzata
                        self.after(50, lambda: [
                            self.update_idletasks(),
                            self.logger.debug("UI refresh completato")
                        ])
                    else:
                        self.logger.debug("UI refresh skipped - finestra non inizializzata")
                except Exception as e:
                    self.logger.error(f"Errore refresh UI: {e}")
            else:
                self.logger.error("PROBLEMA - portfolio_table è None!")
                
            # Aggiorna componenti consumatori del DF filtrato
            if self.charts_ui and hasattr(self.charts_ui, 'refresh_with_filtered_data'):
                self.charts_ui.refresh_with_filtered_data(df, self.filter_state)
            if self.export_ui and hasattr(self.export_ui, 'refresh_with_filtered_data'):
                self.export_ui.refresh_with_filtered_data(df, self.filter_state)
            self._update_navbar_values()
            self._refresh_dashboard(df)
            
        except Exception as e:
            error_msg = ErrorHandler.handle_data_error(e, "caricamento portfolio")
            self.logger.error(f"Errore caricamento dati: {error_msg}")
            import traceback
            self.logger.debug(f"Stack trace: {traceback.format_exc()}")

    def _on_filters_changed(self, payload: Dict[str, Any]):
        """Riceve filtri dalla tabella e aggiorna lo stato globale + refresh componenti"""
        try:
            if not isinstance(payload, dict):
                return
            # Aggiorna stato
            if 'column_filters' in payload and isinstance(payload['column_filters'], dict):
                self.filter_state['column_filters'] = payload['column_filters']
            if 'show_all_records' in payload:
                self.filter_state['show_all_records'] = bool(payload['show_all_records'])
            # Ricarica i dati filtrati e aggiorna UI (nav, table, charts, export)
            self._load_portfolio_data()
        except Exception as e:
            self.logger.error(f"Errore gestione filters_changed: {e}")
    
    def _update_navbar_values(self):
        """Aggiorna i valori mostrati nella navbar"""
        try:
            # Calcola valore totale
            summary = self.portfolio_manager.get_portfolio_summary()
            total_value = summary.get('total_value', 0)
            
            # Calcola valore visibile se tabella disponibile
            visible_value = 0
            percentage = 0
            
            if self.portfolio_table:
                visible_value, _ = self.portfolio_table.get_visible_value()
                if total_value > 0:
                    percentage = (visible_value / total_value * 100)
            
            # Aggiorna navbar
            if self.navbar:
                if visible_value == total_value or percentage >= 99.9:
                    percentage = 100.0
                self.navbar.update_values(total_value, visible_value, percentage)
                # Aggiorna contatori semplici: usa i valori dai pulsanti Portfolio (fonte di verità)
                if self.portfolio_table:
                    total_records_count, current_assets_count = self.portfolio_table.get_counts()
                    # Fallback se i valori non sono ancora stati calcolati
                    if total_records_count == 0 and current_assets_count == 0:
                        total_records_count = len(self.portfolio_manager.load_data())
                        current_assets_count = len(self.portfolio_manager.get_current_assets_only())
                        self.logger.debug(f"Navbar contatori (fallback diretto): Records={total_records_count}, Assets={current_assets_count}")
                    else:
                        self.logger.debug(f"Navbar contatori (da Portfolio buttons): Records={total_records_count}, Assets={current_assets_count}")
                else:
                    total_records_count = 0
                    current_assets_count = 0
                self.navbar.update_counts(total_records_count, current_assets_count)
                
        except Exception as e:
            self.logger.error(f"Errore aggiornamento valori navbar: {e}")

    def _refresh_dashboard(self, dataframe: Optional[pd.DataFrame] = None) -> None:
        """Aggiorna i contenuti della dashboard se disponibile."""
        if not self.roadmap_dashboard:
            return
        try:
            summary = self.portfolio_manager.get_portfolio_summary() if self.portfolio_manager else {}
            df_source = dataframe if dataframe is not None else self._last_filtered_df
            if df_source is None and self.portfolio_manager:
                df_source = self.portfolio_manager.get_current_assets_only()
            # Passa anche il filter_state per mostrare la selezione attiva
            self.roadmap_dashboard.refresh(summary, df_source, self.filter_state)
        except Exception as exc:
            if self.logger:
                self.logger.error(f"Errore aggiornamento dashboard: {exc}")

    def _refresh_portfolio_list(self):
        """Aggiorna la lista dei portfolio disponibili"""
        try:
            app_dir = get_application_directory()
            pattern = os.path.join(app_dir, "*.xlsx")
            excel_files = glob.glob(pattern)
            
            portfolio_files = [os.path.basename(f) for f in excel_files]
            if not portfolio_files:
                portfolio_files = [DatabaseConfig.DEFAULT_PORTFOLIO_FILE]
            
            portfolio_files.sort()
            
            if self.navbar:
                self.navbar.refresh_portfolio_list(portfolio_files, self.current_portfolio_file)
                
        except Exception as e:
            error_msg = ErrorHandler.handle_file_error(e, "refresh lista portfolio")
            self.logger.error(f"Errore refresh portfolio: {error_msg}")
    
    def _switch_portfolio(self, selected_file: str):
        """Cambia il portfolio attivo con validazione sicurezza"""
        try:
            if selected_file != self.current_portfolio_file:
                # Valida il path prima di procedere
                try:
                    app_dir = get_application_directory()
                    full_path = os.path.join(app_dir, selected_file)
                    validated_path = self.path_validator.validate_portfolio_path(full_path)
                    self.current_portfolio_file = selected_file
                    self.logger.info(f"Cambio portfolio validato: {validated_path}")
                    # Aggiorna il PortfolioManager con path sicuro
                    self.portfolio_manager = PortfolioManager(str(validated_path))
                except SecurityError as e:
                    self.logger.error(f"Portfolio non sicuro: {e}")
                    messagebox.showerror("Errore Sicurezza", f"Portfolio non sicuro: {e}")
                    return
                except Exception as e:
                    self.logger.error(f"Errore validazione portfolio: {e}")
                    messagebox.showerror("Errore", f"Errore validazione portfolio: {e}")
                    return
                
                # Pulisce cache e ricarica
                self.data_cache.clear()
                
                # Aggiorna componenti con nuovo portfolio manager
                if self.portfolio_table:
                    self.portfolio_table.portfolio_manager = self.portfolio_manager
                if self.asset_form:
                    self.asset_form.portfolio_manager = self.portfolio_manager
                if self.charts_ui:
                    self.charts_ui.portfolio_manager = self.portfolio_manager
                if self.export_ui:
                    self.export_ui.portfolio_manager = self.portfolio_manager
                if self.roadmap_dashboard:
                    self.roadmap_dashboard.set_portfolio_manager(self.portfolio_manager)
                
                # Ricarica dati
                self._load_portfolio_data()
                
                self.logger.info(f"Portfolio cambiato a: {selected_file}")
                
        except Exception as e:
            error_msg = ErrorHandler.handle_file_error(e, f"cambio portfolio {selected_file}")
            messagebox.showerror("Errore", error_msg)
            # Ripristina selezione precedente
            if self.navbar:
                self.navbar.portfolio_selector.set(self.current_portfolio_file)
    
    def _create_new_portfolio(self):
        """Crea un nuovo portfolio con validazione sicurezza"""
        try:
            dialog = ctk.CTkInputDialog(text="Nome del nuovo portfolio:", title="Nuovo Portfolio")
            portfolio_name = dialog.get_input()
            if portfolio_name:
                try:
                    # Crea path sicuro per il nuovo portfolio
                    safe_path = self.path_validator.create_safe_portfolio_path(portfolio_name)
                    if safe_path.exists():
                        messagebox.showerror("Errore", f"Il portfolio '{safe_path.name}' esiste già!")
                        return
                    self.logger.info(f"Creando nuovo portfolio sicuro: {safe_path}")
                    # Crea nuovo portfolio con path validato
                    new_portfolio_manager = PortfolioManager(str(safe_path))
                    self.current_portfolio_file = safe_path.name
                    self.portfolio_manager = new_portfolio_manager
                    if self.roadmap_dashboard:
                        self.roadmap_dashboard.set_portfolio_manager(self.portfolio_manager)
                except SecurityError as e:
                    self.logger.error(f"Nome portfolio non sicuro: {e}")
                    messagebox.showerror("Errore Sicurezza", f"Nome portfolio non sicuro: {e}")
                    return
                except Exception as e:
                    self.logger.error(f"Errore creazione portfolio sicuro: {e}")
                    messagebox.showerror("Errore", f"Errore creazione portfolio: {e}")
                    return
                
                # Aggiorna interfaccia
                self._refresh_portfolio_list()
                self.data_cache.clear()
                self._load_portfolio_data()
                
                messagebox.showinfo("Successo", f"Nuovo portfolio '{portfolio_name}' creato!")
                
        except Exception as e:
            error_msg = ErrorHandler.handle_file_error(e, "creazione nuovo portfolio")
            messagebox.showerror("Errore", error_msg)
    
    # Event handlers per callbacks componenti
    def _on_view_changed(self, view_type: str):
        """Gestisce il cambio di vista nella tabella portfolio"""
        # Allinea stato globale show_all_records e ricarica
        self.filter_state['show_all_records'] = (view_type == 'records')
        self._load_portfolio_data()
    
    def _on_asset_selected(self, asset_id: int):
        """Gestisce la selezione di un asset"""
        if self.asset_form and self.asset_form.edit_asset(asset_id):
            self.show_page("Asset")
        else:
            messagebox.showerror("Errore", f"Impossibile caricare asset ID {asset_id}")
    
    def _on_asset_saved(self, asset_data: Dict[str, Any]):
        """Gestisce il salvataggio di un asset"""
        self.data_cache.clear()  # Invalida cache
        self._load_portfolio_data()
        
        # Refresh degli altri componenti
        if self.charts_ui:
            self.charts_ui.refresh_charts()
        if self.export_ui:
            self.export_ui.refresh_stats()
    
    def _on_asset_deleted(self, asset_id: int):
        """Gestisce l'eliminazione di un asset"""
        self.data_cache.clear()  # Invalida cache
        self._load_portfolio_data()
        
        # Refresh degli altri componenti
        if self.charts_ui:
            self.charts_ui.refresh_charts()
        if self.export_ui:
            self.export_ui.refresh_stats()
    
    def _on_form_cleared(self):
        """Gestisce la pulizia del form"""
        # Nessuna azione particolare necessaria
        pass
    
    def _on_export_completed(self, export_type: str, filename: str):
        """Gestisce il completamento di un export"""
        print(f"Export {export_type} completato: {filename}")
        # Potrebbe aggiornare statistiche o log di export in futuro

    def _on_market_update_requested(self):
        """Avvia il flusso di aggiornamento prezzi dal pulsante Portfolio."""
        if not self.portfolio_manager:
            return

        if self.portfolio_table and hasattr(self.portfolio_table, 'set_market_update_state'):
            self.portfolio_table.set_market_update_state(True)
        self._perform_market_update()

    def _get_active_selection_ids(self) -> Optional[List[int]]:
        """Determina gli ID degli asset corrispondenti alla selezione corrente."""
        selection_ids: List[int] = []

        if self.portfolio_table:
            try:
                selection_ids.extend(self.portfolio_table.get_selected_asset_ids())
            except Exception as exc:
                self.logger.debug(f"Impossibile recuperare selezione tabella: {exc}")

        if selection_ids:
            return sorted(set(selection_ids))

        if isinstance(self._last_filtered_df, pd.DataFrame) and not self._last_filtered_df.empty:
            if 'id' in self._last_filtered_df.columns:
                collected: List[int] = []
                for raw_id in self._last_filtered_df['id'].dropna().unique():
                    try:
                        collected.append(int(float(raw_id)))
                    except (TypeError, ValueError):
                        continue
                if collected:
                    return sorted(set(collected))

        return None

    def _estimate_update_target_count(self, selection_ids: Optional[List[int]]) -> int:
        """Stima il numero di asset coinvolti nell'aggiornamento prezzi."""
        if selection_ids:
            return len(selection_ids)

        try:
            current_assets = self.portfolio_manager.get_current_assets_only()
            return len(current_assets)
        except Exception as exc:
            self.logger.debug(f"Impossibile stimare il numero di asset per l'aggiornamento: {exc}")
            return 0



    def _perform_market_update(self):
        """Esegue l'aggiornamento prezzi tramite Twelve Data in modo asincrono."""
        try:
            market_service = MarketDataService()
        except MarketDataError as exc:
            messagebox.showwarning("Aggiornamento prezzi", str(exc))
            if self.portfolio_table and hasattr(self.portfolio_table, 'set_market_update_state'):
                self.portfolio_table.set_market_update_state(False)
            return

        selection_ids = self._get_active_selection_ids()
        if selection_ids:
            self.logger.debug(
                "Aggiornamento prezzi limitato a %d asset dalla selezione corrente",
                len(selection_ids),
            )

        total_targets = self._estimate_update_target_count(selection_ids)
        if total_targets <= 0:
            if self.portfolio_table and hasattr(self.portfolio_table, 'set_market_update_state'):
                self.portfolio_table.set_market_update_state(False)
            messagebox.showinfo(
                "Aggiornamento prezzi",
                "Nessun asset disponibile per l'aggiornamento prezzi.",
            )
            return

        progress_dialog = MarketUpdateProgressDialog(self.root, total_targets)
        progress_dialog.handle_event({'stage': 'start', 'total': total_targets})

        def dispatch_progress(event: Dict[str, Any]) -> None:
            self.root.after(0, lambda e=event: progress_dialog.handle_event(e))



        def finalize(result: Optional[Dict[str, Any]], error: Optional[Exception]) -> None:
            progress_dialog.handle_event({'stage': 'done'})
            progress_dialog.close()
            if self.portfolio_table and hasattr(self.portfolio_table, 'set_market_update_state'):
                self.portfolio_table.set_market_update_state(False)
            if error:
                if isinstance(error, MarketDataError):
                    self.logger.error(f"Errore Twelve Data: {error}")
                    messagebox.showerror("Aggiornamento prezzi", str(error))
                else:
                    self.logger.exception("Aggiornamento prezzi: errore inatteso")
                    messagebox.showerror(
                        "Aggiornamento prezzi",
                        f"Errore inatteso: {error}",
                    )
                return
            safe_execute(self._load_portfolio_data)
            safe_execute(self._update_navbar_values)
            if self.charts_ui:
                safe_execute(self.charts_ui.refresh_charts)
            if self.export_ui:
                safe_execute(self.export_ui.refresh_stats)
            result = result or {}
            updated = result.get('updated', 0)
            errors = result.get('errors', [])
            skipped = result.get('skipped', [])
            alerts = result.get('alerts', [])
            nav_issue_ids: List[int] = []
            manual_issue_ids: List[int] = []
            price_alert_ids: List[int] = []
            if alerts:
                for alert in alerts:
                    alert_type = (alert.get('type') or '').lower()
                    raw_id = alert.get('id')
                    try:
                        parsed_id = int(raw_id)
                    except (TypeError, ValueError):
                        parsed_id = None
                        if raw_id is not None:
                            self.logger.debug("ID alert non numerico (%s): %s", alert_type, raw_id)
                    if parsed_id is None:
                        continue
                    if alert_type == 'issuer_nav_unavailable':
                        nav_issue_ids.append(parsed_id)
                    elif alert_type == 'manual_update_required':
                        manual_issue_ids.append(parsed_id)
                    elif alert_type == 'price_alert':
                        price_alert_ids.append(parsed_id)
            # Scrivi le evidenziazioni rosse nel file Excel
            if nav_issue_ids or manual_issue_ids or price_alert_ids:
                try:
                    self._write_alert_colors_to_excel(
                        nav_issue_ids=nav_issue_ids,
                        manual_issue_ids=manual_issue_ids,
                        price_alert_ids=price_alert_ids,
                    )
                    # Ricarica la tabella per mostrare i nuovi colori
                    self._load_portfolio_data()
                except Exception as exc:
                    self.logger.warning("Impossibile scrivere evidenziazioni alert in Excel: %s", exc)

            message_lines = [f"Asset aggiornati: {updated}"]
            if skipped:
                message_lines.append(f"Saltati: {len(skipped)}")
            if alerts:
                message_lines.append(f"Alert anomalie: {len(alerts)}")
            if errors:
                message_lines.append(f"Errori: {len(errors)}")
            summary_message = "\n".join(message_lines)
            if errors:
                messagebox.showerror("Aggiornamento prezzi", summary_message)
            elif alerts:
                messagebox.showwarning("Aggiornamento prezzi", summary_message)
            elif updated == 0:
                messagebox.showinfo("Aggiornamento prezzi", summary_message)
            else:
                messagebox.showinfo("Aggiornamento prezzi", summary_message)
            self.logger.info(
                "Aggiornamento prezzi completato: %s",
                summary_message.replace("\n", " | "),
            )
            self.logger.debug("Aggiornamento prezzi - dettagli: %s", result)
            self._show_market_update_report(result)

        def run_update() -> None:
                result: Optional[Dict[str, Any]] = None
                error: Optional[Exception] = None
                try:
                    result = self.portfolio_manager.update_market_prices(
                        market_service,
                        asset_ids=selection_ids,
                        progress_callback=dispatch_progress,
                    )
                except Exception as exc:
                    error = exc
                finally:
                    self.root.after(0, lambda: finalize(result, error))
    
        threading.Thread(target=run_update, daemon=True).start()

    def _write_alert_colors_to_excel(
        self,
        nav_issue_ids: List[int],
        manual_issue_ids: List[int],
        price_alert_ids: List[int]
    ) -> None:
        """Scrive i colori di evidenziazione rossi nel file Excel per gli alert"""
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        # Unisci tutti gli ID che devono avere sfondo rosso
        all_alert_ids = set(nav_issue_ids + manual_issue_ids + price_alert_ids)

        if not all_alert_ids:
            return

        try:
            wb = load_workbook(self.portfolio_manager.excel_file)
            ws = wb.active

            # Colore rosso chiaro per gli alert (stesso usato prima nei tag)
            red_fill = PatternFill(start_color='FFE5E5', end_color='FFE5E5', fill_type='solid')

            # Applica sfondo rosso a tutte le celle delle righe con alert
            for row_idx in range(2, ws.max_row + 1):
                id_cell = ws.cell(row=row_idx, column=1)
                try:
                    row_id = int(id_cell.value)
                except (TypeError, ValueError):
                    continue

                if row_id in all_alert_ids:
                    # Colora tutte le celle della riga
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = red_fill

            wb.save(self.portfolio_manager.excel_file)
            wb.close()
            self.logger.info(f"Scritti colori alert nel file Excel per {len(all_alert_ids)} righe")

        except Exception as e:
            self.logger.error(f"Errore scrittura colori alert in Excel: {e}")
            raise

    def _generate_update_recommendations(self, errors: list, alerts: list, skipped: list, details: list = None) -> List[str]:
        """Genera raccomandazioni per migliorare futuri aggiornamenti prezzi."""
        recommendations = []
        seen_recommendations = set()

        # Analizza conversioni valuta applicate
        if details:
            converted_assets = []
            for detail in details:
                if detail.get('conversion_applied'):
                    asset_id = detail.get('id')
                    original_price = detail.get('original_price')
                    original_currency = detail.get('original_currency')
                    conversion_rate = detail.get('conversion_rate')
                    converted_price = detail.get('price')
                    if asset_id and original_price and original_currency and conversion_rate:
                        converted_assets.append({
                            'id': asset_id,
                            'original_price': original_price,
                            'original_currency': original_currency,
                            'rate': conversion_rate,
                            'converted_price': converted_price,
                        })

            if converted_assets:
                rec = "CONVERSIONI VALUTA APPLICATE:"
                recommendations.append(rec)
                for conv in converted_assets:
                    rec = f"  ID {conv['id']}: {conv['original_price']:.2f} {conv['original_currency']} -> {conv['converted_price']:.2f} EUR (tasso: {conv['rate']:.4f})"
                    recommendations.append(rec)
                rec = "  Nota: Il ticker quota in valuta estera. Considera di cercare un ticker equivalente quotato in EUR per evitare conversioni."
                recommendations.append(rec)

        # Analizza errori
        for error in errors:
            error_msg = str(error.get('error', '')).lower()
            asset_id = error.get('id')

            if 'pro plan' in error_msg or 'grow plan' in error_msg:
                rec = f"ID {asset_id}: Aggiungere suffisso exchange al ticker (es. .MI, .L, .DE) per usare Yahoo Finance invece di TwelveData Pro"
                if rec not in seen_recommendations:
                    recommendations.append(rec)
                    seen_recommendations.add(rec)

            elif 'symbol' in error_msg and 'invalid' in error_msg:
                rec = f"ID {asset_id}: Correggere il ticker con simbolo valido includendo exchange (es. TICKER.MI)"
                if rec not in seen_recommendations:
                    recommendations.append(rec)
                    seen_recommendations.add(rec)

            elif 'currenttradingperiod' in error_msg or 'yfinance' in error_msg:
                rec = f"ID {asset_id}: Il ticker potrebbe essere delisted o non disponibile su Yahoo Finance - verificare e aggiornare"
                if rec not in seen_recommendations:
                    recommendations.append(rec)
                    seen_recommendations.add(rec)

        # Analizza alert
        for alert in alerts:
            alert_type = (alert.get('type') or '').lower()
            asset_id = alert.get('id')
            message = str(alert.get('message', '')).lower()

            if alert_type == 'issuer_nav_unavailable':
                rec = f"ID {asset_id}: Il NAV dell'emittente non e' disponibile - considerare di aggiungere un ticker Yahoo Finance alternativo"
                if rec not in seen_recommendations:
                    recommendations.append(rec)
                    seen_recommendations.add(rec)

            elif alert_type == 'manual_update_required':
                if 'pro plan' in message or 'grow plan' in message:
                    rec = f"ID {asset_id}: Aggiungere exchange suffix al ticker per evitare limitazioni TwelveData"
                    if rec not in seen_recommendations:
                        recommendations.append(rec)
                        seen_recommendations.add(rec)

            elif 'change_pct' in alert:
                change_pct = alert.get('change_pct')
                if change_pct and abs(change_pct) > 50:
                    rec = f"ID {asset_id}: Variazione {change_pct}% anomala - verificare che il ticker corrisponda all'exchange corretto (EUR vs USD/GBP)"
                    if rec not in seen_recommendations:
                        recommendations.append(rec)
                        seen_recommendations.add(rec)

        # Analizza skipped
        missing_ids_count = 0
        for skip in skipped:
            reason = skip.get('reason', '')
            if reason == 'missing_identifiers':
                missing_ids_count += 1

        if missing_ids_count > 0:
            rec = f"Ci sono {missing_ids_count} asset senza ticker/ISIN - aggiungerli per abilitare aggiornamenti automatici"
            if rec not in seen_recommendations:
                recommendations.append(rec)
                seen_recommendations.add(rec)

        return recommendations

    def _show_market_update_report(self, result: Dict[str, Any]):
        """Mostra un report dettagliato dell'aggiornamento prezzi."""
        if not result:
            return

        try:
            details = result.get('details', [])
            skipped = result.get('skipped', [])
            errors = result.get('errors', [])
            alerts = result.get('alerts', [])
        except AttributeError:
            self.logger.warning("Report aggiornamento prezzi non generato: risultato non valido")
            return

        window = ctk.CTkToplevel(self.root)
        window.title("Report aggiornamento prezzi")
        window.geometry("780x520")
        window.transient(self.root)
        window.grab_set()

        reason_map = {
            'missing_identifiers': 'Identificativi mancanti (ticker/ISIN)',
            'invalid_id': 'ID non valido',
            'missing_amount': 'Quantita assente',
            'price_not_positive': 'Prezzo non positivo',
            'base_record_missing': 'Record originale non trovato',
            'price_outlier': 'Variazione prezzo oltre soglia',
            'issuer_nav_unavailable': 'NAV emittente non disponibile (aggiornare manualmente)',
        }

        report_lines = []
        report_lines.append("Aggiornamento prezzi - report dettagliato")
        report_lines.append("")
        report_lines.append(f"Totale aggiornati: {result.get('updated', 0)}")
        report_lines.append(f"Totale saltati: {len(skipped)}")
        report_lines.append(f"Totale alert: {len(alerts)}")
        report_lines.append(f"Totale errori: {len(errors)}")

        report_lines.append("")
        report_lines.append("SUCCESSI")
        if details:
            for item in details:
                asset_id = item.get('id', 'N/D')
                symbol = item.get('symbol') or 'N/D'
                price = item.get('price')
                currency = item.get('currency') or ''
                new_record = item.get('new_record_id', 'N/D')
                if price is None:
                    price_str = 'n/d'
                elif currency:
                    price_str = f"{price} {currency}"
                else:
                    price_str = str(price)
                manual_suffix = " (aggiornamento manuale richiesto)" if item.get('manual') else ""
                report_lines.append("- ID {asset_id} ({symbol}): prezzo {price_str} -> nuovo record {new_record}{suffix}".format(
                    asset_id=asset_id, symbol=symbol, price_str=price_str, new_record=new_record, suffix=manual_suffix
                ))
        else:
            report_lines.append("- Nessun asset aggiornato")

        report_lines.append("")
        report_lines.append("SALTATI")
        if skipped:
            for item in skipped:
                asset_id = item.get('id', 'N/D')
                reason_code = item.get('reason', 'motivo_non_specificato')
                reason_desc = reason_map.get(reason_code, reason_code.replace('_', ' '))
                change_pct = item.get('change_pct')
                if change_pct is not None:
                    report_lines.append(f"- ID {asset_id}: {reason_desc} (Δ {change_pct}% )")
                else:
                    report_lines.append(f"- ID {asset_id}: {reason_desc}")
        else:
            report_lines.append("- Nessun asset saltato")

        report_lines.append("")
        report_lines.append("ALERT VARIAZIONI")
        if alerts:
            price_alerts = [a for a in alerts if a.get('change_pct') is not None or a.get('change_ratio') is not None]
            nav_alerts = [a for a in alerts if (a.get('type') or '').lower() == 'issuer_nav_unavailable']
            manual_alerts = [a for a in alerts if (a.get('type') or '').lower() == 'manual_update_required']
            for alert in price_alerts:
                asset_id = alert.get('id', 'N/D')
                symbol = alert.get('symbol') or 'N/D'
                prev_price = alert.get('previous_price')
                new_price = alert.get('new_price')
                change_ratio = alert.get('change_ratio')
                currency = alert.get('currency') or ''
                change_pct = alert.get('change_pct')
                if change_pct is None and change_ratio is not None:
                    change_pct = round(change_ratio * 100, 2)
                change_pct_str = f"{change_pct}%" if change_pct is not None else 'n/d'
                if currency:
                    report_lines.append(f"- ID {asset_id} ({symbol}): {prev_price} -> {new_price} {currency} (Δ {change_pct_str})")
                else:
                    report_lines.append(f"- ID {asset_id} ({symbol}): {prev_price} -> {new_price} (Δ {change_pct_str})")
            for alert in nav_alerts:
                asset_id = alert.get('id', 'N/D')
                message = alert.get('message') or 'NAV emittente non disponibile'
                report_lines.append(f"- ID {asset_id}: {message}")
            for alert in manual_alerts:
                asset_id = alert.get('id', 'N/D')
                message = alert.get('message') or 'Aggiornare manualmente il valore'
                report_lines.append(f"- ID {asset_id}: {message}")
            if not price_alerts and not nav_alerts and not manual_alerts:
                report_lines.append("- Nessuna variazione anomala")
        else:
            report_lines.append("- Nessuna variazione anomala")

        report_lines.append("")
        report_lines.append("ERRORI")
        if errors:
            for item in errors:
                asset_id = item.get('id', 'N/D')
                error_desc = item.get('error', 'Errore non specificato')
                report_lines.append(f"- ID {asset_id}: {error_desc}")
        else:
            report_lines.append("- Nessun errore")

        # Genera raccomandazioni per migliorare i futuri aggiornamenti
        recommendations = self._generate_update_recommendations(errors, alerts, skipped, details)
        if recommendations:
            report_lines.append("")
            report_lines.append("RACCOMANDAZIONI PER L'UTENTE")
            report_lines.append("Per migliorare i futuri aggiornamenti automatici, considera di:")
            for rec in recommendations:
                report_lines.append(f"- {rec}")

        report_lines.append("")
        report_lines.append("Nota: i nuovi record sono marcati con ""UPDATED BY AssetMind"" nel campo note.")

        content = "\n".join(report_lines)

        try:
            textbox = ctk.CTkTextbox(window, wrap='word')
        except AttributeError:
            textbox = tk.Text(window, wrap='word')
        textbox.pack(fill='both', expand=True, padx=20, pady=(20, 10))
        textbox.insert('end', content)
        try:
            textbox.configure(state='disabled')
        except (AttributeError, tk.TclError):
            textbox['state'] = 'disabled'

        close_button = ctk.CTkButton(window, text='Chiudi', command=window.destroy)
        close_button.pack(pady=(0, 20))

    def _show_column_filter(self, column: str):
        """Gestisce la richiesta di filtro per una colonna"""
        # Il filtro viene gestito direttamente dal componente PortfolioTable
        pass
    
    def _on_data_filtered(self, filtered_df: pd.DataFrame):
        """Gestisce i dati filtrati aggiornando i valori della navbar"""
        self._update_navbar_values()
    
    def _on_data_changed(self):
        """Gestisce la modifica dei dati (es. dopo riordino) ricaricando tutto"""
        try:
            # Pulisce la cache per forzare il ricaricamento
            self.data_cache.clear()
            # Ricarica i dati del portfolio
            self._load_portfolio_data()
            self.logger.debug("Dati ricaricati dopo modifica")
        except Exception as e:
            self.logger.error(f"Errore nel ricaricamento dati: {e}")
    
    def _center_window(self):
        """Centra la finestra al centro dello schermo"""
        self.root.update_idletasks()
        
        # Usa direttamente le dimensioni della configurazione
        window_width = 1200
        window_height = 800
        
        # Ottieni le dimensioni dello schermo
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Calcola la posizione per centrare la finestra
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        
        # Imposta la posizione della finestra
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
    
    def run(self):
        """Avvia l'applicazione"""
        try:
            self.logger.info("GAB AssetMind (Refactored) avviato con successo")
            self.logger.info(f"Portfolio attivo: {self.current_portfolio_file}")
            component_count = len([c for c in [self.navbar, self.portfolio_table, self.asset_form, self.charts_ui, self.export_ui] if c])
            self.logger.info(f"Componenti caricati: {component_count}/5")
            
            self.root.mainloop()
            
        except KeyboardInterrupt:
            self.logger.info("Applicazione interrotta dall'utente")
        except Exception as e:
            error_msg = ErrorHandler.handle_ui_error(e, "esecuzione principale")
            messagebox.showerror("Errore Critico", error_msg)
        finally:
            self._cleanup()
    
    def _cleanup(self):
        """Pulizia risorse alla chiusura"""
        try:
            if self.charts_ui:
                self.charts_ui.cleanup()
            if self.portfolio_table:
                self.portfolio_table.cleanup_performance_optimizers()
            if self.data_cache:
                self.data_cache.clear()
            self.logger.info("Cleanup completato")
            
        except Exception as e:
            self.logger.error(f"Errore durante cleanup: {e}")

class MarketUpdateProgressDialog:
    """Finestra modale che mostra l'avanzamento dell'aggiornamento prezzi."""

    def __init__(self, parent: ctk.CTk, total_assets: int) -> None:
        self.parent = parent
        self.total = max(int(total_assets), 0)
        self.processed = 0

        self.window = ctk.CTkToplevel(parent)
        self.window.title("Aggiornamento prezzi in corso")
        self.window.geometry("420x220")
        self.window.resizable(False, False)
        self.window.transient(parent)
        self.window.grab_set()
        self.window.protocol("WM_DELETE_WINDOW", lambda: None)

        padding = 20
        self.title_label = ctk.CTkLabel(
            self.window,
            text="Aggiornamento prezzi in corso",
            font=ctk.CTkFont(size=18, weight="bold"),
        )
        self.title_label.pack(pady=(padding, 10))

        self.progress_bar = ctk.CTkProgressBar(self.window, width=320)
        self.progress_bar.pack(pady=10, padx=padding, fill="x")

        if self.total > 0:
            self.progress_bar.set(0.0)
        else:
            self.progress_bar.configure(mode="indeterminate")
            self.progress_bar.start()

        self.status_label = ctk.CTkLabel(
            self.window,
            text="Preparazione...",
            anchor="w",
            justify="left",
        )
        self.status_label.pack(padx=padding, fill="x")

        self.detail_label = ctk.CTkLabel(
            self.window,
            text="",
            anchor="w",
            justify="left",
        )
        self.detail_label.pack(padx=padding, fill="x", pady=(2, 0))

        self.wait_label = ctk.CTkLabel(
            self.window,
            text="",
            anchor="w",
            justify="left",
        )
        self.wait_label.pack(padx=padding, fill="x", pady=(2, 0))

        self.elapsed_label = ctk.CTkLabel(
            self.window,
            text="Tempo trascorso: 0s",
            anchor="w",
            justify="left",
        )
        self.elapsed_label.pack(padx=padding, fill="x", pady=(10, padding))

        self._start_time = time.time()
        self._elapsed_job: Optional[str] = self.window.after(1000, self._update_elapsed)

    def handle_event(self, event: Dict[str, Any]) -> None:
        stage = event.get('stage')

        if stage == 'start':
            total = event.get('total')
            if total is not None:
                self._set_total(total)
            self.status_label.configure(text=f"Avvio aggiornamento ({self.total} asset)...")
            self.detail_label.configure(text="")
            self.wait_label.configure(text="")
            return





        if stage == 'item':
            total = event.get('total', self.total)
            if total is not None:
                self._set_total(total)
            self.processed = event.get('processed', self.processed)
            updated = event.get('updated', 0)
            skipped = event.get('skipped', 0)
            errors = event.get('errors', 0)
            self.status_label.configure(
                text=(
                    f"Elaborati {self.processed}/{self.total} - "
                    f"Aggiornati {updated}, Saltati {skipped}, Errori {errors}"
                )
            )
            asset_id = event.get('asset_id')
            status_raw = (event.get('status') or '').lower()
            status_map = {
                'updated': 'Aggiornato',
                'error': 'Errore',
                'skipped': 'Saltato',
                'invalid_id': 'ID non valido',
                'missing_identifiers': 'Identificativi mancanti',
                'missing_amount': 'Quantita mancante',
                'base_record_missing': 'Record originale non trovato',
                'price_not_positive': 'Prezzo non positivo',
                'manual_update': 'Aggiornamento manuale',
            }
            status = status_map.get(status_raw, status_raw.replace('_', ' ').strip().capitalize())
            message = event.get('message')
            message_map = {
                'missing_amount': 'Quantita mancante',
                'price_not_numeric': 'Prezzo non numerico',
                'price_not_positive': 'Prezzo non positivo',
                'missing_identifiers': 'Identificativi mancanti',
                'base_record_missing': 'Record originale non trovato',
                'price_alert': 'Variazione oltre soglia',
                'issuer_nav_unavailable': 'NAV emittente non disponibile',
                'manual_update_required': 'Aggiornamento manuale richiesto',
                'provider_error': 'Errore provider dati',
                'unexpected_error': 'Errore inatteso',
            }
            if isinstance(message, str):
                message = message_map.get(message.lower(), message)
            if asset_id is not None:
                detail = f"Ultimo asset ID {asset_id}"
                if status:
                    detail += f": {status}"
                if message:
                    detail += f" ({message}"
                    change_pct_event = event.get('change_pct')
                    if change_pct_event is not None:
                        detail += f", Δ {change_pct_event}%"
                    detail += ")"
                self.detail_label.configure(text=detail)
            else:
                if status or message:
                    detail = f"Stato: {status}" if status else "Stato"
                    if message:
                        detail += f" ({message})"
                    self.detail_label.configure(text=detail)
                else:
                    self.detail_label.configure(text="")
            self.wait_label.configure(text="")
            self._update_progress_bar()
            return


        if stage in {'wait_start', 'wait'}:
            remaining = event.get('remaining')
            total_seconds = event.get('total_seconds')
            if stage == 'wait_start' and total_seconds:
                self.wait_label.configure(text=f"Pausa per rate limit: {total_seconds}s")
                return
            if remaining is not None:
                self.wait_label.configure(text=f"Pausa per rate limit: {remaining}s rimanenti")
            else:
                self.wait_label.configure(text="Pausa per rate limit in corso...")
            return

        if stage == 'rate_limit_wait':
            remaining = event.get('remaining_seconds')
            total_seconds = event.get('total_seconds')
            if remaining is not None and total_seconds is not None:
                self.wait_label.configure(text=f"Attesa TwelveData: {remaining}s rimanenti (di {total_seconds}s)")
            elif total_seconds is not None:
                self.wait_label.configure(text=f"Attesa TwelveData: {total_seconds}s")
            else:
                self.wait_label.configure(text="Attesa TwelveData in corso...")
            return

        if stage == 'rate_limit_done':
            self.wait_label.configure(text="")
            return

        if stage == 'wait_end':
            self.wait_label.configure(text="")
            return

        if stage == 'done':
            self.wait_label.configure(text="")
            if self.total > 0 and self.processed >= self.total:
                self.progress_bar.set(1.0)
                self.status_label.configure(text="Aggiornamento completato")
            return

    def close(self) -> None:
        if self.total == 0:
            try:
                self.progress_bar.stop()
            except Exception:
                pass

        if self._elapsed_job is not None:
            try:
                self.window.after_cancel(self._elapsed_job)
            except Exception:
                pass
            self._elapsed_job = None

        try:
            self.window.grab_release()
        except Exception:
            pass

        try:
            self.window.destroy()
        except tk.TclError:
            pass

    def _set_total(self, total: int) -> None:
        if total is None:
            return
        if total <= 0:
            if self.total != 0:
                self.total = 0
                self.progress_bar.configure(mode="indeterminate")
                self.progress_bar.start()
            return

        if self.total == 0:
            try:
                self.progress_bar.stop()
            except Exception:
                pass
            self.progress_bar.configure(mode="determinate")

        self.total = max(total, 1)
        self._update_progress_bar()

    def _update_progress_bar(self) -> None:
        if self.total <= 0:
            return
        fraction = min(1.0, max(0.0, self.processed / self.total))
        self.progress_bar.set(fraction)

    def _update_elapsed(self) -> None:
        elapsed = int(time.time() - self._start_time)
        self.elapsed_label.configure(text=f"Tempo trascorso: {elapsed}s")
        self._elapsed_job = self.window.after(1000, self._update_elapsed)


def main():
    """Funzione principale di avvio"""
    try:
        # Inizializza logging prima di tutto
        logger = get_logger('startup')

        # Configura debug mode da variabile ambiente o argomenti
        debug_mode = (os.getenv('GAB_DEBUG', 'false').lower() == 'true' or
                     '--debug' in sys.argv or '-d' in sys.argv)
        if debug_mode:
            set_debug_mode(True)
            logger.info("Debug mode ATTIVATO")

        # Verifica requisiti di sistema
        logger.info("Avvio GAB AssetMind (Refactored Version)...")
        logger.info(f"Python: {sys.version}")
        logger.info(f"Working Directory: {os.getcwd()}")
        
        # Avvia applicazione
        app = GABAssetMind()
        app.run()
        
    except Exception as e:
        # Gestione errori critici
        import tkinter as tk
        from tkinter import messagebox
        
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(
            "Errore Critico GAB AssetMind", 
            f"Impossibile avviare l'applicazione:\n\n{e}\n\nContattare il supporto tecnico."
        )
        try:
            logger = get_logger('startup')
            logger.critical(f"Errore critico: {e}")
            import traceback
            logger.critical(f"Stack trace: {traceback.format_exc()}")
        except:
            # Fallback se anche il logging fallisce
            print(f"Errore critico: {e}")
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    main()


