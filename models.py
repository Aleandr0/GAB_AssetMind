#!/usr/bin/env python3
"""
Modelli dati per GAB AssetMind
Gestisce la persistenza dei dati portfolio in formato Excel

Classi principali:
- Asset: Rappresenta un singolo asset del portfolio
- PortfolioManager: Gestisce il database Excel e le operazioni CRUD
"""

import pandas as pd
import os
from datetime import datetime
from typing import Optional, List, Dict, Any, Callable, Set
from logging_config import get_logger
from security_validation import PathSecurityValidator, SecurityError
from date_utils import get_date_manager, format_for_storage, parse_date, get_today_formatted
from market_data import MarketDataError

PRICE_OUTLIER_THRESHOLD = 0.05
MANUAL_UPDATE_NOTE = "AssetMind - Da aggiornare manualmente"
MANUAL_UPDATE_MESSAGE = "Aggiornamento manuale richiesto"


def apply_global_filters(df: pd.DataFrame, column_filters: Optional[Dict[str, Set[str]]]) -> pd.DataFrame:
    """Applica filtri di colonna in modo coerente con la tabella portfolio."""
    if df is None or getattr(df, "empty", True) or not column_filters:
        return df if df is not None else pd.DataFrame()

    filtered = df.copy()
    try:
        for column, allowed_values in column_filters.items():
            if not allowed_values or column not in filtered.columns:
                continue
            col_values = filtered[column].fillna('N/A').astype(str)
            allowed = {str(value) for value in allowed_values}
            filtered = filtered[col_values.isin(allowed)]
        return filtered
    except Exception:
        return filtered


class Asset:
    """
    Rappresenta un singolo asset del portfolio
    
    Supporta tutti i tipi di asset: ETF, Azioni, Fondi di investimento, Buoni del Tesoro,
    PAC, Criptovalute, Liquidità, Immobiliare, Oggetti
    
    Attributi:
        asset_id: ID univoco dell'asset
        category: Categoria (ETF, Azioni, etc.)
        asset_name: Nome dell'asset
        position: Posizione/descrizione (testo libero)
        risk_level: Livello di rischio 1-5
        ticker: Codice ticker di borsa
        isin: Codice ISIN
        created_at/updated_at: Date di creazione/aggiornamento
        created_amount/updated_amount: Quantità iniziale/attuale
        created_unit_price/updated_unit_price: Prezzo unitario iniziale/attuale
        created_total_value/updated_total_value: Valore totale iniziale/attuale
        accumulation_plan: Descrizione piano di accumulo
        accumulation_amount: Importo accumulo mensile
        income_per_year: Reddito annuale da investimento
        rental_income: Reddito annuale da affitto
        note: Note libere
    """
    
    def __init__(self, asset_id: int = None, category: str = "", asset_name: str = "", 
                 position: str = "", risk_level: int = 1, ticker: str = "", 
                 isin: str = "", created_at: str = "", created_amount: float = 0.0, 
                 created_unit_price: float = 0.0, created_total_value: float = 0.0,
                 updated_at: str = "", updated_amount: float = 0.0, 
                 updated_unit_price: float = 0.0, updated_total_value: float = 0.0,
                 accumulation_plan: str = "", accumulation_amount: float = 0.0,
                 income_per_year: float = 0.0, rental_income: float = 0.0, note: str = ""):
        self.id = asset_id
        self.category = category
        self.asset_name = asset_name
        self.position = position
        self.risk_level = risk_level
        self.ticker = ticker
        self.isin = isin
        self.created_at = created_at
        self.created_amount = created_amount
        self.created_unit_price = created_unit_price
        self.created_total_value = created_total_value
        self.updated_at = updated_at
        self.updated_amount = updated_amount
        self.updated_unit_price = updated_unit_price
        self.updated_total_value = updated_total_value
        self.accumulation_plan = accumulation_plan
        self.accumulation_amount = accumulation_amount
        self.income_per_year = income_per_year
        self.rental_income = rental_income
        self.note = note

    def to_dict(self) -> Dict[str, Any]:
        return {
            'id': self.id,
            'category': self.category,
            'asset_name': self.asset_name,
            'position': self.position,
            'risk_level': self.risk_level,
            'ticker': self.ticker,
            'isin': self.isin,
            'created_at': self.created_at,
            'created_amount': self.created_amount,
            'created_unit_price': self.created_unit_price,
            'created_total_value': self.created_total_value,
            'updated_at': self.updated_at,
            'updated_amount': self.updated_amount,
            'updated_unit_price': self.updated_unit_price,
            'updated_total_value': self.updated_total_value,
            'accumulation_plan': self.accumulation_plan,
            'accumulation_amount': self.accumulation_amount,
            'income_per_year': self.income_per_year,
            'rental_income': self.rental_income,
            'note': self.note
        }

class PortfolioManager:
    """
    Gestisce il portfolio e la persistenza dei dati in Excel
    
    Fornisce operazioni CRUD (Create, Read, Update, Delete) per gli asset
    e calcoli di sommario per il portfolio.
    
    Attributes:
        excel_file: Nome del file Excel per la persistenza
        categories: Lista delle categorie di asset supportate
    """
    
    def __init__(self, excel_file: str = "portfolio_data.xlsx"):
        """
        Inizializza il gestore del portfolio con validazione sicurezza

        Args:
            excel_file: Nome del file Excel da usare (default: portfolio_data.xlsx)
        """
        self.path_validator = PathSecurityValidator()
        self.logger = get_logger('PortfolioManager')

        # Valida il path del file in modo sicuro
        try:
            self.excel_file = str(self.path_validator.validate_portfolio_path(excel_file))
            self.logger.info(f"Path portfolio validato: {self.excel_file}")
        except SecurityError as e:
            self.logger.error(f"Path portfolio non sicuro: {e}")
            # Fallback su path sicuro di default
            self.excel_file = str(self.path_validator.create_safe_portfolio_path("portfolio_data.xlsx"))
            self.logger.warning(f"Usando path sicuro di fallback: {self.excel_file}")
        except Exception as e:
            self.logger.error(f"Errore validazione path: {e}")
            self.excel_file = excel_file  # Fallback senza validazione

        self.categories = [
            "ETF", "Azioni", "Fondi di investimento", "Buoni del Tesoro",
            "PAC", "Criptovalute", "Liquidità", "Immobiliare", "Oggetti"
        ]
        self._initialize_excel()
    
    def _initialize_excel(self):
        if not os.path.exists(self.excel_file):
            # Usa schema centralizzato per evitare disallineamenti
            from config import DatabaseConfig
            columns = DatabaseConfig.DB_COLUMNS
            df = pd.DataFrame(columns=columns)
            df.to_excel(self.excel_file, index=False)
    
    def load_data(self) -> pd.DataFrame:
        """Loader semplice e collaudato: legge il foglio attivo/default,
        pulisce solo le date e calcola i totali mancanti, senza rinomine aggressive."""
        try:
            df = pd.read_excel(self.excel_file, keep_default_na=False, na_values=[''])
            self.logger.debug(f"load_data: columns={list(df.columns)} rows={len(df)}")

            # Pulisce le date (rimuove l'ora se presente)
            for col in ['created_at', 'updated_at']:
                if col in df.columns:
                    df[col] = df[col].apply(self._clean_date_from_excel)

            # Calcola i totali se mancanti
            if 'created_total_value' in df.columns:
                mask = pd.isna(df['created_total_value'])
                df.loc[mask, 'created_total_value'] = df.loc[mask, 'created_amount'].fillna(0) * df.loc[mask, 'created_unit_price'].fillna(0)

            if 'updated_total_value' in df.columns:
                mask = pd.isna(df['updated_total_value'])
                df.loc[mask, 'updated_total_value'] = df.loc[mask, 'updated_amount'].fillna(0) * df.loc[mask, 'updated_unit_price'].fillna(0)

            return df
        except Exception as e:
            self.logger.error(f"Errore nel caricamento dati: {e}")
            return pd.DataFrame()
    



    def update_market_prices(
        self,
        market_provider,
        asset_ids: Optional[List[int]] = None,
        progress_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
    ) -> Dict[str, Any]:
        """Aggiorna i prezzi duplicando i record e gestendo fallback multipli."""

        summary: Dict[str, Any] = {
            'updated': 0,
            'skipped': [],
            'errors': [],
            'alerts': [],
            'details': [],
            'manual_updates': [],
        }

        def _notify_progress(event: Dict[str, Any]) -> None:
            if not progress_callback:
                return
            try:
                progress_callback(event)
            except Exception:
                pass

        def _coerce_id(raw: Any) -> Optional[int]:
            try:
                return int(float(raw))
            except (TypeError, ValueError):
                return None

        processed_assets = 0

        def _notify_item(status: str, asset_id: Optional[int] = None, extra: Optional[Dict[str, Any]] = None) -> None:
            event: Dict[str, Any] = {
                'stage': 'item',
                'status': status,
                'processed': processed_assets,
                'updated': summary['updated'],
                'skipped': len(summary['skipped']),
                'errors': len(summary['errors']),
            }
            if asset_id is not None:
                event['asset_id'] = asset_id
            if extra:
                event.update(extra)
            _notify_progress(event)

        def _clean_str(value: Any) -> str:
            if pd.isna(value) or value is None:
                return ''
            return str(value).strip()

        def _to_positive_float(value: Any) -> float:
            try:
                if value in (None, '') or pd.isna(value):
                    return 0.0
                numeric = float(value)
                return numeric if numeric > 0 else 0.0
            except (TypeError, ValueError):
                return 0.0

        def _create_manual_placeholder(
            asset_id: Optional[int],
            raw_id: Any,
            row: pd.Series,
            reason_code: str,
            message: Optional[str] = None,
        ) -> bool:
            nonlocal next_id
            if asset_id is None:
                return False

            try:
                mask = df_all['id'].astype(str) == str(raw_id)
            except Exception:
                return False
            if not mask.any():
                return False

            placeholder = df_all[mask].iloc[-1].copy()
            placeholder['id'] = next_id
            placeholder['updated_at'] = today

            existing_note = _clean_str(placeholder.get('note'))
            note_parts = [part.strip() for part in existing_note.split('|') if part.strip()] if existing_note else []
            if MANUAL_UPDATE_NOTE not in note_parts:
                note_parts.append(MANUAL_UPDATE_NOTE)
            if MANUAL_UPDATE_MESSAGE not in note_parts:
                note_parts.append(MANUAL_UPDATE_MESSAGE)
            placeholder['note'] = ' | '.join(note_parts) if note_parts else f'{MANUAL_UPDATE_NOTE} | {MANUAL_UPDATE_MESSAGE}'

            new_rows.append(placeholder)
            summary['updated'] += 1
            summary['manual_updates'].append({
                'id': asset_id,
                'new_record_id': next_id,
                'reason': reason_code,
            })

            symbol = _clean_str(placeholder.get('ticker')) or _clean_str(placeholder.get('isin')) or _clean_str(row.get('asset_name'))
            price_value = _to_positive_float(placeholder.get('updated_unit_price')) or _to_positive_float(placeholder.get('created_unit_price')) or None

            detail_entry: Dict[str, Any] = {
                'id': asset_id,
                'symbol': symbol,
                'price': price_value if price_value is not None else None,
                'currency': placeholder.get('currency'),
                'new_record_id': next_id,
                'manual': True,
                'reason': reason_code,
            }
            summary['details'].append(detail_entry)

            alert_entry: Dict[str, Any] = {
                'id': asset_id,
                'type': 'manual_update_required',
                'reason': reason_code,
                'symbol': symbol,
            }
            if message:
                alert_entry['message'] = message
            summary['alerts'].append(alert_entry)

            next_id += 1
            return True

        def _maybe_pause(request_attempted: bool, row_position: int) -> None:
            """
            Implementa rate limiting per rispettare i limiti API TwelveData
            TwelveData Free Plan: 8 richieste/minuto (480/ora)

            Args:
                request_attempted: True se è stata effettuata una richiesta API
                row_position: Posizione corrente (1-indexed)
            """
            if not request_attempted:
                return  # Nessuna richiesta = nessuna pausa necessaria

            # TwelveData: 8 richieste/minuto
            # Strategia: pausa di 60+ secondi ogni 8 richieste per reset completo
            REQUESTS_PER_MINUTE = 8
            PAUSE_SECONDS = 65  # 60s reset + 5s safety margin

            if row_position % REQUESTS_PER_MINUTE == 0:
                import time
                self.logger.info(f"Rate limiting TwelveData: pausa di {PAUSE_SECONDS}s dopo {row_position} richieste per reset quota...")

                # Notifica l'inizio del countdown
                _notify_progress({
                    'stage': 'rate_limit_wait',
                    'total_seconds': PAUSE_SECONDS,
                    'remaining_seconds': PAUSE_SECONDS,
                })

                # Countdown con aggiornamenti ogni secondo
                for remaining in range(PAUSE_SECONDS, 0, -1):
                    time.sleep(1)
                    _notify_progress({
                        'stage': 'rate_limit_wait',
                        'total_seconds': PAUSE_SECONDS,
                        'remaining_seconds': remaining - 1,
                    })

                # Notifica fine attesa
                _notify_progress({'stage': 'rate_limit_done'})

        try:
            df_all = self.load_data()
        except Exception as exc:
            self.logger.error(f"update_market_prices: impossibile caricare dati - {exc}")
            summary['errors'].append({'id': None, 'error': str(exc)})
            _notify_progress({'stage': 'complete', 'summary': summary})
            return summary

        if df_all.empty:
            self.logger.info("update_market_prices: nessun dato disponibile")
            _notify_progress({'stage': 'start', 'total': 0})
            _notify_progress({'stage': 'complete', 'summary': summary})
            return summary

        try:
            current_df = self.get_current_assets_only()
        except Exception as exc:
            self.logger.error(f"update_market_prices: errore calcolo asset correnti - {exc}")
            summary['errors'].append({'id': None, 'error': str(exc)})
            _notify_progress({'stage': 'complete', 'summary': summary})
            return summary

        market_update_categories = {
            'Criptovalute',
            'ETF',
            'Fondi di investimento',
            'Azioni',
            'Titoli di stato',
        }
        if not current_df.empty and 'category' in current_df.columns:
            current_df = current_df[current_df['category'].isin(market_update_categories)]

        if current_df.empty:
            self.logger.info("update_market_prices: nessun asset corrente da aggiornare nelle categorie supportate")
            _notify_progress({'stage': 'start', 'total': 0})
            _notify_progress({'stage': 'complete', 'summary': summary})
            return summary

        id_filter: Optional[Set[int]] = None
        if asset_ids:
            cleaned_ids: Set[int] = set()
            for raw in asset_ids:
                coerced = _coerce_id(raw)
                if coerced is not None:
                    cleaned_ids.add(coerced)
            if cleaned_ids:
                id_filter = cleaned_ids
                current_df = current_df[current_df['id'].apply(lambda raw: _coerce_id(raw) in id_filter)]
                if current_df.empty:
                    self.logger.info("update_market_prices: nessun asset corrispondente alla selezione")
                    _notify_progress({'stage': 'start', 'total': 0})
                    _notify_progress({'stage': 'complete', 'summary': summary})
                    return summary

        total_rows = len(current_df)
        _notify_progress({'stage': 'start', 'total': total_rows})

        numeric_ids = pd.to_numeric(df_all['id'], errors='coerce').dropna()
        next_id = int(numeric_ids.max()) + 1 if not numeric_ids.empty else 1
        today = get_today_formatted('storage')

        new_rows: List[pd.Series] = []

        for row_index, (_, row) in enumerate(current_df.iterrows(), start=1):
            raw_id = row.get('id')
            asset_id = _coerce_id(raw_id)
            if asset_id is None:
                self.logger.warning(f"update_market_prices: ID non valido {raw_id}")
                summary['skipped'].append({'id': raw_id, 'reason': 'invalid_id'})
                processed_assets += 1
                _notify_item('invalid_id', asset_id=raw_id)
                continue

            if id_filter is not None and asset_id not in id_filter:
                continue

            ticker = _clean_str(row.get('ticker'))
            isin = _clean_str(row.get('isin'))

            if not ticker and not isin:
                manual_created = _create_manual_placeholder(
                    asset_id=asset_id,
                    raw_id=raw_id,
                    row=row,
                    reason_code='missing_identifiers',
                    message='Identificativi mancanti (ticker/ISIN)',
                )
                processed_assets += 1
                if manual_created:
                    _notify_item('manual_update', asset_id=asset_id, extra={'message': 'missing_identifiers'})
                else:
                    summary['errors'].append({'id': asset_id, 'error': 'manual_update_failed_missing_identifiers'})
                    _notify_item('error', asset_id=asset_id, extra={'message': 'missing_identifiers'})
                continue

            request_attempted = False
            try:
                request_attempted = True
                quote_data = market_provider.get_latest_price(
                    ticker=ticker or None,
                    isin=isin or None,
                    asset_name=_clean_str(row.get('asset_name')) or None,
                )
            except MarketDataError as exc:
                error_msg = str(exc)
                lowered_error = error_msg.lower()
                reason_code = 'issuer_nav_unavailable' if ('nav blackrock' in lowered_error or 'nav emittente' in lowered_error) else 'provider_error'
                manual_created = _create_manual_placeholder(
                    asset_id=asset_id,
                    raw_id=raw_id,
                    row=row,
                    reason_code=reason_code,
                    message=error_msg,
                )
                processed_assets += 1
                if manual_created:
                    _notify_item('manual_update', asset_id=asset_id, extra={'message': reason_code})
                else:
                    if reason_code == 'issuer_nav_unavailable':
                        summary['alerts'].append({
                            'id': asset_id,
                            'type': 'issuer_nav_unavailable',
                            'symbol': ticker or isin or row.get('asset_name'),
                            'message': error_msg,
                        })
                        summary['skipped'].append({
                            'id': asset_id,
                            'reason': 'issuer_nav_unavailable',
                        })
                        _notify_item('skipped', asset_id=asset_id, extra={'message': 'issuer_nav_unavailable'})
                    else:
                        summary['errors'].append({'id': asset_id, 'error': error_msg})
                        _notify_item('error', asset_id=asset_id, extra={'message': error_msg})
                _maybe_pause(request_attempted, row_index)
                continue
            except Exception as exc:
                error_msg = str(exc)
                self.logger.error(
                    f"update_market_prices: errore inatteso per ID {asset_id}: {exc}",
                    exc_info=True,
                )
                manual_created = _create_manual_placeholder(
                    asset_id=asset_id,
                    raw_id=raw_id,
                    row=row,
                    reason_code='unexpected_error',
                    message=error_msg,
                )
                processed_assets += 1
                if manual_created:
                    _notify_item('manual_update', asset_id=asset_id, extra={'message': 'unexpected_error'})
                else:
                    summary['errors'].append({'id': asset_id, 'error': error_msg})
                    _notify_item('error', asset_id=asset_id, extra={'message': error_msg})
                _maybe_pause(request_attempted, row_index)
                continue

            price = quote_data.get('price')
            try:
                price_value = float(price)
            except (TypeError, ValueError):
                manual_created = _create_manual_placeholder(
                    asset_id=asset_id,
                    raw_id=raw_id,
                    row=row,
                    reason_code='price_not_numeric',
                    message='Prezzo non numerico',
                )
                processed_assets += 1
                if manual_created:
                    _notify_item('manual_update', asset_id=asset_id, extra={'message': 'price_not_numeric'})
                else:
                    summary['errors'].append({'id': asset_id, 'error': 'price_not_numeric'})
                    _notify_item('error', asset_id=asset_id, extra={'message': 'price_not_numeric'})
                _maybe_pause(request_attempted, row_index)
                continue

            # Conversione automatica valuta se necessario
            quote_currency = (quote_data.get('currency') or '').upper()
            expected_currency = 'EUR'  # Valuta di default dell'applicazione
            conversion_rate = None
            original_price = None
            original_currency = None

            if quote_currency and quote_currency != expected_currency:
                self.logger.info(
                    f"ID {asset_id}: prezzo in {quote_currency}, conversione automatica in {expected_currency}"
                )
                try:
                    # Usa yfinance per ottenere il tasso di cambio
                    import yfinance as yf
                    currency_pair = f"{quote_currency}{expected_currency}=X"
                    ticker_fx = yf.Ticker(currency_pair)
                    fx_rate = ticker_fx.fast_info.last_price

                    if fx_rate and fx_rate > 0:
                        original_price = price_value
                        original_currency = quote_currency
                        price_value = price_value * fx_rate
                        conversion_rate = fx_rate
                        self.logger.info(
                            f"ID {asset_id}: conversione {original_price} {quote_currency} -> {price_value:.4f} {expected_currency} (tasso: {fx_rate:.4f})"
                        )
                        quote_data['currency'] = expected_currency
                        quote_data['conversion_applied'] = True
                        quote_data['original_price'] = original_price
                        quote_data['original_currency'] = original_currency
                        quote_data['conversion_rate'] = conversion_rate
                    else:
                        self.logger.warning(
                            f"ID {asset_id}: tasso di cambio {currency_pair} non disponibile, uso prezzo originale"
                        )
                except Exception as fx_exc:
                    self.logger.warning(
                        f"ID {asset_id}: impossibile convertire {quote_currency} -> {expected_currency}: {fx_exc}"
                    )

            if price_value <= 0:
                manual_created = _create_manual_placeholder(
                    asset_id=asset_id,
                    raw_id=raw_id,
                    row=row,
                    reason_code='price_not_positive',
                    message='Prezzo non positivo',
                )
                processed_assets += 1
                if manual_created:
                    _notify_item('manual_update', asset_id=asset_id, extra={'message': 'price_not_positive'})
                else:
                    summary['skipped'].append({'id': asset_id, 'reason': 'price_not_positive'})
                    _notify_item('skipped', asset_id=asset_id, extra={'message': 'price_not_positive'})
                _maybe_pause(request_attempted, row_index)
                continue

            amount = _to_positive_float(row.get('updated_amount'))
            amount_override = None
            if amount == 0:
                fallback_amount = _to_positive_float(row.get('created_amount'))
                if fallback_amount == 0:
                    manual_created = _create_manual_placeholder(
                        asset_id=asset_id,
                        raw_id=raw_id,
                        row=row,
                        reason_code='missing_amount',
                        message='Quantita mancante',
                    )
                    processed_assets += 1
                    if manual_created:
                        _notify_item('manual_update', asset_id=asset_id, extra={'message': 'missing_amount'})
                    else:
                        summary['skipped'].append({'id': asset_id, 'reason': 'missing_amount'})
                        _notify_item('skipped', asset_id=asset_id, extra={'message': 'missing_amount'})
                    _maybe_pause(request_attempted, row_index)
                    continue
                amount = fallback_amount
                amount_override = fallback_amount

            mask = df_all['id'].astype(str) == str(raw_id)
            if not mask.any():
                self.logger.warning(
                    f"update_market_prices: record originale non trovato per ID {asset_id}"
                )
                summary['skipped'].append({'id': asset_id, 'reason': 'base_record_missing'})
                processed_assets += 1
                _notify_item('skipped', asset_id=asset_id, extra={'message': 'base_record_missing'})
                _maybe_pause(request_attempted, row_index)
                continue

            # IMPORTANTE: Usa l'ULTIMO record (iloc[-1]) per avere i dati più recenti
            # Questo è fondamentale per il confronto del prezzo precedente
            matching_records = df_all[mask]
            last_record = matching_records.iloc[-1].copy()

            # Log per debug: mostra quanti record esistono per questo asset
            num_records = len(matching_records)
            if num_records > 1:
                self.logger.debug(f"ID {asset_id}: trovati {num_records} record, usando l'ultimo (ID record: {last_record.get('id')})")

            base_row = last_record.copy()
            base_row['id'] = next_id
            base_row['updated_at'] = today
            base_row['updated_unit_price'] = round(price_value, 6)
            if amount_override is not None:
                base_row['updated_amount'] = amount_override
            total_value = round(amount * price_value, 2)
            base_row['updated_total_value'] = total_value

            note_tag = 'UPDATED BY AssetMind'
            alert_tag = 'PRICE ALERT'
            existing_note = _clean_str(base_row.get('note'))
            note_parts = [part.strip() for part in existing_note.split('|') if part.strip()] if existing_note else []
            if note_tag not in note_parts:
                note_parts.append(note_tag)

            # Confronta con il prezzo dell'ULTIMO record, non del primo
            price_alert_info: Optional[Dict[str, Any]] = None
            previous_price = _to_positive_float(last_record.get('updated_unit_price')) or _to_positive_float(last_record.get('created_unit_price'))
            if previous_price > 0:
                change_ratio = abs(price_value - previous_price) / previous_price
                if change_ratio > PRICE_OUTLIER_THRESHOLD:
                    change_pct = round(change_ratio * 100, 2)
                    price_alert_info = {
                        'id': asset_id,
                        'type': 'price_alert',
                        'symbol': quote_data.get('symbol'),
                        'currency': quote_data.get('currency'),
                        'previous_price': round(previous_price, 6),
                        'new_price': round(price_value, 6),
                        'change_ratio': round(change_ratio, 6),
                        'change_pct': change_pct,
                    }
                    summary['alerts'].append(dict(price_alert_info))
                    if alert_tag not in note_parts:
                        note_parts.append(alert_tag)
                    self.logger.warning(
                        "update_market_prices: variazione prezzo anomala per ID %s (da %s a %s, %s%%)",
                        asset_id,
                        previous_price,
                        price_value,
                        change_pct,
                    )

            if note_parts:
                base_row['note'] = ' | '.join(note_parts)

            if not ticker and quote_data.get('symbol'):
                base_row['ticker'] = quote_data['symbol']

            new_rows.append(base_row)
            summary['updated'] += 1
            detail_entry = {
                'id': asset_id,
                'symbol': quote_data.get('symbol'),
                'price': round(price_value, 6),
                'currency': quote_data.get('currency'),
                'new_record_id': next_id,
            }
            if price_alert_info:
                detail_entry['alert'] = True
                detail_entry['change_pct'] = price_alert_info.get('change_pct')
            if quote_data.get('conversion_applied'):
                detail_entry['conversion_applied'] = True
                detail_entry['original_price'] = quote_data.get('original_price')
                detail_entry['original_currency'] = quote_data.get('original_currency')
                detail_entry['conversion_rate'] = quote_data.get('conversion_rate')
            summary['details'].append(detail_entry)

            next_id += 1
            processed_assets += 1
            notify_extra = None
            if price_alert_info:
                notify_extra = {
                    'message': 'price_alert',
                    'change_pct': price_alert_info.get('change_pct'),
                }
            _notify_item('updated', asset_id=asset_id, extra=notify_extra)
            _maybe_pause(request_attempted, row_index)

        if not new_rows:
            _notify_progress({'stage': 'complete', 'summary': summary})
            return summary

        try:
            df_updated = pd.concat([df_all, pd.DataFrame(new_rows)], ignore_index=True)
            self.save_data(df_updated)
        except Exception as exc:
            self.logger.error(f"update_market_prices: errore salvataggio dati - {exc}")
            summary['errors'].append({'id': None, 'error': str(exc)})
            _notify_progress({'stage': 'complete', 'summary': summary})
            return summary

        _notify_progress({'stage': 'complete', 'summary': summary})
        return summary

    def _clean_date_from_excel(self, date_value):
        """Pulisce le date caricate da Excel usando il sistema centralizzato"""
        try:
            date_manager = get_date_manager()
            result = date_manager.format_for_storage(date_value)
            return result if result else ""
        except Exception as e:
            self.logger.error(f"Errore pulizia data da Excel: {e}")
            return ""
    
    def save_data(self, df: pd.DataFrame):
        try:
            # Salva usando openpyxl per supportare le formule Excel
            self.save_data_with_formulas(df)
            return True
        except Exception as e:
            self.logger.error(f"Errore nel salvataggio dati: {e}")
            return False
    
    def save_data_with_formulas(self, df: pd.DataFrame):
        """Salva i dati con formule Excel per i calcoli automatici"""
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from datetime import datetime
        
        # Crea copia del DataFrame e formatta le date come solo giorno
        df_clean = df.copy()
        
        # Converte le colonne date per rimuovere l'ora
        date_columns = ['created_at', 'updated_at']
        for col in date_columns:
            if col in df_clean.columns:
                df_clean[col] = df_clean[col].apply(self._format_date_for_excel)
        
        # Crea workbook
        wb = Workbook()
        ws = wb.active
        
        # Inserisce i dati dal DataFrame pulito
        for r in dataframe_to_rows(df_clean, index=False, header=True):
            ws.append(r)
        
        # Trova le colonne dei valori totali
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header_list = list(header_row)
        
        try:
            created_amount_idx = header_list.index('created_amount')
            created_price_idx = header_list.index('created_unit_price')
            created_total_idx = header_list.index('created_total_value')
            
            updated_amount_idx = header_list.index('updated_amount')
            updated_price_idx = header_list.index('updated_unit_price')
            updated_total_idx = header_list.index('updated_total_value')
            
            # Converti gli indici in lettere di colonna
            from openpyxl.utils import get_column_letter
            
            created_amount_col = get_column_letter(created_amount_idx + 1)
            created_price_col = get_column_letter(created_price_idx + 1)
            created_total_col = get_column_letter(created_total_idx + 1)
            
            updated_amount_col = get_column_letter(updated_amount_idx + 1)
            updated_price_col = get_column_letter(updated_price_idx + 1)
            updated_total_col = get_column_letter(updated_total_idx + 1)
            
            # Applica le formule a tutte le righe (dalla riga 2 in poi)
            for row in range(2, ws.max_row + 1):
                # Formula per Valore Totale Iniziale = Quantità * Prezzo Unitario
                formula1 = f'={created_amount_col}{row}*{created_price_col}{row}'
                ws[f'{created_total_col}{row}'] = formula1
                
                # Formula per Valore Totale Attuale = Quantità * Prezzo Unitario
                formula2 = f'={updated_amount_col}{row}*{updated_price_col}{row}'
                ws[f'{updated_total_col}{row}'] = formula2
                
        except (ValueError, IndexError) as e:
            self.logger.error(f"Errore nell'applicazione delle formule: {e}")
        
        # Salva il file
        wb.save(self.excel_file)
    
    def _format_date_for_excel(self, date_value):
        """Formatta le date per Excel usando il sistema centralizzato"""
        try:
            date_manager = get_date_manager()
            return date_manager.format_for_excel(date_value)
        except Exception as e:
            self.logger.error(f"Errore formattazione data per Excel: {e}")
            # Fallback per compatibilità
            if pd.isna(date_value) or date_value == "" or date_value is None:
                return None
            date_str = str(date_value)
            return date_str.split()[0] if " " in date_str else date_str
    
    def add_asset(self, asset: Asset) -> bool:
        """
        Aggiunge un nuovo asset al portfolio
        
        Args:
            asset: Oggetto Asset da aggiungere
            
        Returns:
            True se il salvataggio è riuscito, False altrimenti
        """
        df = self.load_data()
        
        # Assegna ID automatico se non specificato
        if asset.id is None:
            asset.id = len(df) + 1 if not df.empty else 1
        
        # Assegna data corrente se non specificata usando sistema centralizzato
        if asset.created_at == "":
            from date_utils import get_today_formatted
            asset.created_at = get_today_formatted('storage')
        
        # Aggiunge l'asset al DataFrame
        new_row = pd.DataFrame([asset.to_dict()])
        df = pd.concat([df, new_row], ignore_index=True)
        
        return self.save_data(df)
    
    def update_asset(self, asset_id: int, updated_data: Dict[str, Any]) -> bool:
        df = self.load_data()
        
        if asset_id not in df['id'].values:
            return False
        
        # Non forzare la data di aggiornamento - viene gestita dal chiamante
        
        for key, value in updated_data.items():
            df.loc[df['id'] == asset_id, key] = value
        
        return self.save_data(df)
    
    def delete_asset(self, asset_id: int) -> bool:
        df = self.load_data()
        
        if asset_id not in df['id'].values:
            return False
        
        df = df[df['id'] != asset_id]
        return self.save_data(df)
    
    def get_asset(self, asset_id: int) -> Optional[Asset]:
        df = self.load_data()
        
        if asset_id not in df['id'].values:
            return None
        
        row = df[df['id'] == asset_id].iloc[0]
        return Asset(
            asset_id=row['id'],
            category=row['category'],
            asset_name=row['asset_name'],
            position=row['position'],
            risk_level=row['risk_level'],
            ticker=row['ticker'],
            isin=row['isin'],
            created_at=row['created_at'],
            created_amount=row['created_amount'],
            created_unit_price=row['created_unit_price'],
            created_total_value=row['created_total_value'],
            updated_at=row['updated_at'],
            updated_amount=row['updated_amount'],
            updated_unit_price=row['updated_unit_price'],
            updated_total_value=row['updated_total_value'],
            accumulation_plan=row['accumulation_plan'],
            accumulation_amount=row['accumulation_amount'],
            income_per_year=row['income_per_year'],
            rental_income=row['rental_income'],
            note=row['note']
        )
    
    def get_assets_by_category(self, category: str) -> List[Asset]:
        df = self.load_data()
        filtered_df = df[df['category'] == category]
        
        assets = []
        for _, row in filtered_df.iterrows():
            assets.append(Asset(
                asset_id=row['id'],
                category=row['category'],
                asset_name=row['asset_name'],
                position=row['position'],
                risk_level=row['risk_level'],
                ticker=row['ticker'],
                isin=row['isin'],
                created_at=row['created_at'],
                created_amount=row['created_amount'],
                created_unit_price=row['created_unit_price'],
                created_total_value=row['created_total_value'],
                updated_at=row['updated_at'],
                updated_amount=row['updated_amount'],
                updated_unit_price=row['updated_unit_price'],
                updated_total_value=row['updated_total_value'],
                accumulation_plan=row['accumulation_plan'],
                accumulation_amount=row['accumulation_amount'],
                income_per_year=row['income_per_year'],
                rental_income=row['rental_income'],
                note=row['note']
            ))
        
        return assets
    
    def get_portfolio_summary(self) -> Dict[str, Any]:
        df = self.load_data()

        if df.empty:
            return {
                'total_value': 0,
                'total_income': 0,
                'categories_count': {},
                'risk_distribution': {},
                'monthly_accumulation': 0
            }

        # NUOVA LOGICA: Considera solo record più recenti per ogni asset unico
        # Asset identificato da: category + asset_name + position + isin

        # IMPORTANTE: Usa stessa normalizzazione di get_current_assets_only() per coerenza
        def _norm(s):
            if pd.isna(s):
                return ''
            val = str(s).strip()
            if val.lower() in {'na','n/a','none','null','nan',''}:
                return ''
            return val

        # Normalizza campi chiave per deduplica (stessa logica di get_current_assets_only)
        for key_col in ['category','asset_name','position','isin']:
            if key_col in df.columns:
                df[key_col] = df[key_col].apply(_norm)
            else:
                df[key_col] = ''

        # Crea colonna per identificare univocamente ogni asset
        df['asset_key'] = (df['category'] + '|' +
                          df['asset_name'] + '|' +
                          df['position'] + '|' +
                          df['isin'])

        # Converte date per ordinamento (usa updated_at se disponibile, altrimenti created_at)
        df['effective_date'] = pd.to_datetime(
            df['updated_at'].replace(['', 'NA', 'N/A', 'na'], pd.NA).fillna(df['created_at']),
            format='%Y-%m-%d', errors='coerce'
        )

        # Per ogni asset unico, prende solo il record con data più recente
        latest_records = df.sort_values('effective_date', ascending=False).groupby('asset_key').first().reset_index()

        # Calcola i totali sui record più recenti
        total_value = latest_records['updated_total_value'].fillna(latest_records['created_total_value']).sum()
        total_income = (latest_records['income_per_year'].fillna(0).sum() +
                       latest_records['rental_income'].fillna(0).sum())
        categories_count = latest_records['category'].value_counts().to_dict()
        risk_distribution = latest_records['risk_level'].value_counts().to_dict()
        monthly_accumulation = latest_records['accumulation_amount'].fillna(0).sum()

        return {
            'total_value': total_value,
            'total_income': total_income,
            'categories_count': categories_count,
            'risk_distribution': risk_distribution,
            'monthly_accumulation': monthly_accumulation
        }
    
    def get_current_assets_only(self):
        """
        Ritorna solo gli asset più recenti (un record per asset unico)
        """
        df = self.load_data()
        if df.empty:
            return df

        # Normalizza campi chiave per deduplica
        def _norm(s):
            if pd.isna(s):
                return ''
            val = str(s).strip()
            if val.lower() in {'na','n/a','none','null','nan',''}:
                return ''
            return val
        for key_col in ['category','asset_name','position','isin']:
            if key_col in df.columns:
                df[key_col] = df[key_col].apply(_norm)
            else:
                df[key_col] = ''

        df['asset_key'] = (df['category'] + '|' + df['asset_name'] + '|' + df['position'] + '|' + df['isin'])

        # Ordina per data effettiva (updated_at preferito), poi prendi il primo per asset_key
        df['effective_date'] = pd.to_datetime(
            df['updated_at'].replace(['', 'NA', 'N/A', 'na'], pd.NA).fillna(df['created_at']),
            format='%Y-%m-%d', errors='coerce'
        )

        df['original_order'] = df.index
        latest_records = df.sort_values(['effective_date','original_order'], ascending=[False, True])\
                           .groupby('asset_key').first().reset_index()
        latest_records = latest_records.sort_values('original_order')
        latest_records = latest_records.drop(columns=[c for c in ['asset_key','effective_date','original_order'] if c in latest_records.columns])

        self.logger.debug(f"current_assets_only: input_rows={len(df)} output_rows={len(latest_records)}")
        return latest_records
    
    def get_filtered_assets(self, filters: Dict[str, Any] = None) -> pd.DataFrame:
        """
        Ritorna gli asset filtrati secondo i criteri specificati
        
        Args:
            filters: Dizionario con i filtri da applicare
                    {
                        'category': str | None,
                        'position': str | None, 
                        'risk_level': List[int] | None,
                        'value_range': (min, max) | None,
                        'name_search': str | None
                    }
        """
        df = self.get_current_assets_only()
        
        if df.empty or not filters:
            return df
        
        # Filtro per categoria
        if filters.get('category') and filters['category'] != 'All':
            df = df[df['category'] == filters['category']]
        
        # Filtro per posizione (contiene testo)
        if filters.get('position'):
            position_text = filters['position'].lower()
            df = df[df['position'].fillna('').str.lower().str.contains(position_text, na=False)]
        
        # Filtro per livello di rischio
        if filters.get('risk_level'):
            df = df[df['risk_level'].isin(filters['risk_level'])]
        
        # Filtro per range di valore
        if filters.get('value_range'):
            min_val, max_val = filters['value_range']
            # Usa updated_total_value se disponibile, altrimenti created_total_value
            df['current_value'] = df['updated_total_value'].fillna(df['created_total_value'])
            if min_val is not None:
                df = df[df['current_value'] >= min_val]
            if max_val is not None:
                df = df[df['current_value'] <= max_val]
            df = df.drop('current_value', axis=1)
        
        # Filtro per nome asset (contiene testo)
        if filters.get('name_search'):
            name_text = filters['name_search'].lower()
            df = df[df['asset_name'].fillna('').str.lower().str.contains(name_text, na=False)]
        
        return df
    
    def get_unique_positions(self) -> List[str]:
        """Ritorna tutte le posizioni uniche nel portfolio"""
        df = self.get_current_assets_only()
        positions = df['position'].fillna('').unique()
        return [pos for pos in positions if pos != '']
    
    def color_historical_records(self):
        """
        Colora i record storici di azzurro direttamente nel file Excel.
        I record storici sono quelli non più attuali per un dato asset.
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font
        except ImportError:
            self.logger.error("openpyxl non disponibile. Installa con: pip install openpyxl")
            return
        
        try:
            # Carica tutti i dati per identificare i record storici
            df = self.load_data()
            if df.empty:
                return
            
            # Identifica gli asset correnti
            current_df = self.get_current_assets_only()
            current_ids = set(current_df['id'].astype(int))
            
            # Record storici sono quelli non nell'insieme degli attuali
            historical_ids = set(df['id'].astype(int)) - current_ids
            
            self.logger.info(f"Record storici da colorare: {len(historical_ids)}")
            self.logger.info(f"Record attuali: {len(current_ids)}")
            
            # Apri il file Excel con openpyxl
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Definisci il colore azzurro per il testo
            blue_font = Font(color="0066CC")  # Azzurro
            
            # Itera sulle righe (skip header row 1)
            for row_idx in range(2, ws.max_row + 1):
                # La colonna A contiene l'ID (primo valore)
                id_cell = ws.cell(row=row_idx, column=1)
                
                try:
                    record_id = int(id_cell.value)
                    
                    # Se è un record storico, colora tutta la riga di azzurro
                    if record_id in historical_ids:
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.font = blue_font
                except (ValueError, TypeError):
                    # Se l'ID non è convertibile, skip
                    continue
            
            # Salva il file
            wb.save(self.excel_file)
            self.logger.info(f"File Excel aggiornato con {len(historical_ids)} record storici colorati di azzurro")
            
        except Exception as e:
            self.logger.error(f"Errore nella colorazione dei record storici: {e}")
            import traceback
            self.logger.debug(f"Stack trace: {traceback.format_exc()}")


