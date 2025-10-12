#!/usr/bin/env python3
"""
Componente interfaccia export per GAB AssetMind
Gestisce l'esportazione di dati in vari formati (CSV, PDF, Excel backup)
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.patches import FancyBboxPatch
import numpy as np
from datetime import datetime
import os
from typing import Optional

from config import UIConfig, Messages
from utils import ErrorHandler, safe_execute
from models import PortfolioManager, apply_global_filters
from ui_components import BaseUIComponent
from security_validation import PathSecurityValidator, SecurityError
from logging_config import get_logger

class ExportUI(BaseUIComponent):
    """Componente per l'esportazione di dati"""
    
    def __init__(self, parent, portfolio_manager: PortfolioManager):
        super().__init__(parent, portfolio_manager)
        self.export_frame = None
        self.path_validator = PathSecurityValidator()
        self.logger = get_logger('ExportUI')
        self._external_filtered_df = None
        self._filter_info = None
    
    def create_export_interface(self) -> ctk.CTkFrame:
        """Crea l'interfaccia completa per l'export"""
        self.export_frame = ctk.CTkFrame(self.parent)
        self.export_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        self._create_header()
        self._create_export_options()
        self._create_info_section()
        
        return self.export_frame
    
    def _create_header(self):
        """Crea l'header della pagina export"""
        self.header_frame = ctk.CTkFrame(self.export_frame, height=80)
        self.header_frame.pack(fill="x", padx=10, pady=10)
        self.header_frame.pack_propagate(False)

        # Container per titolo e filtri (layout orizzontale)
        title_container = ctk.CTkFrame(self.header_frame)
        title_container.pack(fill="both", expand=True)

        # Titolo a sinistra
        title_label = ctk.CTkLabel(
            title_container,
            text="📄 Esportazione Dati",
            font=ctk.CTkFont(**UIConfig.FONTS['title'])
        )
        title_label.pack(side="left", padx=20, pady=10)

        # Label filtri attivi a destra (multilinea con wrapping)
        base_size = UIConfig.FONTS['text'].get('size', 12)
        reduced_size = int(base_size * 0.7)

        self.filter_label = ctk.CTkLabel(
            title_container,
            text="",
            font=ctk.CTkFont(family=UIConfig.FONTS['text'].get('family', 'Arial'), size=reduced_size),
            text_color=UIConfig.COLORS['secondary'],
            justify="left",
            wraplength=600
        )
        self.filter_label.pack(side="left", padx=20, pady=10, fill="x", expand=True)

        # Sottotitolo sotto
        subtitle_label = ctk.CTkLabel(
            self.header_frame,
            text="Esporta i dati del portfolio in diversi formati",
            font=ctk.CTkFont(**UIConfig.FONTS['text']),
            text_color=UIConfig.COLORS['secondary']
        )
        subtitle_label.pack()
    
    def _create_export_options(self):
        """Crea le opzioni di esportazione"""
        options_frame = ctk.CTkFrame(self.export_frame)
        options_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Grid layout per le opzioni
        options_frame.grid_columnconfigure(0, weight=1)
        options_frame.grid_columnconfigure(1, weight=1)
        options_frame.grid_columnconfigure(2, weight=1)
        
        # Opzione 1: Export CSV
        self._create_export_card(
            options_frame, 0, 0,
            "📊 Export CSV",
            "Esporta tutti i dati in formato CSV\nCompatibile con Excel e altri software",
            "Esporta CSV",
            self._export_csv,
            UIConfig.COLORS['success']
        )
        
        # Opzione 2: Export PDF (placeholder per future implementation)
        self._create_export_card(
            options_frame, 0, 1,
            "📑 Report PDF",
            "Genera un report professionale\ncon grafici e analisi dettagliate",
            "Genera PDF",
            self._export_pdf,
            UIConfig.COLORS['primary']
        )
        
        # Opzione 3: Backup Excel
        self._create_export_card(
            options_frame, 0, 2,
            "💾 Backup Excel",
            "Crea una copia di sicurezza\ndel database Excel completo",
            "Backup",
            self._backup_excel,
            UIConfig.COLORS['warning']
        )
        
        # Opzione 4: Export Asset Correnti
        self._create_export_card(
            options_frame, 1, 0,
            "🎯 Asset Attuali",
            "Esporta solo gli asset più recenti\n(senza storico duplicati)",
            "Export Attuali",
            self._export_current_assets,
            UIConfig.COLORS['info']
        )
        
        # Opzione 5: Export per Categoria
        self._create_export_card(
            options_frame, 1, 1,
            "🏷️ Per Categoria",
            "Esporta dati filtrati\nper una categoria specifica",
            "Export Categoria",
            self._export_by_category,
            UIConfig.COLORS['purple']
        )
        
        # Opzione 6: Import Dati (future feature)
        self._create_export_card(
            options_frame, 1, 2,
            "📥 Import Dati",
            "Importa dati da file CSV\n(Funzionalità futura)",
            "Import",
            self._import_data,
            UIConfig.COLORS['secondary'],
            enabled=False
        )
    
    def _create_export_card(self, parent, row: int, col: int, title: str, 
                           description: str, button_text: str, command: callable,
                           color: str, enabled: bool = True):
        """Crea una card per un'opzione di export"""
        card_frame = ctk.CTkFrame(parent, corner_radius=10)
        card_frame.grid(row=row, column=col, padx=15, pady=15, sticky="nsew", ipadx=20, ipady=20)
        
        # Titolo
        title_label = ctk.CTkLabel(
            card_frame,
            text=title,
            font=ctk.CTkFont(**UIConfig.FONTS['subheader'])
        )
        title_label.pack(pady=(10, 5))
        
        # Descrizione
        desc_label = ctk.CTkLabel(
            card_frame,
            text=description,
            font=ctk.CTkFont(**UIConfig.FONTS['text']),
            text_color=UIConfig.COLORS['secondary'],
            justify="center"
        )
        desc_label.pack(pady=(0, 15))
        
        # Bottone
        button = ctk.CTkButton(
            card_frame,
            text=button_text,
            command=command if enabled else None,
            **UIConfig.BUTTON_SIZES['medium'],
            font=ctk.CTkFont(**UIConfig.FONTS['button']),
            fg_color=color if enabled else UIConfig.COLORS['secondary'],
            hover_color=f"{color}_hover" if enabled else UIConfig.COLORS['secondary_hover'],
            state="normal" if enabled else "disabled"
        )
        button.pack(pady=(0, 10))
    
    def _create_info_section(self):
        """Crea la sezione informativa"""
        info_frame = ctk.CTkFrame(self.export_frame, height=100)
        info_frame.pack(fill="x", padx=10, pady=10)
        info_frame.pack_propagate(False)

        # Banner filtri attivi
        self._add_filter_banner(info_frame)

        # Statistiche portfolio
        self._update_portfolio_stats(info_frame)
    
    def _update_portfolio_stats(self, parent):
        """Aggiorna le statistiche del portfolio"""
        try:
            # Carica statistiche
            summary = self.portfolio_manager.get_portfolio_summary()
            current_assets_count = len(self.portfolio_manager.get_current_assets_only())
            total_records_count = len(self.portfolio_manager.load_data())
            
            # Layout orizzontale per le stats
            stats_container = ctk.CTkFrame(parent, fg_color="transparent")
            stats_container.pack(expand=True, fill="both", padx=20, pady=20)
            
            stats = [
                ("📊 Asset Attuali", f"{current_assets_count}"),
                ("📈 Record Totali", f"{total_records_count}"),
                ("💰 Valore Totale", f"€{summary.get('total_value', 0):,.2f}"),
                ("📅 Ultimo Export", self._get_last_export_time())
            ]
            
            for i, (label, value) in enumerate(stats):
                stat_frame = ctk.CTkFrame(stats_container)
                stat_frame.pack(side="left", expand=True, fill="both", padx=5)
                
                label_widget = ctk.CTkLabel(
                    stat_frame,
                    text=label,
                    font=ctk.CTkFont(**UIConfig.FONTS['text'])
                )
                label_widget.pack(pady=(10, 0))
                
                value_widget = ctk.CTkLabel(
                    stat_frame,
                    text=value,
                    font=ctk.CTkFont(**UIConfig.FONTS['subheader'])
                )
                value_widget.pack(pady=(0, 10))
                
        except Exception as e:
            print(f"Errore aggiornamento statistiche: {e}")

    def _format_filter_summary(self) -> str:
        """Ritorna una stringa riassuntiva dei filtri attivi.
        Formato: Campo: valore1, valore2, valore3
        Ogni filtro su una riga separata.
        Ritorna "Patrimonio Complessivo" se non ci sono filtri."""
        try:
            info = self._filter_info or {}
            col_filters = info.get('column_filters') or {}

            # Nessun filtro attivo - mostra "Patrimonio Complessivo"
            if not col_filters:
                return 'Patrimonio Complessivo'

            from config import FieldMapping
            lines = []

            for col, values in col_filters.items():
                disp = FieldMapping.DB_TO_DISPLAY.get(col, col)
                vals = list(sorted({str(v) for v in values}))
                shown = ', '.join(vals)
                lines.append(f"{disp}: {shown}")

            return '\n'.join(lines)
        except Exception:
            return 'Patrimonio Complessivo'

    def _add_filter_banner(self, parent):
        """Mostra un'etichetta con i filtri attivi se presenti."""
        summary = self._format_filter_summary()
        if not summary:
            return
        try:
            label = ctk.CTkLabel(
                parent,
                text=f"🔍 {summary}",
                font=ctk.CTkFont(**UIConfig.FONTS['text']),
                text_color=UIConfig.COLORS['secondary']
            )
            label.pack(fill="x", padx=10, pady=(6, 0))
        except Exception:
            pass

    def _build_export_header(self, df: pd.DataFrame, extra: str = '') -> str:
        """Costruisce le righe di intestazione per i file esportati (CSV)."""
        try:
            import os
            db_name = os.path.basename(self.portfolio_manager.excel_file)
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            summary = self._format_filter_summary()
            lines = [
                "# GAB AssetMind Export",
                f"# Database: {db_name}",
                f"# Data export: {timestamp}",
                f"# Selezione: {summary if summary else 'Nessun filtro (Asset correnti)'}",
                f"# Righe esportate: {len(df) if isinstance(df, pd.DataFrame) else 0}"
            ]
            if extra:
                lines.append(f"# {extra}")
            lines.append("")
            return "\n".join(lines)
        except Exception as e:
            return ""
    
    def _get_last_export_time(self) -> str:
        """Restituisce l'ora dell'ultimo export (placeholder)"""
        # In una versione futura, questo potrebbe essere salvato in un file di configurazione
        return "Mai"
    
    def _export_csv(self):
        """Esporta tutti i dati in formato CSV"""
        try:
            filename = filedialog.asksaveasfilename(
                title="Salva Export CSV",
                defaultextension=".csv",
                filetypes=[("File CSV", "*.csv"), ("Tutti i file", "*.*")],
                initialfile=f"portfolio_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            
            if filename:
                try:
                    # Valida il path di export per sicurezza
                    validated_path = self.path_validator.validate_export_path(filename)
                    self.logger.info(f"Path export CSV validato: {validated_path}")

                    if isinstance(self._external_filtered_df, pd.DataFrame) and not self._external_filtered_df.empty:
                        df = self._external_filtered_df.copy()
                    else:
                        df = self.portfolio_manager.load_data()
                    if df.empty:
                        messagebox.showwarning("Avviso", "Nessun dato da esportare")
                        return

                    # Intestazione + CSV (UTF-8 BOM, separatore ';')
                    header_text = self._build_export_header(df)
                    with open(str(validated_path), 'w', encoding='utf-8-sig', newline='') as f:
                        if header_text:
                            f.write(header_text)
                        df.to_csv(f, index=False, sep=';')
                    filename = str(validated_path)

                except SecurityError as e:
                    self.logger.error(f"Path export non sicuro: {e}")
                    messagebox.showerror("Errore Sicurezza", f"Path non sicuro: {e}")
                    return
                except Exception as e:
                    self.logger.error(f"Errore validazione path export: {e}")
                    messagebox.showerror("Errore", f"Errore validazione path: {e}")
                    return
                
                messagebox.showinfo("Successo", 
                                  f"Dati esportati con successo!\n\nFile: {os.path.basename(filename)}\n"
                                  f"Record esportati: {len(df)}\n"
                                  f"Percorso: {filename}")
                
                self.trigger_callback('export_completed', 'CSV', filename)
                
        except Exception as e:
            error_msg = ErrorHandler.handle_file_error(e, "export CSV")
            messagebox.showerror("Errore Export", error_msg)
    
    def _export_current_assets(self):
        """Esporta solo gli asset più recenti"""
        try:
            filename = filedialog.asksaveasfilename(
                title="Salva Asset Attuali",
                defaultextension=".csv",
                filetypes=[("File CSV", "*.csv"), ("Tutti i file", "*.*")],
                initialfile=f"asset_attuali_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            
            if filename:
                if isinstance(self._external_filtered_df, pd.DataFrame) and not self._external_filtered_df.empty:
                    df = self._external_filtered_df.copy()
                else:
                    df = self.portfolio_manager.get_current_assets_only()
                if df.empty:
                    messagebox.showwarning("Avviso", "Nessun asset attuale da esportare")
                    return
                header_text = self._build_export_header(df, extra='Vista: Asset Attuali')
                with open(filename, 'w', encoding='utf-8-sig', newline='') as f:
                    if header_text:
                        f.write(header_text)
                    df.to_csv(f, index=False, sep=';')
                
                messagebox.showinfo("Successo",
                                  f"Asset attuali esportati con successo!\n\n"
                                  f"File: {os.path.basename(filename)}\n"
                                  f"Asset esportati: {len(df)}")
                
                self.trigger_callback('export_completed', 'Current Assets', filename)
                
        except Exception as e:
            error_msg = ErrorHandler.handle_file_error(e, "export asset attuali")
            messagebox.showerror("Errore Export", error_msg)
    
    def _export_by_category(self):
        """Esporta dati filtrati per categoria"""
        try:
            # Dialog per selezione categoria
            from config import AssetConfig
            
            dialog = ctk.CTkInputDialog(
                text=f"Inserisci categoria da esportare:\n{', '.join(AssetConfig.CATEGORIES)}",
                title="Export per Categoria"
            )
            category = dialog.get_input()
            
            if not category:
                return
                
            if category not in AssetConfig.CATEGORIES:
                messagebox.showerror("Errore", f"Categoria '{category}' non valida")
                return
            
            filename = filedialog.asksaveasfilename(
                title=f"Salva Export {category}",
                defaultextension=".csv",
                filetypes=[("File CSV", "*.csv"), ("Tutti i file", "*.*")],
                initialfile=f"{category.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            
            if filename:
                if isinstance(self._external_filtered_df, pd.DataFrame) and not self._external_filtered_df.empty:
                    df = self._external_filtered_df.copy()
                else:
                    df = self.portfolio_manager.load_data()
                filtered_df = df[df['category'] == category]
                
                if filtered_df.empty:
                    messagebox.showwarning("Avviso", f"Nessun asset trovato per la categoria '{category}'")
                    return
                header_text = self._build_export_header(filtered_df, extra=f'Categoria: {category}')
                with open(filename, 'w', encoding='utf-8-sig', newline='') as f:
                    if header_text:
                        f.write(header_text)
                    filtered_df.to_csv(f, index=False, sep=';')
                
                messagebox.showinfo("Successo",
                                  f"Export categoria '{category}' completato!\n\n"
                                  f"File: {os.path.basename(filename)}\n"
                                  f"Asset esportati: {len(filtered_df)}")
                
                self.trigger_callback('export_completed', f'Category {category}', filename)
                
        except Exception as e:
            error_msg = ErrorHandler.handle_data_error(e, "export per categoria")
            messagebox.showerror("Errore Export", error_msg)
    
    def _backup_excel(self):
        """Crea un backup del database Excel"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = filedialog.asksaveasfilename(
                title="Salva Backup Excel",
                defaultextension=".xlsx",
                filetypes=[("File Excel", "*.xlsx"), ("Tutti i file", "*.*")],
                initialfile=f"portfolio_backup_{timestamp}.xlsx"
            )
            
            if filename:
                try:
                    # Valida il path di backup per sicurezza
                    validated_path = self.path_validator.validate_export_path(filename)
                    self.logger.info(f"Path backup validato: {validated_path}")

                    # IMPORTANTE: Il backup deve SEMPRE includere TUTTI i record (asset correnti + storici)
                    # Non usare _external_filtered_df che contiene solo gli asset filtrati
                    df = self.portfolio_manager.load_data()
                    if df.empty:
                        messagebox.showwarning("Avviso", "Nessun dato da fare backup")
                        return

                    # Salva usando il metodo del PortfolioManager per mantenere le formule
                    backup_manager = PortfolioManager(str(validated_path))
                    backup_manager.save_data(df)

                    # Applica la colorazione azzurra ai record storici nel backup
                    backup_manager.color_historical_records()

                    filename = str(validated_path)  # Usa path validato

                    # Aggiungi foglio 'Export_Info' con metadati
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(filename)
                        if 'Export_Info' in wb.sheetnames:
                            ws = wb['Export_Info']
                            wb.remove(ws)
                        ws = wb.create_sheet('Export_Info', 0)
                        import os
                        db_name = os.path.basename(self.portfolio_manager.excel_file)
                        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        summary = self._format_filter_summary() or 'Nessun filtro (Asset correnti)'
                        rows = [
                            ['GAB AssetMind - Export Info'],
                            ['Database', db_name],
                            ['Data export', timestamp],
                            ['Selezione', summary],
                            ['Righe esportate', len(df)]
                        ]
                        for r in rows:
                            ws.append(r)
                        wb.save(filename)
                    except Exception as e:
                        self.logger.debug(f'Impossibile aggiungere Export_Info: {e}')

                except SecurityError as e:
                    self.logger.error(f"Path backup non sicuro: {e}")
                    messagebox.showerror("Errore Sicurezza", f"Path non sicuro: {e}")
                    return
                except Exception as e:
                    self.logger.error(f"Errore validazione path backup: {e}")
                    messagebox.showerror("Errore", f"Errore validazione path: {e}")
                    return
                
                messagebox.showinfo("Successo",
                                  f"Backup creato con successo!\n\n"
                                  f"File: {os.path.basename(filename)}\n"
                                  f"Record inclusi: {len(df)}\n"
                                  f"Include formule Excel originali")
                
                self.trigger_callback('export_completed', 'Excel Backup', filename)
                
        except Exception as e:
            error_msg = ErrorHandler.handle_file_error(e, "backup Excel")
            messagebox.showerror("Errore Backup", error_msg)
    
    def _export_pdf(self):
        """Genera un report PDF completo sugli asset selezionati, includendo grafici."""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = filedialog.asksaveasfilename(
                title="Salva Report PDF",
                defaultextension=".pdf",
                filetypes=[("File PDF", "*.pdf"), ("Tutti i file", "*.*")],
                initialfile=f"report_{timestamp}.pdf"
            )

            if not filename:
                return

            # Valida path
            validated_path = self.path_validator.validate_export_path(filename)

            # Determina il DataFrame filtrato corrente
            filter_info = self._filter_info if isinstance(self._filter_info, dict) else {}
            col_filters = filter_info.get('column_filters') if filter_info else None
            show_all = bool(filter_info.get('show_all_records')) if filter_info else False
            has_column_filters = False
            if isinstance(col_filters, dict):
                has_column_filters = any(bool(values) for values in col_filters.values())
            elif col_filters:
                has_column_filters = True

            if isinstance(self._external_filtered_df, pd.DataFrame) and not self._external_filtered_df.empty:
                df = self._external_filtered_df.copy()
            else:
                try:
                    base_df = self.portfolio_manager.load_data() if show_all else self.portfolio_manager.get_current_assets_only()
                    df = apply_global_filters(base_df, col_filters)
                except Exception:
                    df = self.portfolio_manager.get_current_assets_only()

            if show_all and not has_column_filters:
                df = self.portfolio_manager.get_current_assets_only()

            if df.empty:
                messagebox.showwarning("Avviso", "Nessun dato da includere nel PDF")
                return

            # Metadati
            import os
            db_name = os.path.basename(self.portfolio_manager.excel_file)
            ts_human = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            selection_summary = self._format_filter_summary() or 'Nessun filtro (Asset correnti)'

            # Colonna valore corrente
            df = df.copy()
            df['current_value'] = df['updated_total_value'].fillna(df['created_total_value']).fillna(0)

            with PdfPages(str(validated_path)) as pdf:
                from math import ceil
                page_num = 1

                # Helper header/footer uniformi
                def _pdf_header_footer(fig, page_no: int):
                    try:
                        fig.text(0.02, 0.98, "GAB AssetMind – Report", fontsize=9, color='gray', va='top')
                        fig.text(0.98, 0.98, f"{db_name} | {ts_human}", fontsize=8, color='gray', ha='right', va='top')
                        fig.text(0.98, 0.02, f"Pagina {page_no}", ha='right', fontsize=8, color='gray')
                    except Exception:
                        pass

                # Precomputazioni per Sommario (pagine)
                try:
                    # Sezioni condizionali
                    pre_cat_series = df.groupby('category')['current_value'].sum()
                    pre_cat_series = pre_cat_series[pre_cat_series > 0]
                    has_cat = not pre_cat_series.empty

                    pre_risk_series = df.groupby('risk_level')['current_value'].sum() if 'risk_level' in df.columns else pd.Series(dtype=float)
                    has_risk = not pre_risk_series.empty

                    df_time = df.copy()
                    df_time['created_year'] = pd.to_datetime(df_time.get('created_at'), errors='coerce').dt.year
                    df_time['updated_year'] = pd.to_datetime(df_time.get('updated_at'), errors='coerce').dt.year
                    years = sorted(set([y for y in df_time['created_year'].dropna().astype(int).tolist()] +
                                       [y for y in df_time['updated_year'].dropna().astype(int).tolist()]))
                    has_timeline = len(years) > 0

                    # Dettaglio
                    detail_cols_pref = ['id', 'category', 'position', 'asset_name', 'isin',
                                        'created_at', 'updated_at', 'created_amount', 'updated_amount',
                                        'created_unit_price', 'updated_unit_price', 'current_value']
                    detail_cols = [c for c in detail_cols_pref if c in df.columns]
                    detail_df_pre = df[detail_cols].copy()
                    rows_per_page = 28
                    detail_pages = (len(detail_df_pre) + rows_per_page - 1) // rows_per_page if len(detail_df_pre) > 0 else 0

                    # Mappa sezioni -> pagine
                    cover_page = 1
                    summary_page = 2
                    kpi_page = 3
                    next_page = 4
                    cat_page = next_page if has_cat else None
                    next_page = next_page + 1 if has_cat else next_page
                    risk_page = next_page if has_risk else None
                    next_page = next_page + 1 if has_risk else next_page
                    timeline_page = next_page if has_timeline else None
                    next_page = next_page + 1 if has_timeline else next_page
                    top_page = next_page
                    next_page += 1
                    detail_first_page = next_page if detail_pages > 0 else None
                except Exception:
                    # Fallback semplice
                    cover_page = 1
                    summary_page = 2
                    kpi_page = 3
                    cat_page = 4
                    risk_page = 5
                    timeline_page = 6
                    top_page = 7
                    detail_first_page = 8

                # Copertina (Pagina 1)
                fig = plt.figure(figsize=(8.27, 11.69))
                fig.suptitle("GAB AssetMind - Report Portfolio", fontsize=20, fontweight='bold', y=0.92)
                # Logo opzionale
                try:
                    possible = [os.path.join(os.getcwd(), 'logo.png'), os.path.join(os.getcwd(), 'assets', 'logo.png')]
                    for p in possible:
                        if os.path.exists(p):
                            import matplotlib.image as mpimg
                            logo = mpimg.imread(p)
                            ax_logo = fig.add_axes([0.78, 0.88, 0.18, 0.1])
                            ax_logo.axis('off')
                            ax_logo.imshow(logo)
                            break
                except Exception:
                    pass
                cover_txt = (
                    f"Database: {db_name}\n"
                    f"Data export: {ts_human}\n"
                )
                fig.text(0.08, 0.80, cover_txt, fontsize=12)

                # Riquadro 'Note sulla selezione' con filtri attivi (elenco puntato)
                try:
                    from config import FieldMapping
                    note_x, note_y, note_w, note_h = 0.08, 0.58, 0.84, 0.22
                    box = FancyBboxPatch((note_x, note_y), note_w, note_h,
                                         boxstyle="round,pad=0.02,rounding_size=0.02",
                                         fc="#f9fafb", ec="#d1d5db", transform=fig.transFigure, clip_on=False)
                    fig.add_artist(box)

                    fig.text(note_x + 0.02, note_y + note_h - 0.05,
                             "Note sulla selezione", fontsize=13, fontweight='bold')

                    # Elenco puntato dei filtri
                    bullet_lines = []
                    base_line = f"Base: {'Tutti i record' if (self._filter_info or {}).get('show_all_records') else 'Asset correnti'}"
                    bullet_lines.append(base_line)
                    col_filters = (self._filter_info or {}).get('column_filters') or {}
                    if col_filters:
                        for col, values in col_filters.items():
                            disp = FieldMapping.DB_TO_DISPLAY.get(col, col)
                            vals = list(sorted({str(v) for v in values}))
                            shown = (', '.join(vals[:5]) + (f" +{len(vals)-5}" if len(vals) > 5 else '')) if vals else '—'
                            bullet_lines.append(f"• {disp}: {shown}")
                    else:
                        bullet_lines.append("• Nessun filtro di colonna attivo")

                    ty = note_y + note_h - 0.10
                    for line in bullet_lines:
                        fig.text(note_x + 0.02, ty, line, fontsize=10, color='#374151')
                        ty -= 0.05
                except Exception:
                    pass
                fig.text(0.08, 0.10, "Report generato da GAB AssetMind", fontsize=9, color='gray')
                # Niente header sulla copertina per evitare sovrapposizione
                pdf.savefig(fig)
                plt.close(fig)
                page_num += 1
                # Pagina KPI sintetici
                try:
                    total_value = df['current_value'].sum()
                    asset_count = len(df)
                    category_count = df['category'].nunique() if 'category' in df.columns else 0
                    # Rischio medio ponderato per valore se possibile, altrimenti semplice
                    if 'risk_level' in df.columns and asset_count > 0:
                        try:
                            weighted = (df['risk_level'].fillna(0) * df['current_value']).sum()
                            avg_risk = (weighted / total_value) if total_value > 0 else df['risk_level'].fillna(0).mean()
                        except Exception:
                            avg_risk = df['risk_level'].fillna(0).mean()
                    else:
                        avg_risk = 0

                    top_cat = None
                    try:
                        if 'category' in df.columns:
                            top_cat = df.groupby('category')['current_value'].sum().sort_values(ascending=False).head(1)
                    except Exception:
                        pass

                    fig, ax = plt.subplots(figsize=(11.69, 8.27))
                    ax.axis('off')
                    ax.set_title("Indicatori Sintetici (Selezione)", fontsize=16, fontweight='bold', pad=20)

                    # Disegna riquadri KPI
                    kpi_texts = [
                        ("Valore Totale", f"€{total_value:,.2f}"),
                        ("N. Asset", f"{asset_count}"),
                        ("N. Categorie", f"{category_count}"),
                        ("Rischio Medio", f"{avg_risk:.2f}"),
                        ("Top Categoria", f"{top_cat.index[0]} ( €{float(top_cat.iloc[0]):,.0f} )" if top_cat is not None and len(top_cat)>0 else "N/A"),
                    ]

                    x0, y0 = 0.05, 0.65
                    w, h = 0.27, 0.18
                    dx = 0.32
                    dy = 0.25
                    idx = 0
                    for row in range(2):
                        for col in range(3):
                            if idx >= len(kpi_texts):
                                break
                            label, value = kpi_texts[idx]
                            px = x0 + col*dx
                            py = y0 - row*dy
                            box = FancyBboxPatch((px, py), w, h, boxstyle="round,pad=0.02,rounding_size=0.02", fc="#f5f7fb", ec="#d1d5db")
                            ax.add_patch(box)
                            ax.text(px+0.02, py+h-0.06, label, fontsize=11, color="#6b7280")
                            ax.text(px+0.02, py+h/2-0.06, value, fontsize=16, fontweight='bold', color="#111827")
                            idx += 1

                    ax.text(0.01, 0.01, selection_summary, transform=ax.transAxes, fontsize=9, color='gray')
                    _pdf_header_footer(fig, page_num)
                    pdf.savefig(fig)
                    plt.close(fig)
                    page_num += 1
                except Exception:
                    pass
                # (Pagina ridondante rimossa)

                # (Pagina Sommario e Categoria rimosse su richiesta)

                # Distribuzione per rischio
                try:
                    risk_series = df.groupby('risk_level')['current_value'].sum().sort_index()
                    if not risk_series.empty:
                        fig, ax = plt.subplots(figsize=(11.69, 8.27))
                        ax.bar(risk_series.index.astype(str), risk_series.values,
                               color=plt.cm.Blues(np.linspace(0.4, 0.9, len(risk_series))))
                        ax.set_title("Distribuzione Valore per Livello di Rischio (selezione)", fontsize=14, fontweight='bold')
                        ax.set_xlabel("Livello di Rischio")
                        ax.set_ylabel("Valore (€)")
                        ax.grid(axis='y', alpha=0.3)
                        ax.text(0.01, 0.01, selection_summary, transform=ax.transAxes, fontsize=9, color='gray')
                        _pdf_header_footer(fig, page_num)
                        pdf.savefig(fig)
                        plt.close(fig)
                        page_num += 1
                except Exception:
                    pass

                # Evoluzione temporale (per categoria e totale)
                try:
                    # Parsing anni
                    df_time = df.copy()
                    df_time['created_year'] = pd.to_datetime(df_time.get('created_at'), errors='coerce').dt.year
                    df_time['updated_year'] = pd.to_datetime(df_time.get('updated_at'), errors='coerce').dt.year
                    # Valori
                    df_time['created_value'] = df_time.get('created_amount', 0).fillna(0) * df_time.get('created_unit_price', 0).fillna(0)
                    df_time['updated_value'] = df_time.get('updated_amount', 0).fillna(0) * df_time.get('updated_unit_price', 0).fillna(0)

                    years = sorted(set([y for y in df_time['created_year'].dropna().astype(int).tolist()] +
                                       [y for y in df_time['updated_year'].dropna().astype(int).tolist()]))
                    if years:
                        # Serie per categoria
                        categories = sorted([c for c in df_time['category'].dropna().unique()]) if 'category' in df_time.columns else []
                        series_by_cat = {}
                        for cat in categories:
                            sub = df_time[df_time['category'] == cat]
                            vals = []
                            for yr in years:
                                mc = sub['created_year'].notna() & (sub['created_year'] <= yr)
                                mu = sub['updated_year'].notna() & (sub['updated_year'] <= yr)
                                v = sub.loc[mu, 'updated_value'].sum() + sub.loc[mc & (~mu), 'created_value'].sum()
                                vals.append(v)
                            series_by_cat[cat] = vals

                        # Totale
                        total_values = [sum(series_by_cat[cat][i] for cat in series_by_cat) for i in range(len(years))] if series_by_cat else []

                        fig, ax = plt.subplots(figsize=(11.69, 8.27))
                        # Linee per categoria
                        if series_by_cat:
                            colors = plt.cm.Set3(np.linspace(0, 1, len(series_by_cat)))
                            for (cat, vals), col in zip(series_by_cat.items(), colors):
                                ax.plot(years, vals, marker='o', linewidth=1.8, label=str(cat), color=col)
                        # Linea totale
                        if total_values:
                            ax.plot(years, total_values, marker='s', color='black', linewidth=3.0, label='TOTALE')

                        ax.set_title("Evoluzione Temporale del Valore (selezione)", fontsize=14, fontweight='bold')
                        ax.set_xlabel("Anno")
                        ax.set_ylabel("Valore (€)")
                        ax.grid(True, alpha=0.3)
                        ax.legend(loc='upper left', bbox_to_anchor=(1.02, 1))
                        ax.text(0.01, 0.01, selection_summary, transform=ax.transAxes, fontsize=9, color='gray')
                        _pdf_header_footer(fig, page_num)
                        pdf.savefig(fig)
                        plt.close(fig)
                        page_num += 1
                except Exception:
                    pass

                # Tabella top asset
                try:
                    cols = ['id', 'category', 'asset_name', 'position', 'isin', 'current_value']
                    show_cols = [c for c in cols if c in df.columns]
                    table_df = df[show_cols].copy().sort_values('current_value', ascending=False)
                    try:
                        total_value_sel = df['current_value'].sum()
                        table_df['Share %'] = table_df['current_value'].apply(lambda v: (v/total_value_sel*100) if total_value_sel>0 else 0).round(2)
                    except Exception:
                        table_df['Share %'] = 0.0
                    top_n = min(30, len(table_df))
                    table_df = table_df.head(top_n)
                    fig, ax = plt.subplots(figsize=(11.69, 8.27))
                    ax.axis('off')
                    ax.set_title("Top Asset per Valore (selezione)", fontsize=14, fontweight='bold', pad=20)
                    tbl = ax.table(cellText=table_df.values,
                                   colLabels=[str(c).title().replace('_', ' ') for c in table_df.columns],
                                   loc='center')
                    tbl.auto_set_font_size(False)
                    tbl.set_fontsize(8)
                    tbl.scale(1, 1.4)
                    _pdf_header_footer(fig, page_num)
                    pdf.savefig(fig)
                    plt.close(fig)
                    page_num += 1
                except Exception:
                    pass

                # Dettaglio Asset: tabella impaginata (tutti gli asset selezionati) - A3 Landscape
                try:
                    from config import FieldMapping

                    # Colonne fedeli alla tabella Portfolio (stesso ordine)
                    detail_cols_pref = ['id', 'category', 'position', 'asset_name', 'isin', 'ticker', 'risk_level',
                                        'created_at', 'created_amount', 'created_unit_price', 'created_total_value',
                                        'updated_at', 'updated_amount', 'updated_unit_price', 'updated_total_value',
                                        'accumulation_plan', 'accumulation_amount', 'income_per_year', 'rental_income', 'note']
                    detail_cols = [c for c in detail_cols_pref if c in df.columns]
                    detail_df = df[detail_cols].copy()

                    # Helper per word wrap
                    def wrap_text(text, max_len=30):
                        """Spezza testo lungo su più righe"""
                        if pd.isna(text) or text == '':
                            return ''
                        text = str(text)
                        if len(text) <= max_len:
                            return text
                        # Spezza ogni max_len caratteri
                        lines = []
                        for i in range(0, len(text), max_len):
                            lines.append(text[i:i+max_len])
                        return '\n'.join(lines)

                    # Formattazioni base
                    for col in detail_df.columns:
                        if col in ['created_total_value', 'updated_total_value']:
                            detail_df[col] = detail_df[col].map(lambda x: f"€{x:,.2f}" if pd.notna(x) else '')
                        elif col in ['created_amount', 'updated_amount']:
                            detail_df[col] = detail_df[col].map(lambda x: f"{x:,.4f}" if pd.notna(x) else '')
                        elif col in ['created_unit_price', 'updated_unit_price', 'accumulation_amount', 'income_per_year', 'rental_income']:
                            detail_df[col] = detail_df[col].map(lambda x: f"€{x:,.2f}" if pd.notna(x) else '')
                        elif col in ['created_at', 'updated_at']:
                            detail_df[col] = detail_df[col].map(lambda x: str(x)[:10] if pd.notna(x) else '')
                        elif col == 'asset_name':
                            detail_df[col] = detail_df[col].map(lambda x: wrap_text(x, 40))
                        elif col == 'note':
                            detail_df[col] = detail_df[col].map(lambda x: wrap_text(x, 50))
                        else:
                            detail_df[col] = detail_df[col].map(lambda x: str(x) if pd.notna(x) else '')

                    # Headers con nomi display
                    col_labels = [FieldMapping.DB_TO_DISPLAY.get(c, c.replace('_', ' ').title()) for c in detail_df.columns]

                    rows_per_page = 22  # Meno righe per A3 landscape per lasciare spazio al word wrap
                    total_rows = len(detail_df)
                    for start in range(0, total_rows, rows_per_page):
                        chunk = detail_df.iloc[start:start+rows_per_page]

                        # A3 Landscape: 16.54 x 11.69 inches
                        fig, ax = plt.subplots(figsize=(16.54, 11.69))
                        ax.axis('off')
                        ax.set_title("Dettaglio Asset (selezione)", fontsize=14, fontweight='bold', pad=20)

                        tbl = ax.table(cellText=chunk.values,
                                       colLabels=col_labels,
                                       loc='center',
                                       cellLoc='left')  # Allineamento a sinistra per leggibilità
                        tbl.auto_set_font_size(False)
                        tbl.set_fontsize(6)  # Font più piccolo per più colonne
                        tbl.scale(1, 1.8)  # Più altezza per il word wrap

                        # Stile header
                        for (i, j), cell in tbl.get_celld().items():
                            if i == 0:  # Header
                                cell.set_facecolor('#4472C4')
                                cell.set_text_props(weight='bold', color='white')
                            else:
                                if i % 2 == 0:
                                    cell.set_facecolor('#F2F2F2')
                                else:
                                    cell.set_facecolor('white')

                        _pdf_header_footer(fig, page_num)
                        pdf.savefig(fig, dpi=150)  # DPI più alta per leggibilità
                        plt.close(fig)
                        page_num += 1
                except Exception as e:
                    self.logger.error(f"Errore generazione tabella dettaglio: {e}")
                    pass

            messagebox.showinfo("Report PDF", f"Report generato: {os.path.basename(str(validated_path))}")
            self.trigger_callback('export_completed', 'PDF', str(validated_path))

        except SecurityError as e:
            messagebox.showerror("Errore Sicurezza", f"Path non sicuro: {e}")
        except Exception as e:
            error_msg = ErrorHandler.handle_file_error(e, "export PDF")
            messagebox.showerror("Errore Export PDF", error_msg)

    def _import_data(self):
        """Importa dati da file esterno (funzionalità futura)"""
        messagebox.showinfo(
            "Funzionalità Futura",
            "L'importazione di dati da file esterni sarà implementata in una versione futura.\n\n"
            "Supporterà:\n"
            "• Import da file CSV con mapping automatico\n"
            "• Validazione dati durante l'import\n"
            "• Preview e conferma prima dell'inserimento\n"
            "• Merge intelligente con dati esistenti"
        )
    
    def refresh_stats(self):
        """Aggiorna le statistiche visualizzate"""
        # Rimuove e ricrea la sezione info
        for widget in self.export_frame.winfo_children():
            if isinstance(widget, ctk.CTkFrame):
                children = widget.winfo_children()
                if len(children) > 0 and any("Asset Attuali" in str(child) for child in children if hasattr(child, 'cget')):
                    widget.destroy()
                    self._create_info_section()
                    break

    def _update_filter_label(self):
        """Aggiorna la label dei filtri nell'header."""
        if not hasattr(self, 'filter_label') or not self.filter_label:
            return

        summary = self._format_filter_summary()
        if summary:
            self.filter_label.configure(text=summary)
        else:
            self.filter_label.configure(text="")

    def refresh_with_filtered_data(self, df: pd.DataFrame, filter_info: Optional[dict] = None):
        """Imposta DF filtrato e info filtri correnti, poi aggiorna stats."""
        try:
            self._external_filtered_df = df.copy() if isinstance(df, pd.DataFrame) else None
            if isinstance(filter_info, dict):
                self._filter_info = filter_info
            self._update_filter_label()
        except Exception:
            self._external_filtered_df = None
        finally:
            safe_execute(self.refresh_stats)








